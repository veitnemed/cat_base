import json
import os
import sys
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter import simpledialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *

import tkinter as tk
from tkinter import messagebox
import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog  # уже должно быть
import traceback



# --- Tooltip infra (минимальная) ---
if "_bind_tooltip" not in globals():
    class _HoverTip:


        def __init__(self, widget, text_provider):
            self.widget = widget
            self.text_provider = text_provider
            self.tip = None
            widget.bind("<Enter>", self._show, add="+")
            widget.bind("<Leave>", self._hide, add="+")
            widget.bind("<Motion>", self._move, add="+")
        def _show(self, e=None):
            try:
                txt = self.text_provider() or ""
            except Exception:
                txt = ""
            if not txt:
                return
            if self.tip:
                return
            import tkinter as tk
            self.tip = tk.Toplevel(self.widget)
            self.tip.wm_overrideredirect(True)
            self.tip.wm_attributes("-topmost", True)
            lbl = tk.Label(self.tip, text=txt, justify="left",
                           background="#ffffe0", relief="solid", borderwidth=1,
                           font=("Segoe UI", 9))
            lbl.pack(ipadx=4, ipady=2)
            self._move(e)
        def _move(self, e=None):
            if not self.tip:
                return
            x = (e.x_root if e else self.widget.winfo_rootx()) + 12
            y = (e.y_root if e else self.widget.winfo_rooty()) + 12
            self.tip.wm_geometry(f"+{x}+{y}")
        def _hide(self, e=None):
            if self.tip:
                try:
                    self.tip.destroy()
                except Exception:
                    pass
                self.tip = None

    def _bind_tooltip(widget, text_provider):
        try:
            _HoverTip(widget, text_provider)
        except Exception:
            pass

# --- Специальный тултип "112: xxxx" для ВС7 ---
def attach_storage_112_tooltip(widget, key):
    def _text():
        return f"112: {_vs7_112_last4_for_storage(key)}"
    _bind_tooltip(widget, _text)


def _center_window(win, parent=None):
    try:
        win.update_idletasks()
        p = parent if parent else root
        px, py = p.winfo_rootx(), p.winfo_rooty()
        pw, ph = p.winfo_width(), p.winfo_height()
        ww, wh = win.winfo_width(), win.winfo_height()
        if not ww or not wh:
            win.update_idletasks()
            ww, wh = win.winfo_width(), win.winfo_height()
        x = px + max(0, (pw - ww)//2)
        y = py + max(0, (ph - wh)//2)
        win.geometry(f"+{x}+{y}")
    except Exception:
        win.geometry("+80+80")

def _ask_vs7_112_last4(key, prefill=""):
    # ---------- ручные настройки ----------
    WIN_W, WIN_H     = 380, 230      # размеры окна (пиксели)
    PADX, PADY       = 18, 14        # внешние отступы контейнеров
    CELL_WIDTH_CH    = 1            # ширина ячейки в символах
    CELL_IPADY       = 2             # «высота» ячейки (внутренний отступ)
    CELL_H_GAP       = 4            # расстояние между ячейками
    TOP_LINE_FONT    = ("Segoe UI", 12)          # «ВС7 №xxxx [z]»
    TITLE_FONT       = ("Segoe UI", 18, "bold")  # «Номер ССБ 112:»
    DIGIT_FONT       = ("Segoe UI", 14, "bold")  # цифры в ячейках
    # --------------------------------------

    top = tk.Toplevel(root)
    top.title("Номер ССБ 112")
    top.transient(root)
    top.grab_set()
    top.geometry(f"{WIN_W}x{WIN_H}")
    top.resizable(False, False)

    # центрирование
    try:
        top.update_idletasks()
        sw = top.winfo_screenwidth()
        sh = top.winfo_screenheight()
        x = (sw - WIN_W) // 2
        y = (sh - WIN_H) // 2
        top.geometry(f"{WIN_W}x{WIN_H}+{x}+{y}")
    except Exception:
        pass

    try:
        _, z, num4 = serial_parts_from_key(key)
    except Exception:
        z, num4 = "?", "????"
    tb.Label(top, text=f"ВС7 №{num4}",
             font=TOP_LINE_FONT, anchor="center").pack(fill="x", padx=PADX, pady=(PADY, 6))

    # заголовок по центру
    tb.Label(top, text="Номер ССБ 112:", font=("Segoe UI", 12, "bold"),
             anchor="center").pack(fill="x", padx=18, pady=(0, 8))

    # ряд ячеек
    row = tb.Frame(top);
    row.pack(padx=20, pady=(0, 20))
    for i in range(4):
        row.grid_columnconfigure(i, weight=1, uniform="cells")

    prefill = prefill if (isinstance(prefill, str) and prefill.isdigit() and len(prefill) == 4) else ""
    vars4, entries = [], []

    def focus_next(i):
        if i < 3:
            entries[i + 1].focus_set();
            entries[i + 1].icursor("end")

    def focus_prev(i):
        if i > 0:
            entries[i - 1].focus_set();
            entries[i - 1].icursor("end")

    # ВАЖНО: реагируем только на цифры на KeyPress
    def on_keypress(ev, idx):
        ch = ev.char or ""
        if ch.isdigit() and len(ch) == 1:
            vars4[idx].set(ch)
            focus_next(idx)
            return "break"  # не даём «отпрыгивать»
        # Игнор всего остального (стрелки, шифт, и т.д.) — ничего не делаем
        return None

    def on_backspace(ev, idx):
        if vars4[idx].get():
            vars4[idx].set("")
        else:
            if idx > 0:
                focus_prev(idx)
                vars4[idx - 1].set("")
        return "break"

    def on_delete(ev, idx):
        vars4[idx].set("")
        return "break"

    def on_left(ev, idx):
        focus_prev(idx);  return "break"

    def on_right(ev, idx):
        focus_next(idx);  return "break"

    def on_paste(ev, idx):
        try:
            text = top.clipboard_get()
        except Exception:
            return "break"
        digs = "".join(ch for ch in text if ch.isdigit())[:4]
        if not digs:
            return "break"
        i = idx
        for ch in digs:
            if i > 3: break
            vars4[i].set(ch);
            i += 1
        entries[min(i, 3)].focus_set();
        entries[min(i, 3)].icursor("end")
        return "break"

    for i in range(4):
        sv = tk.StringVar(value=(prefill[i] if i < len(prefill) else ""))
        e = tb.Entry(row, width=3, justify="center", textvariable=sv)
        try:
            e.configure(font=("Segoe UI", 12, "bold"))
        except:
            pass
        e.grid(row=0, column=i, padx=(6 if i > 0 else 0, 6 if i < 3 else 0),
               ipady=6, sticky="nsew")

        # БИНДИМ KeyPress вместо KeyRelease
        e.bind("<KeyPress>", lambda ev, j=i: on_keypress(ev, j))
        e.bind("<BackSpace>", lambda ev, j=i: on_backspace(ev, j))
        e.bind("<Delete>", lambda ev, j=i: on_delete(ev, j))
        e.bind("<Left>", lambda ev, j=i: on_left(ev, j))
        e.bind("<Right>", lambda ev, j=i: on_right(ev, j))
        e.bind("<<Paste>>", lambda ev, j=i: on_paste(ev, j))
        e.bind("<Control-v>", lambda ev, j=i: on_paste(ev, j))
        vars4.append(sv);
        entries.append(e)

    if entries:
        entries[0].focus_set();
        entries[0].icursor("end")

    # Кнопки + глобальные бинды Enter/Esc
    btns = tb.Frame(top);
    btns.pack(pady=(0, 14))
    result = {"val": None}

    def _collect():
        return "".join(v.get() for v in vars4)

    def on_ok():
        s = _collect()
        if len(s) != 4 or not s.isdigit():
            messagebox.showwarning("ССБ 112", "Введите ровно 4 цифры или нажмите «Отмена».", parent=top)
            return
        result["val"] = s
        top.destroy()

    def on_cancel():
        result["val"] = None
        top.destroy()

    tb.Button(btns, text="ОК", width=12, bootstyle="primary", command=on_ok).pack(side="left", padx=8)
    tb.Button(btns, text="Отмена", width=12, bootstyle="secondary", command=on_cancel).pack(side="left", padx=8)

    # Enter / NumPad Enter подтверждают, Esc — отмена
    top.bind("<Return>", lambda e: on_ok())
    top.bind("<KP_Enter>", lambda e: on_ok())
    top.bind("<Escape>", lambda e: on_cancel())

    top.wait_window()
    return result["val"]


def _vs7_112_last4_for_storage(key) -> str:
    """
    Возвращает последние 4 цифры 112 для ВС7 в хранилище:
    - если meta['from_vs6d'] == True -> берём xxxx из номера СЕРПа;
    - иначе берём meta['last4'] (если валидны 4 цифры);
    - иначе '????'.
    """
    try:
        # поддержка как tuple-ключей, так и строковых (на всякий случай)
        meta = (vs7_112_meta.get(key)
                or vs7_112_meta.get("|".join(map(str, key)), {})) or {}
        if meta.get("from_vs6d") is True:
            return serial_parts_from_key(key)[2]  # xxxx
        last4 = str(meta.get("last4", "")).strip()
        return last4 if (last4.isdigit() and len(last4) == 4) else "????"
    except Exception:
        return "????"

enable_vs13_mode = True
product_mode = "ВС6Д"
vs6d_btn = None
vs13_btn = None
vs6d_only_mode = False
vs6d_only_mode_var = None
storage_sort_mode = "dates"  # "dates" | "models"
storage_sort_button = None
work_sort_mode = "factories"
sort_buttons = {}
toggle_storage_button = None
GAP_AFTER_NUMBER_PX = 12

entry_header_label = None
BG_MAIN = BG_WORK_AREA = BG_HEADER = BG_COUNTER = "#ededed"
BG_ENTRY, BG_SETTINGS = "#ffffff", "#ededed"
TEXT_MAIN = TEXT_LABEL = TEXT_HEADER = TEXT_FACTORY = "#212529"
TEXT_BUTTON, TEXT_DATE, TEXT_COUNTER = "#ffffff", "#555555", "#25405c"
XL_MAX_ROWS_LIMIT = 1500          # верхний предел в настройке
xl_max_rows = 120                 # по умолчанию (можешь поставить 300/500)
xl_settings_win = None
BORDER_MAIN = BORDER_INPUT = BORDER_WORK_AREA = "#212529"
AUTO_COMPACT_ON_SMALL_WIDTH = False
BUTTON_SERP = BUTTON_SAVE = BUTTON_SEARCH = "#505b65"
BUTTON_SUCCESS, BUTTON_DANGER = "#dde6dd", "#d5c0c0"
SECTION_BORDER = "#cfd4da"
BUTTON_SECONDARY, BUTTON_DRAFT, BUTTON_NOT_READY = "#6c757d", "#ce9713", "#972f2f"
FONT_SIZE_SMALL, FONT_SIZE_MEDIUM, FONT_SIZE_LARGE = 10, 11, 12
FONT_SIZE_XL = FONT_SIZE_XXL = 12
FONT_PRIMARY = ("Arial", FONT_SIZE_MEDIUM)
FONT_BOLD = ("Arial", FONT_SIZE_MEDIUM, "bold")
FONT_COUNTER, FONT_BLOCK = ("Arial", 12), ("Arial", 12)
FONT_SERP_NUMBER = FONT_SAVE_BUTTON = ("Arial", 12)
FONT_HEADER1 = FONT_HEADER2 = ("Arial", FONT_SIZE_XL, "bold")
FONT_HEADER3, FONT_FACTORY = ("Arial", 11, "italic"), ("Arial", FONT_SIZE_LARGE, "bold")
FONT_SETTINGS, FONT_BUTTON = ("Arial", FONT_SIZE_XL, "bold"), (
    "Arial",
    FONT_SIZE_MEDIUM,
)
vs7_112_meta = {}
variant_overrides = {}
# === НАСТРОЙКИ АРХИВА ===
ARCHIVE_FONT = ("Arial", 12)          # строки списка
ARCHIVE_HEADER_FONT = ("Arial", 13, "bold")  # заголовки дат/вкладок
ARCHIVE_ROW_PADY = 6                             # отступ между строками
FONT_STATUS_BUTTON = ("Arial", 11, "bold")
blockcount_mode = False

display_full_serials = False
display_full_serials_var = None
display_full_serials_storage = True

storage_count_lbl = None
request_table = None
filter_var = None
work_header_label = None
storage_header_label = None
manual_override_ping = False
last_window_state = None
work6_widgets = {}
work7_widgets = {}
work_scroll_position = 0.0
work7_scroll_position = 0.0
work_top_element = None
work7_top_element = None
storage_visible = False
storage_count_cache = 0
request_filter = "Все"
TEXT_ADD_SERP, TEXT_SAVE, TEXT_SEARCH = "+", "Сохранить", "Поиск"
TEXT_WORK_TITLE, TEXT_STORAGE_TITLE = "СЕРП'ы в работе", "Ждут отправки"
TEXT_WORK_LABEL, TEXT_STORAGE_LABEL = "В работе", "Собранно"
TEXT_YEAR, TEXT_NUMBER, TEXT_SEARCH_LABEL = "Год:", "Номер:", "Поиск:"
storage_count_cache = 0
BORDER_WIDTH, ENTRY_BORDER_WIDTH = 10, 10
ROW_HEIGHT, PADY_HEADER, PADY_PRODUCT = 40, 5, 3
repair_blocks = set()
REPAIR_TEXT = "РЕМОНТ"
hide_added_date = False
REPAIR_BOOTSTYLE = "secondary"
base_path = (
    os.path.dirname(sys.executable)
    if getattr(sys, "frozen", False)
    else os.path.dirname(os.path.abspath(__file__))
)
# --- ФИКСИРОВАННАЯ БАЗОВАЯ ПАПКА ДЛЯ ВСЕХ ДАННЫХ И ОТЧЁТА ---
# --- БАЗОВЫЕ ПАПКИ ---
SERP_BASE_DIR = r"C:\serp_base"
os.makedirs(SERP_BASE_DIR, exist_ok=True)

# Новая схема файлов
BASE_DATA_FILE    = os.path.join(SERP_BASE_DIR, "base_data.json")
ARCHIVE_FILE      = os.path.join(SERP_BASE_DIR, "base_archive.json")

# Бэкапы в отдельной папке
BACKUP_DIR = r"C:\base_basckup"
os.makedirs(BACKUP_DIR, exist_ok=True)
BACKUP_DATA_FILE    = os.path.join(BACKUP_DIR, "base_data_backup.json")
BACKUP_ARCHIVE_FILE = os.path.join(BACKUP_DIR, "base_archive_backup.json")

# ===================== XL-подсветка (112/114/161) + НОРМАЛИЗАТОР =====================
# Папка: C:\serp_base\xl
# Файлы:
#   - xl.xlsx           (входной, может быть "особый")
#   - xl_norm.xlsx      (нормализованная копия, создаётся автоматически)
#   - top_ssb.jason     (кэш подсказок для подсветки)

XL_DIR = r"C:\serp_base\xl"
XL_XLSX_PATH = os.path.join(XL_DIR, "xl.xlsx")
XL_NORM_XLSX_PATH = os.path.join(XL_DIR, "xl_norm.xlsx")
XL_CACHE_DIR = os.path.join(SERP_BASE_DIR, "cache_xl")
XL_JSON_PATH = os.path.join(XL_CACHE_DIR, "top_ssb.jason")

XL_DEBUG_LOG = os.path.join(XL_DIR, "xl_debug.txt")

# ВРЕМЕННАЯ ОТЛАДКА:
# True  -> пишем подробный лог в XL_DEBUG_LOG
# False -> лог не пишется (перед релизом поставишь False)
XL_DEBUG = True
xl_pass = {"ССБ 112": set(), "ССБ 114": set(), "ССБ 161": set()}
xl_fail = {"ССБ 112": set(), "ССБ 114": set(), "ССБ 161": set()}

XL_TARGET_SHEETS = {
    "112": ["ssb112", "ssb 112", "SSB112", "SSB 112"],
    "114": ["ssb114", "ssb 114", "SSB114", "SSB 114"],
    "161": ["ssb161", "ssb 161", "SSB161", "SSB 161"],
}

XL_MAX_KEEP_PER_TYPE = 120   # берём чуть больше "≈100", как просили
XL_SCAN_ROW_LIMIT = 5000     # страховка, чтобы не читать бесконечно

# Глобальные наборы подсказок (по 4-ключу: (num4, factory_name, "ВС6Д", yy))
xl_hints_112 = set()
xl_hints_114 = set()
xl_hints_161 = set()


def _xl_log(msg: str):
    # ЛОГИ XL ОТКЛЮЧЕНЫ: ничего не пишем на диск
    return


def _write_json_atomic(path: str, obj) -> bool:
    """Запись JSON через temp-файл + replace (меньше шансов словить битый файл)."""
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
        return True
    except Exception as e:
        _xl_log(f"write_json_atomic ERROR: {e}\n{traceback.format_exc()}")
        return False


def _pick_sheet_name(wb, variants):
    """Выбрать реальное имя листа из вариантов (без учёта регистра)."""
    try:
        names = list(getattr(wb, "sheetnames", []) or [])
        low = {str(n).strip().casefold(): n for n in names}
        for v in variants:
            k = str(v).strip().casefold()
            if k in low:
                return low[k]
    except Exception:
        pass
    return None


def _is_ws_looks_empty(ws) -> bool:
    """
    Главный детектор твоей проблемы:
    если openpyxl видит только A1, то max_row/max_col = 1.
    """
    try:
        mr = int(getattr(ws, "max_row", 0) or 0)
        mc = int(getattr(ws, "max_column", 0) or 0)
        if mr <= 1 and mc <= 1:
            return True
        return False
    except Exception:
        return True


def _excel_make_normalized_copy(src_path: str, dst_path: str, sheet_names_to_copy) -> bool:
    """
    Нормализация через Excel COM:
    - открываем src в Excel
    - создаём новый workbook
    - на каждый нужный лист переносим USED RANGE как VALUES
    - сохраняем dst как обычный .xlsx

    Это повторяет твою "ручную копию", из-за которой всё заработало.
    """
    _xl_log("=== NORMALIZE VIA EXCEL START ===")
    _xl_log(f"src={src_path}")
    _xl_log(f"dst={dst_path}")

    try:
        import win32com.client  # нужен pywin32
    except Exception as e:
        _xl_log(f"pywin32 not installed / import failed: {e}")
        return False

    excel = None
    wb = None
    out_wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        # ReadOnly=True, чтобы не ругалось, если файл держит другая программа
        wb = excel.Workbooks.Open(src_path, ReadOnly=True)

        # если есть внешние обновления — пробуем подтянуть
        try:
            wb.RefreshAll()
            # если есть async-запросы — дождаться (если метода нет — просто пропускаем)
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
        except Exception:
            pass

        out_wb = excel.Workbooks.Add()

        # удалим лишние дефолтные листы (в Excel их обычно 1-3)
        try:
            while out_wb.Worksheets.Count > 0:
                ws0 = out_wb.Worksheets(1)
                ws0.Delete()
        except Exception:
            pass

        copied_any = False

        for sname in sheet_names_to_copy:
            try:
                src_ws = wb.Worksheets(sname)
            except Exception:
                _xl_log(f"sheet not found in Excel: {sname}")
                continue

            try:
                used = src_ws.UsedRange
                rows = used.Rows.Count
                cols = used.Columns.Count
                _xl_log(f"[{sname}] UsedRange rows={rows} cols={cols}")

                # создаём лист в выходной книге
                new_ws = out_wb.Worksheets.Add()
                new_ws.Name = str(sname)

                # читаем значения пачкой и пишем пачкой
                values = used.Value
                if values is None:
                    _xl_log(f"[{sname}] used.Value is None -> skip")
                    continue

                # диапазон назначения
                new_ws.Range(new_ws.Cells(1, 1), new_ws.Cells(rows, cols)).Value = values
                copied_any = True
            except Exception as e:
                _xl_log(f"[{sname}] copy ERROR: {e}\n{traceback.format_exc()}")

        if not copied_any:
            _xl_log("normalize: nothing copied -> FAIL")
            return False

        os.makedirs(os.path.dirname(dst_path), exist_ok=True)
        # 51 = xlOpenXMLWorkbook (.xlsx)
        out_wb.SaveAs(dst_path, FileFormat=51)
        _xl_log("normalize: saved OK")
        return True

    except Exception as e:
        _xl_log(f"NORMALIZE ERROR: {e}\n{traceback.format_exc()}")
        return False

    finally:
        try:
            if out_wb is not None:
                out_wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        _xl_log("=== NORMALIZE VIA EXCEL END ===")


def _parse_block_serial(raw) -> tuple:
    """
    raw -> ('aa','yy','z','xxxx') или (None,...)
    Ожидаем 9 цифр: aa + yy + z + xxxx
    Примеры:
      612520430 -> aa=61 yy=25 z=2 xxxx=0430
      122520432 -> aa=12 yy=25 z=2 xxxx=0432
    """
    try:
        s = str(raw).strip()
        # часто Excel даёт float/инт — приведём:
        if s.endswith(".0"):
            s = s[:-2]
        s = "".join(ch for ch in s if ch.isdigit())
        if len(s) != 9:
            return (None, None, None, None)
        aa, yy, z, num4 = s[0:2], s[2:4], s[4:5], s[5:9]
        if aa not in ("12", "14", "61"):
            return (None, None, None, None)
        if not (yy.isdigit() and z.isdigit() and num4.isdigit()):
            return (None, None, None, None)
        return (aa, yy, z, num4)
    except Exception:
        return (None, None, None, None)


def _block_serial_to_key(raw_serial):
    """
    9-значный серийник блока -> 4-ключ СЕРПа:
      (num4, factory_name, "ВС6Д", yy)
    """
    aa, yy, z, num4 = _parse_block_serial(raw_serial)
    if not aa:
        return None

    # z — это код завода (1..6) -> имя завода
    try:
        factory_name = factory_mapping.get(z)
    except Exception:
        factory_name = None
    if not factory_name:
        return None

    return (num4, factory_name, "ВС6Д", yy)


def _collect_verified_from_ws(ws, want_prefix: str):
    """
    Чтение из openpyxl Worksheet.
    Берём колонки:
      D = Block serial number
      E = Result
    """
    collected = []
    scanned = 0
    parsed_fail = 0
    result1_seen = 0
    sample = []

    try:
        max_row = int(getattr(ws, "max_row", 0) or 0)
        max_col = int(getattr(ws, "max_column", 0) or 0)
    except Exception:
        max_row, max_col = 0, 0

    _xl_log(f"[{ws.title}] max_row={max_row} max_col={max_col}")

    if max_row < 2:
        return collected

    # D=4, E=5
    for r in range(2, min(max_row, XL_SCAN_ROW_LIMIT) + 1):
        scanned += 1
        serial = ws.cell(row=r, column=4).value
        res = ws.cell(row=r, column=5).value

        aa, yy, z, num4 = _parse_block_serial(serial)
        if not aa:
            parsed_fail += 1
            continue

        # фильтр по типу листа (на всякий)
        if aa != want_prefix:
            continue

        # Result должен быть 1
        ok = False
        try:
            if isinstance(res, str):
                ok = res.strip() == "1"
            else:
                ok = int(res) == 1
        except Exception:
            ok = False

        if ok:
            result1_seen += 1
            collected.append(f"{aa}{yy}{z}{num4}")
            if len(sample) < 5:
                sample.append(f"{aa}{yy}{z}{num4}")

        if len(collected) >= XL_MAX_KEEP_PER_TYPE:
            break

    _xl_log(f"[{ws.title}] scanned={scanned} parsed_fail={parsed_fail} result1_seen={result1_seen} collected={len(collected)} sample={sample}")
    return collected


def _load_xl_cache(silent=True):
    """Читаем top_ssb.jason при старте, чтобы подсветка появлялась сразу."""
    global xl_hints_112, xl_hints_114, xl_hints_161
    xl_hints_112, xl_hints_114, xl_hints_161 = set(), set(), set()

    try:
        if not os.path.isfile(XL_JSON_PATH):
            return
        with open(XL_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f) or {}

        def _read_list(name):
            out = set()
            for item in data.get(name, []) or []:
                # ожидаем список вида ["0432","ВЕКТОР","ВС6Д","25"]
                if isinstance(item, (list, tuple)) and len(item) >= 4:
                    out.add((str(item[0]).zfill(4), str(item[1]), "ВС6Д", str(item[3])[:2].zfill(2)))
            return out

        xl_hints_112 = _read_list("112")
        xl_hints_114 = _read_list("114")
        xl_hints_161 = _read_list("161")

    except Exception as e:
        _xl_log(f"_load_xl_cache ERROR: {e}\n{traceback.format_exc()}")
        if not silent:
            messagebox.showerror("XL", f"Не удалось прочитать {XL_JSON_PATH}:\n{e}")


def _show_xl_updated_info(event=None):
    """
    ПКМ по кнопке 📄: показать, когда обновлялся xl.xlsx (и кэш top_ssb.jason, если есть).
    """
    try:
        lines = []

        # 1) Время обновления Excel-файла
        if os.path.isfile(XL_XLSX_PATH):
            ts = os.path.getmtime(XL_XLSX_PATH)
            dt = datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M:%S")
            lines.append(f"xl.xlsx: {dt}")
        else:
            lines.append(f"xl.xlsx: файл не найден ({XL_XLSX_PATH})")

        # 2) Время обновления JSON-кэша (если используешь)
        if "XL_JSON_PATH" in globals() and os.path.isfile(XL_JSON_PATH):
            tsj = os.path.getmtime(XL_JSON_PATH)
            dtj = datetime.fromtimestamp(tsj).strftime("%d.%m.%Y %H:%M:%S")
            lines.append(f"top_ssb.jason: {dtj}")

            # 3) meta.updated_at из JSON (если есть)
            try:
                with open(XL_JSON_PATH, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                meta = payload.get("meta", {})
                if meta.get("updated_at"):
                    lines.append(f"meta.updated_at: {meta.get('updated_at')}")
            except Exception:
                pass
        else:
            # если кэша нет — не страшно
            pass

        # 4) Сколько подсказок сейчас в памяти (если есть наборы)
        try:
            n112 = len(globals().get("xl_hints_112", set()) or set())
            n114 = len(globals().get("xl_hints_114", set()) or set())
            n161 = len(globals().get("xl_hints_161", set()) or set())
            lines.append(f"В памяти: 112={n112}, 114={n114}, 161={n161}")
        except Exception:
            pass

        messagebox.showinfo("XL", "\n".join(lines), parent=root)
    except Exception as e:
        messagebox.showerror("XL", f"Не удалось получить информацию об XL:\n{e}", parent=root)

    return "break"

def _xl_file_mtime_dt():
    try:
        if not os.path.isfile(XL_XLSX_PATH):
            return None
        return datetime.fromtimestamp(os.path.getmtime(XL_XLSX_PATH))
    except Exception:
        return None


def _xl_file_mtime_str():
    dt = _xl_file_mtime_dt()
    if not dt:
        return "??.?? ??:??:??"
    return dt.strftime("%d.%m %H:%M:%S")


def _xl_is_stale(hours: int = 24) -> bool:
    dt = _xl_file_mtime_dt()
    if not dt:
        return True
    try:
        age_sec = (datetime.now() - dt).total_seconds()
        return age_sec > hours * 3600
    except Exception:
        return True


def update_xl_button_visual():
    """Красная/зелёная кнопка 📄 по возрасту xl.xlsx."""
    btn = globals().get("xl_button_top")
    if not btn or not btn.winfo_exists():
        return
    try:
        btn.configure(bootstyle=("danger" if _xl_is_stale(24) else "success"))
    except Exception:
        pass


def _load_xl_hints(silent=False):
    """
    Главная функция:
      1) пытаемся прочитать xl.xlsx через openpyxl
      2) если листы "пустые" (max_row/max_col=1) -> нормализуем через Excel COM в xl_norm.xlsx
      3) читаем снова уже xl_norm.xlsx
      4) собираем Result=1 и пишем top_ssb.jason
    """
    global xl_hints_112, xl_hints_114, xl_hints_161

    _xl_log("=== _load_xl_hints START ===")
    _xl_log(f"XL_XLSX_PATH={XL_XLSX_PATH}")

    try:
        os.makedirs(XL_DIR, exist_ok=True)
    except Exception:
        pass

    if not os.path.isfile(XL_XLSX_PATH):
        if not silent:
            messagebox.showerror("XL", f"Файл не найден:\n{XL_XLSX_PATH}")
        _xl_log("xlsx not found -> END")
        return False

    def _try_read_workbook(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            return wb
        except Exception as e:
            _xl_log(f"openpyxl load_workbook ERROR: {e}\n{traceback.format_exc()}")
            return None

    # 1) читаем оригинал
    wb = _try_read_workbook(XL_XLSX_PATH)
    if wb is None:
        if not silent:
            messagebox.showerror("XL", "Не удалось открыть xl.xlsx (openpyxl). См. xl_debug.txt")
        _xl_log("wb is None -> END")
        return False

    _xl_log(f"sheetnames={getattr(wb, 'sheetnames', [])}")

    # выберем реальные имена листов
    sheet_real = {}
    for k, variants in XL_TARGET_SHEETS.items():
        sheet_real[k] = _pick_sheet_name(wb, variants)

    # проверка "пустоты"
    looks_empty = False
    for k in ("112", "114", "161"):
        sname = sheet_real.get(k)
        if not sname:
            continue
        try:
            ws = wb[sname]
            if _is_ws_looks_empty(ws):
                looks_empty = True
        except Exception:
            looks_empty = True

    # 2) если пусто — нормализуем через Excel COM
    if looks_empty:
        _xl_log("looks_empty=True -> try normalize")
        ok = _excel_make_normalized_copy(
            XL_XLSX_PATH,
            XL_NORM_XLSX_PATH,
            [v for v in sheet_real.values() if v],
        )
        if not ok:
            if not silent:
                messagebox.showerror(
                    "XL",
                    "Excel-файл выглядит пустым для чтения.\n"
                    "Автонормализация не сработала (нужен Excel + pywin32).\n\n"
                    "Варианты:\n"
                    "1) Установить pywin32: pip install pywin32\n"
                    "2) Либо сделать копию руками (как ты уже делал) и сохранить как xl.xlsx.\n\n"
                    "Подробности: xl_debug.txt",
                )
            _xl_log("normalize failed -> END")
            return False

        # перечитываем уже нормализованный файл
        wb = _try_read_workbook(XL_NORM_XLSX_PATH)
        if wb is None:
            if not silent:
                messagebox.showerror("XL", "Не удалось открыть xl_norm.xlsx после нормализации. См. xl_debug.txt")
            _xl_log("wb norm None -> END")
            return False
        _xl_log(f"norm sheetnames={getattr(wb, 'sheetnames', [])}")
        # листы могли переименоваться чуть иначе (но мы их сохранили с теми же именами)
        for k, variants in XL_TARGET_SHEETS.items():
            sheet_real[k] = _pick_sheet_name(wb, variants)

    # 3) сбор
    out_serials = {"112": [], "114": [], "161": []}

    for k, prefix in (("112", "12"), ("114", "14"), ("161", "61")):
        sname = sheet_real.get(k)
        if not sname:
            _xl_log(f"[{k}] sheet missing")
            continue
        try:
            ws = wb[sname]
            out_serials[k] = _collect_verified_from_ws(ws, want_prefix=prefix, max_collect=xl_max_rows)
        except Exception as e:
            _xl_log(f"[{k}] read ERROR: {e}\n{traceback.format_exc()}")

    # 4) конвертим серийники блоков в 4-ключи
    def _to_key_list(serial_list):
        keys = []
        for s in serial_list:
            kk = _block_serial_to_key(s)
            if kk:
                keys.append(kk)
        # уникальность
        uniq = []
        seen = set()
        for kk in keys:
            if kk not in seen:
                uniq.append(kk)
                seen.add(kk)
        return uniq

    keys112 = _to_key_list(out_serials["112"])
    keys114 = _to_key_list(out_serials["114"])
    keys161 = _to_key_list(out_serials["161"])

    xl_hints_112 = set(keys112)
    xl_hints_114 = set(keys114)
    xl_hints_161 = set(keys161)

    _xl_log(f"keys112={len(keys112)} keys114={len(keys114)} keys161={len(keys161)}")

    # 5) пишем JSON (важно: формат списками, чтобы json.dump не упал)
    payload = {
        "meta": {
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "source": XL_XLSX_PATH,
            "normalized_used": bool(looks_empty),
            "normalized_path": XL_NORM_XLSX_PATH if looks_empty else "",
        },
        "112": [list(k) for k in keys112],
        "114": [list(k) for k in keys114],
        "161": [list(k) for k in keys161],
    }

    ok = _write_json_atomic(XL_JSON_PATH, payload)
    if not ok:
        if not silent:
            messagebox.showerror(
                "XL",
                "Не удалось записать top_ssb.jason.\n"
                "Скорее всего файл открыт/залочен.\n"
                "Закрой его и попробуй снова.\n\n"
                "Подробности: xl_debug.txt",
            )
        _xl_log("write_json failed -> END")
        return False

    _xl_log("=== _load_xl_hints END (OK) ===")
    return True


def _xl_hint_on(key, block_type, installed: bool) -> bool:
    """
    True -> подсвечиваем bootstyle='info'
    Сейчас: только ВС6Д, только если блок не установлен.
    """
    try:
        if installed:
            return False
        if str(get_variant(key)).strip() != "ВС6Д":
            return False

        bt = str(block_type).strip()
        if bt == "ССБ 112":
            return key in xl_hints_112
        if bt == "ССБ 114":
            return key in xl_hints_114
        if bt == "ССБ 161":
            return key in xl_hints_161
        return False
    except Exception:
        return False

def _xl_blue_pairs_in_program():
    """
    Возвращает множество пар (key, block_type), которые ДОЛЖНЫ быть подсвечены синим
    по XL прямо сейчас (с учётом installed, variant и т.п.).
    Считаем только те изделия, которые реально есть в products и не в storage/assembled.
    """
    pairs = set()
    try:
        for key, blocks in products.items():
            # подсветка нужна в "работе", не для склада/архива
            if key in storage_products or key in assembled_products:
                continue

            needed_blocks = block_types_for(key)
            for bt in needed_blocks:
                installed = bool(blocks.get(bt, False))
                if _xl_hint_on(key, bt, installed):
                    pairs.add((key, bt))
    except Exception:
        pass
    return pairs

def _open_xl_settings_popup(_e=None):
    global xl_settings_win, xl_max_rows

    try:
        # если окно уже есть — просто поднять
        if xl_settings_win is not None and xl_settings_win.winfo_exists():
            xl_settings_win.deiconify()
            xl_settings_win.lift()
            xl_settings_win.focus_force()
            return
    except Exception:
        pass

    win = tk.Toplevel(root)
    xl_settings_win = win
    win.title("Чтение данных из Excel")
    try:
        win.configure(bg=BG_MAIN)
    except Exception:
        pass

    container = tb.Frame(win, style="Main.TFrame", padding=10)
    container.pack(fill="both", expand=True)

    tb.Label(
        container,
        text="Чтение данных из Excel",
        style="Header2.TLabel",
    ).pack(anchor="w", padx=8, pady=(0, 8))

    # подсказка о файле
    stale = _xl_is_stale(24)
    info_text = f"Файл: {XL_XLSX_PATH}\nОбновлён: {_xl_file_mtime_str()}"
    if stale:
        info_text = "Файл не обновлялся более 24 часов назад\n" + info_text

    tb.Label(
        container,
        text=info_text,
        style="Main.TLabel",
        justify="left",
    ).pack(anchor="w", padx=8, pady=(0, 10))

    row = tb.Frame(container, style="Main.TFrame")
    row.pack(fill="x", padx=8, pady=(0, 10))

    tb.Label(row, text="Количество записей (max 1500):", style="Main.TLabel").pack(side="left")

    xl_max_rows_var = tk.StringVar(value=str(xl_max_rows))

    ent_box = _wrap_with_border(row)
    ent_box.pack(side="left", padx=(10, 0))

    ent = tb.Entry(ent_box, textvariable=xl_max_rows_var, width=8, bootstyle="light", font=FONT_PRIMARY)
    ent.pack(padx=4, pady=3)

    btns = tb.Frame(container, style="Main.TFrame")
    btns.pack(fill="x", padx=8)

    def _apply():
        global xl_max_rows
        try:
            v = int((xl_max_rows_var.get() or "").strip())
            v = max(1, min(XL_MAX_ROWS_LIMIT, v))
            xl_max_rows = v
            save_data(False)
            messagebox.showinfo("XL", f"Применено.\nКоличество записей: {xl_max_rows}", parent=win)
        except Exception:
            messagebox.showerror("XL", "Введите число от 1 до 1500.", parent=win)

    tb.Button(btns, text="Сохранить", bootstyle="success", command=_apply).pack(side="right", padx=(8, 0))
    tb.Button(btns, text="Закрыть", bootstyle="secondary", command=win.destroy).pack(side="right")

    # при закрытии сбросить ссылку
    def _on_close():
        global xl_settings_win
        try:
            xl_settings_win = None
        except Exception:
            pass
        try:
            win.destroy()
        except Exception:
            pass

    win.protocol("WM_DELETE_WINDOW", _on_close)

def _xl_current_hint_pairs():
    """
    Считаем пары (key, block_type), которые СЕЙЧАС должны подсвечиваться XL-ом.
    Это именно то, что пользователь видит как "голубые кнопки".
    """
    pairs = set()
    try:
        for key, blocks in products.items():
            # подсветка сейчас делается только для ВС6Д — как у тебя в _xl_hint_on
            if str(get_variant(key)).strip() != "ВС6Д":
                continue
            if key in assembled_products or key in storage_products:
                continue

            for bt in block_types_for(key):
                installed = bool(blocks.get(bt, False))
                if _xl_hint_on(key, bt, installed):
                    pairs.add((key, bt))
    except Exception:
        pass
    return pairs


def _safe_apply_xl():
    """
    ЛКМ по 📄:
      - грузим подсказки из xl.xlsx
      - обновляем кэш/список
      - показываем уведомление: сколько НОВЫХ подсветок появилось
      - красим кнопку по давности файла (24ч)
    """
    try:
        # что уже подсвечено ДО нажатия
        before_pairs = _xl_current_hint_pairs()

        ok = _load_xl_hints(silent=False)
        if ok:
            _load_xl_cache(silent=True)
            update_product_list(preserve_scroll=True)

        # что подсвечено ПОСЛЕ
        after_pairs = _xl_current_hint_pairs()
        new_pairs = after_pairs - before_pairs

        # сообщение
        lines = []
        if _xl_is_stale(24):
            lines.append("Файл не обновлялся более 24 часов назад")
            lines.append("")

        lines.append(f"Найдено {len(new_pairs)} новых блоков")
        lines.append(f"Excel файл обновлён: {_xl_file_mtime_str()}")

        messagebox.showinfo("XL", "\n".join(lines), parent=root)

    except Exception as e:
        try:
            messagebox.showerror("XL", f"Ошибка применения XL:\n{e}", parent=root)
        except Exception:
            pass
    finally:
        # всегда обновляем цвет кнопки
        update_xl_button_visual()


def _save_112_registry(data: dict):
    try:
        os.makedirs(BLOCK_DATA_DIR, exist_ok=True)
        with open(BLOCK112_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# --- Совместимость со старым кодом, где могли остаться эти имена ---
DATA_FILE  = BASE_DATA_FILE       # чтобы старые обращения не падали
BACKUP_FILE = BACKUP_DATA_FILE

def _112_triplet_for_key(key):
    yy, z, num4 = serial_parts_from_key(key)
    v = get_variant(key)
    if v == "ВС6Д":
        last4 = num4
    else:
        meta = (vs7_112_meta.get(key) or {})
        last4 = num4 if meta.get("from_vs6d") is True else meta.get("last4")
    if isinstance(last4, str) and last4.isdigit() and len(last4) == 4:
        return (yy, z, last4)
    return None

def _load_block112_index():
    """Построить индекс 112 по (yy, z, last4) → key на основе текущих products."""
    global block112_index
    block112_index = {}
    for key, blocks in products.items():
        if blocks.get("ССБ 112", False):
            t = _112_triplet_for_key(key)
            if t:
                block112_index[t] = key

def _format_112_triplet(t):
    yy, z, last4 = t
    return f"12{yy}{z}{last4}"

def _112_conflict_for(key, candidate_last4: str):
    """Проверка конфликта (yy, z, last4) → другой key."""
    yy, z, _ = serial_parts_from_key(key)
    last4 = str(candidate_last4).zfill(4)
    t = (yy, z, last4)
    other = block112_index.get(t)
    return (t, other) if other and other != key else (t, None)


def _save_block112_index():
    try:
        os.makedirs(BLOCK_DATA_DIR, exist_ok=True)
        with open(BLOCK112_FILE, "w", encoding="utf-8") as f:
            json.dump(block112_index, f, ensure_ascii=False, indent=2)
    except Exception:
        # не мешаем основной логике, просто тихо игнорируем
        pass

def _prompt_vs7_112_last4_and_set(key):
    # спрашиваем 4 цифры
    s = simpledialog.askstring("ССБ 112 (ВС7)", "Введите последние 4 цифры (0000-9999):", parent=root)
    if not s:
        return
    if not (s.isdigit() and len(s) == 4):
        messagebox.showerror("Ошибка", "Нужно ввести ровно 4 цифры.")
        return

    # жёсткая проверка по (yy, z, last4)
    t, other = _112_conflict_for(key, s)
    if other and other != key:
        messagebox.showerror(
            "Дубликат 112",
            f"Блок {_format_112_triplet(t)} уже установлен на {format_key_long(other)}."
        )
        return

    # записываем meta для ВС7 и ставим флаг блока
    meta = vs7_112_meta.setdefault(key, {})
    meta["from_vs6d"] = False
    meta["last4"] = s
    products.setdefault(key, {})["ССБ 112"] = True

    save_data(False)
    _load_block112_index()   # перестроим индекс
    update_product_list()

def _register_vs7_112_assignment(key, last4: str):
    """Фиксируем, что у ВС7 (key) установлен 112 с указанными 4 цифрами."""
    yy, z, num4 = serial_parts_from_key(key)  # yy, z, xxxx из вашего кода
    rec = {
        "variant": "ВС7",
        "serp_number": num4,     # номер СЕРПа для сообщения 'ВС7 №xxxx'
        "factory": z,
        "year2": yy,
        "serp_key": list(key) if isinstance(key, tuple) else key,
        "ts": datetime.now().isoformat(timespec="seconds"),
    }
    lst = [r for r in block112_index.get(last4, []) if r.get("serp_key") != rec["serp_key"]]
    lst.append(rec)
    block112_index[last4] = lst
    _save_block112_index()

def _warn_if_112_already_used(last4: str):
    """
    Показываем информационное сообщение, если last4 совпадает:
     - с номером СЕРПа ВС6Д (в работе/хранилище/архив)
     - либо с ранее записанным 112 на ВС7 (из файла)
    Ничего НЕ блокируем, просто информируем.
    """
    last4 = str(last4).zfill(4)

    # 1) поиск по ВС6Д среди всех ключей (работа/хранилище/архив)
    all_keys = set(products.keys()) | set(storage_products) | set(assembled_products)
    hits_vs6d = []
    for k in all_keys:
        try:
            if get_variant(k) == "ВС6Д" and _split_key_safe(k)[0] == last4:
                factory_name = _split_key_safe(k)[1]
                z = factory_reverse_mapping.get(factory_name, "?")
                hits_vs6d.append(f"№{last4} [{z}]")
        except Exception:
            continue

    if hits_vs6d:
        messagebox.showwarning("ССБ 112", f"Данный блок установлен на ВС6Д {', '.join(hits_vs6d)}.")
        return

    # 2) проверка по ранее сохранённым установкам 112 на ВС7
    recs = block112_index.get(last4) or []
    recs_vs7 = [r for r in recs if r.get("variant") == "ВС7"]
    if recs_vs7:
        # показываем последнюю запись
        r = recs_vs7[-1]
        num = r.get("serp_number", "????")
        messagebox.showwarning("ССБ 112", f"Данный блок установлен на ВС7 №{num}.")


MONTHS_GENITIVE = [
    "января",
    "февраля",
    "марта",
    "апреля",
    "мая",
    "июня",
    "июля",
    "августа",
    "сентября",
    "октября",
    "ноября",
    "декабря",
]
marked_statuses = {}
show_draft_group = False
request_filter = "Все"
up112_hints = set()
highlight_112 = set()
UP112_PATH = r"C:\ssb_data\global_data\up_112"
BLOCKS_BY_VARIANT = {
    "ВС6Д": ["ССБ 112", "ССБ 114", "ССБ 161"],
    "ВС7": ["ССБ 112", "ССБ 116"],
}

# --- helpers ---
def _as_bool(x):
    if isinstance(x, bool):
        return x
    if isinstance(x, (int, float)):
        return bool(x)
    return str(x).strip().lower() in ("1", "true", "yes", "y", "on")


def _needs_section_frame(hid: str) -> bool:
    """
    Какие заголовки считаем «секциями» (нужна тонкая рамка вокруг группы).
    - В режиме 'Заготовки': сами секции «Заготовки», «Остальные», «В ремонте»
    - В режиме 'Блоки': группы по количеству (Готовы к сборке, 2 блока ... 0 блоков)
    """
    return (
        (":hdr:drafts" in hid) or
        (":hdr:others" in hid) or
        (":hdr:repair" in hid) or
        (":grp:" in hid)          # это заголовки вида WВС6Д:grp:3, WВС7:grp:0 и т.п.
    )

def get_variant(key) -> str:
    return key[2] if isinstance(key, tuple) and len(key) >= 3 else "ВС6Д"


def with_variant(key, variant: str):
    if len(key) == 3:
        return (key[0], key[1], variant)
    return (key[0], key[1], variant)


def key2to3(key):
    return key if len(key) == 3 else (key[0], key[1], "ВС6Д")

# --- tuple key <-> string for JSON ---
_KEY_SEP = "|"

# --- Хелперы сериализации ключей (если своих нет) ---
def _k2s(k):
    # tuple key -> string для JSON
    if isinstance(k, tuple):
        return "|".join(map(str, k))
    return str(k)

def _s2k(s):
    # string -> tuple key (под твою схему ключей: number, factory, variant, yy)
    if isinstance(s, str) and "|" in s:
        parts = s.split("|")
        if len(parts) >= 4:
            return (parts[0], parts[1], parts[2], parts[3])
        if len(parts) == 3:
            return (parts[0], parts[1], parts[2])
        if len(parts) == 2:
            return (parts[0], parts[1])
    return s



def block_types_for(key):
    return BLOCKS_BY_VARIANT.get(get_variant(key), BLOCKS_BY_VARIANT["ВС6Д"])


def installed_count_for(key, blocks):
    """
    Возвращает (cnt, need):
      cnt  — сколько нужных блоков установлено,
      need — сколько блоков всего требуется для этого ключа (по варианту).
    Учтено: блок в состоянии 'РЕМОНТ' не считается установленным.
    """
    needed = block_types_for(key)
    cnt = sum(1 for t in needed
              if blocks.get(t, False) and (key, t) not in repair_blocks)
    return cnt, len(needed)



def get_product_prefix_from_key(key, compact: bool = False) -> str:
    v = get_variant(key)
    return f"{v}" if compact else f"{v}"


def prefix_with_spacing_before_num(key) -> str:
    """Возвращает 'ВС6Д ' или 'ВС7  ' — два пробела для ВС7, один для ВС6Д."""
    pref = get_product_prefix_from_key(key, compact=True)  # 'ВС6Д' или 'ВС7'
    return f"{pref} " if get_variant(key) == "ВС7" else f"{pref} "


def _split_key_safe(key):
    if len(key) == 3:
        return key[0], key[1], key[2]
    return key[0], key[1], "ВС6Д"


def format_key_short(key):
    # Короткая форма тоже с полным серийником
    return f"{get_product_prefix_from_key(key)} №{serial_for_key(key)}"

# --- Tooltip для ССБ 112 (ВС7) ----------------------------------------------

def _112_last4_for(key) -> str | None:
    """Вернёт last4 для ВС7 из vs7_112_meta, если задано корректно (4 цифры)."""
    try:
        meta = vs7_112_meta.get(key) or {}
        last4 = meta.get("last4", "")
        if isinstance(last4, str) and last4.isdigit() and len(last4) == 4:
            return last4
    except Exception:
        pass
    return None

def _show_tooltip(widget, text: str):
    # уже открыт? — перерисуем
    tip = getattr(widget, "_tip_win", None)
    if tip and tip.winfo_exists():
        try:
            tip.destroy()
        except Exception:
            pass

    tip = tk.Toplevel(widget)
    tip.wm_overrideredirect(True)
    tip.attributes("-topmost", True)

    # позиция — под кнопкой
    try:
        x = widget.winfo_rootx()
        y = widget.winfo_rooty() + widget.winfo_height() + 2
    except Exception:
        x = y = 0
    tip.wm_geometry(f"+{x}+{y}")

    lbl = tk.Label(
        tip, text=text,
        bg="#333333", fg="#ffffff",
        padx=6, pady=3,
        relief="solid", borderwidth=1,
        font=("Segoe UI", 9)
    )
    lbl.pack()
    widget._tip_win = tip

def _hide_tooltip(widget):
    tip = getattr(widget, "_tip_win", None)
    if tip and tip.winfo_exists():
        try:
            tip.destroy()
        except Exception:
            pass
    widget._tip_win = None

def attach_vs7_112_tooltip(btn, key):
    """Привязать ховер только для ВС7 + ССБ 112. Показывает last4, если блок включён."""
    def _on_enter(_e=None):
        try:
            v = str(get_variant(key)).strip()
            if v not in ("ВС7", "ВС13"):
                return
            # показываем только если 112 активен (зелёный)
            if not products.get(key, {}).get("ССБ 112", False):
                return
            last4 = _112_last4_for(key)
            if not last4:
                return  # не задано — не шумим
            _show_tooltip(btn, f"№ {last4}")
        except Exception:
            pass

    def _on_leave(_e=None):
        _hide_tooltip(btn)

    # снимем старые Enter/Leave, чтобы не копилось
    try:
        btn.unbind("<Enter>")
        btn.unbind("<Leave>")
    except Exception:
        pass

    btn.bind("<Enter>", _on_enter)
    btn.bind("<Leave>", _on_leave)
# ---------------------------------------------------------------------------

def _get_vs7_ssb112_serial(key):
    """
    Возвращает серийник блока ССБ112 для изделия ВС7, если он сохранён в vs7_112_meta.
    Пытаемся по key, потом по key2to3(key).
    Формат возвращаемой строки: '122520432' (как в Excel), либо None.
    """
    try:
        # пробуем ключи в разных форматах
        candidates = []
        candidates.append(key)
        try:
            candidates.append(key2to3(key))
        except Exception:
            pass

        val = None
        for kk in candidates:
            if kk in vs7_112_meta:
                val = vs7_112_meta.get(kk)
                break

        if val is None:
            return None

        # val может быть строкой
        if isinstance(val, str):
            s = val.strip()
            return s if s else None

        # или словарём
        if isinstance(val, dict):
            # самые вероятные ключи (под твою структуру потом подгоним точно)
            for fld in ("ssb112", "ssb_112", "block112", "block_112", "serial", "serial112"):
                if fld in val:
                    s = str(val.get(fld, "")).strip()
                    if s:
                        return s

        # что-то неизвестное
        return None
    except Exception:
        return None


def format_key_long(key):
    # "СЕРП ВС6Д №yyz0001"
    number, factory_name, _ = _split_key_safe(key)
    return f"{get_product_prefix_from_key(key)} №{serial_for_key(key)}"


def on_window_state_change(event=None):
    # Больше не трогаем storage_visible при zoom/normal
    pass


def resource_path(relative_path):
    return (
        os.path.join(sys._MEIPASS, relative_path)
        if hasattr(sys, "_MEIPASS")
        else os.path.join(os.path.abspath("."), relative_path)
    )



products, assembled_products, storage_products = {}, set(), set()
draft_products, redy_products, assembly_dates, storage_dates = set(), set(), {}, {}
locked_storage = set()
assembly_years, comments, product_dates = {}, {}, {}
sort_order, archive_view_mode = (
    "old_first",
    "journal",
)
FACTORY_ORDER_FIXED = ["ВЕКТОР", "ИНТЕГРАЛ", "РЗП", "КНИИТМУ", "СВТ", "СИГНАЛ"]
factory_order = FACTORY_ORDER_FIXED[:]  # используем копию
data_path, work_scroll_position, storage_scroll_position = base_path, 0, 0
archive_scroll_position, work_top_element, storage_top_element = 0, None, None
archive_top_element, long_press_timers = None, {}
work_widgets, storage_widgets, header_widgets, date_header_widgets = {}, {}, {}, {}
scaling_factor = 2.0
factory_mapping = {
    "1": "ВЕКТОР",
    "2": "ИНТЕГРАЛ",
    "3": "РЗП",
    "4": "КНИИТМУ",
    "5": "СВТ",
    "6": "СИГНАЛ",
}
factory_reverse_mapping = {v: k for k, v in factory_mapping.items()}
type_mapping, all_block_types = {"61": "ССБ 161", "12": "ССБ 112", "14": "ССБ 114"}, [
    "ССБ 161",
    "ССБ 112",
    "ССБ 114",
]


def parse_serial_number(serial):
    if not serial.isdigit() or len(serial) != 9:
        return None
    code, item_number, factory_code, year_code = (
        serial[:2],
        serial[-4:],
        serial[-5],
        serial[-7:-5],
    )
    if code not in type_mapping or factory_code not in factory_mapping:
        return None
    return {
        "Тип изделия": type_mapping[code],
        "Год выпуска": year_code,
        "Завод": factory_mapping[factory_code],
        "Номер изделия": item_number,
        "Ключ изделия": (item_number, factory_mapping[factory_code]),
        "Серийный номер": serial,
    }

def _empty_base_payload():
    return {
        "products": {},
        "storage": [],
        "draft": [],
        "redy": [],
        "comments": {},
        "product_dates": {},
        "assembly_dates": {},   # на базе обычно пусто, но оставим для совместимости
        "storage_dates": {},
        "assembly_years": {},
        "marked_statuses": {},
        "up112_hints": [],
        "repair_blocks": [],
        "locked_storage": [],
        "settings": {
            "factory_order": ["ВЕКТОР", "ИНТЕГРАЛ", "РЗП", "КНИИТМУ", "СВТ", "СИГНАЛ"],
            "sort_order": "new_first",
            "archive_view_mode": "journal",
            "data_path": SERP_BASE_DIR,
            "scaling_factor": 2.0,
            "enable_vs13_mode": False,
            "show_draft_group": False,
            "last_year": None,
        },
    }

def _empty_archive_payload():
    return {
        "assembled": [],
        "assembly_dates": {},
        "storage_dates": {},
        "assembly_years": {},
        "comments": {},
        "product_dates": {},
    }

def _safe_dump_json(path, payload):
    # Пишем атомарно + сразу делаем бэкап в BACKUP_DIR
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def _write_backups(base_payload, arch_payload):
    try:
        _safe_dump_json(BACKUP_DATA_FILE, base_payload)
    except Exception as e:
        print("Бэкап base_data.json не выполнен:", e)
    try:
        _safe_dump_json(BACKUP_ARCHIVE_FILE, arch_payload)
    except Exception as e:
        print("Бэкап base_archive.json не выполнен:", e)

def _ensure_new_files_or_migrate():
    """
    Если новых файлов нет, а LEGACY_DATA_FILE существует — мигрируем.
    Если ничего нет — создаём пустые заготовки.
    """
    have_base = os.path.exists(BASE_DATA_FILE)
    have_arch = os.path.exists(ARCHIVE_FILE)
    if have_base and have_arch:
        return

    if os.path.exists(LEGACY_DATA_FILE) and not (have_base or have_arch):
        # --- миграция со старого единого data.json ---
        with open(LEGACY_DATA_FILE, "r", encoding="utf-8") as f:
            old = json.load(f)

        # Соберём набор архивных ключей
        assembled = set(tuple(x) for x in old.get("assembled", []))

        # Разделим словари по принадлежности ключа к архиву
        def split_map(dct):
            left, right = {}, {}
            for k_json, v in dct.items():
                k = tuple(json.loads(k_json))
                (right if k in assembled else left)[k_json] = v
            return left, right

        # Что у нас было
        products_old         = old.get("products", {})
        comments_old         = old.get("comments", {})
        product_dates_old    = old.get("product_dates", {})
        assembly_dates_old   = old.get("assembly_dates", {})
        storage_dates_old    = old.get("storage_dates", {})
        assembly_years_old   = old.get("assembly_years", {})
        marked_statuses_old  = old.get("marked_statuses", {})
        up112_hints_old      = old.get("up112_hints", [])
        repair_blocks_old    = old.get("repair_blocks", [])
        locked_storage_old   = old.get("locked_storage", [])
        storage_list_old     = old.get("storage", [])
        draft_list_old       = old.get("draft", [])
        redy_list_old        = old.get("redy", [])
        settings_old         = old.get("settings", {})

        # Сплиты
        comments_base,  comments_arch  = split_map(comments_old)
        prod_dates_base, prod_dates_arch = split_map(product_dates_old)
        stor_dates_base, stor_dates_arch = split_map(storage_dates_old)
        assem_years_base, assem_years_arch = split_map(assembly_years_old)

        # База
        new_base = _empty_base_payload()
        new_base["products"]        = products_old
        new_base["storage"]         = storage_list_old
        new_base["draft"]           = draft_list_old
        new_base["redy"]            = redy_list_old
        new_base["comments"]        = comments_base
        new_base["product_dates"]   = prod_dates_base
        new_base["assembly_dates"]  = {}  # в старом формате это по сути даты «отправки» (ушли в архив)
        new_base["storage_dates"]   = stor_dates_base
        new_base["assembly_years"]  = assem_years_base
        new_base["marked_statuses"] = marked_statuses_old
        new_base["up112_hints"]     = up112_hints_old
        new_base["repair_blocks"]   = repair_blocks_old
        new_base["locked_storage"]  = locked_storage_old
        # Settings переносим как есть
        if settings_old:
            new_base["settings"].update(settings_old)
            new_base["settings"]["data_path"] = SERP_BASE_DIR  # фиксированный путь

        # Архив
        new_arch = _empty_archive_payload()
        new_arch["assembled"]       = list(assembled)
        new_arch["assembly_dates"]  = {k:v for k,v in assembly_dates_old.items()}  # всё, что было
        new_arch["storage_dates"]   = stor_dates_arch
        new_arch["assembly_years"]  = assem_years_arch
        new_arch["comments"]        = comments_arch
        new_arch["product_dates"]   = prod_dates_arch

        _safe_dump_json(BASE_DATA_FILE, new_base)
        _safe_dump_json(ARCHIVE_FILE,   new_arch)
        _write_backups(new_base, new_arch)
        return

    # Если вообще ничего не было — создаём пустые файлы
    if not have_base:
        _safe_dump_json(BASE_DATA_FILE, _empty_base_payload())
    if not have_arch:
        _safe_dump_json(ARCHIVE_FILE, _empty_archive_payload())
    _write_backups(_empty_base_payload(), _empty_archive_payload())


def save_data(show_message=False):
    r"""
    Пишем два файла:
      - base_data.json (рабочие данные)
      - base_archive.json (архив)
    Плюс бэкапы в C:\base_basckup\
    """
    global data_path, scaling_factor, show_draft_group, up112_hints, last_selected_year
    data_path = SERP_BASE_DIR

    # Настройки
    settings = {
        "factory_order": factory_order,
        "sort_order": sort_order,
        "archive_view_mode": archive_view_mode,
        "data_path": data_path,
        "scaling_factor": scaling_factor,
        "enable_vs13_mode": enable_vs13_mode,
        "display_full_serials": display_full_serials,
        "display_full_serials_storage": display_full_serials_storage,
        "show_draft_group": show_draft_group,
        "variant_overrides": {_k2s(k): v for k, v in variant_overrides.items()},
        "vs7_112_meta": { _k2s(k): v for k, v in vs7_112_meta.items() },
        "vs6d_only_mode": vs6d_only_mode,
        "hide_added_date": hide_added_date,
        "xl_max_rows": xl_max_rows,
        "last_year": (last_selected_year or (year_var.get().strip() if 'year_var' in globals() else None)),
    }

    # Разделяем словари по принадлежности к архиву
    archived = set(assembled_products)

    def split_dict(src: dict):
        left, right = {}, {}
        for k, v in src.items():
            (right if k in archived else left)[k] = v
        return left, right

    comments_base,       comments_arch        = split_dict(comments)
    product_dates_base,  product_dates_arch   = split_dict(product_dates)
    storage_dates_base,  storage_dates_arch   = split_dict(storage_dates)
    assembly_years_base, assembly_years_arch  = split_dict(assembly_years)
    assembly_dates_base, assembly_dates_arch  = split_dict(assembly_dates)

    # --- базовый файл ---
    base_payload = {
        "products":        {json.dumps(k): v for k, v in products.items()},
        "storage":         list(storage_products),
        "draft":           list(draft_products),
        "redy":            list(redy_products),
        "comments":        {json.dumps(k): v for k, v in comments_base.items()},
        "product_dates":   {json.dumps(k): v for k, v in product_dates_base.items()},
        "assembly_dates":  {json.dumps(k): v for k, v in assembly_dates_base.items()},
        "storage_dates":   {json.dumps(k): v for k, v in storage_dates_base.items()},
        "assembly_years":  {json.dumps(k): v for k, v in assembly_years_base.items()},
        "marked_statuses": {json.dumps(k): v for k, v in marked_statuses.items()},
        "up112_hints":     list(up112_hints),
        "repair_blocks":   [[json.dumps(k), b] for (k, b) in repair_blocks],
        "locked_storage":  list(locked_storage),
        "settings":        settings,
    }

    # --- архивный файл ---
    arch_payload = {
        "assembled":       list(assembled_products),
        "assembly_dates":  {json.dumps(k): v for k, v in assembly_dates_arch.items()},
        "storage_dates":   {json.dumps(k): v for k, v in storage_dates_arch.items()},
        "assembly_years":  {json.dumps(k): v for k, v in assembly_years_arch.items()},
        "comments":        {json.dumps(k): v for k, v in comments_arch.items()},
        "product_dates":   {json.dumps(k): v for k, v in product_dates_arch.items()},
    }

    # запись + бэкапы
    _safe_dump_json(BASE_DATA_FILE, base_payload)
    _safe_dump_json(ARCHIVE_FILE,   arch_payload)
    _write_backups(base_payload, arch_payload)

    if show_message:
        messagebox.showinfo("Сообщение", "Данные сохранены!")

def load_data():
    global products, assembled_products, storage_products
    global comments, product_dates, assembly_dates, storage_dates, assembly_years
    global draft_products, redy_products, marked_statuses, up112_hints, repair_blocks, locked_storage
    global factory_order, sort_order, archive_view_mode, data_path, scaling_factor, show_draft_group, enable_vs13_mode, last_selected_year, display_full_serials, display_full_serials_storage
    global block112_index
    _ensure_new_files_or_migrate()
    _load_block112_index()

    # читаем оба файла
    with open(BASE_DATA_FILE, "r", encoding="utf-8") as f:
        base = json.load(f)
    with open(ARCHIVE_FILE, "r", encoding="utf-8") as f:
        arch = json.load(f)

    # --- база ---
    products = {tuple(json.loads(k)): v for k, v in base.get("products", {}).items()}
    storage_products = set(tuple(x) for x in base.get("storage", []))
    draft_products   = set(tuple(x) for x in base.get("draft", []))
    redy_products    = set(tuple(x) for x in base.get("redy", []))

    comments_base      = {tuple(json.loads(k)): v for k, v in base.get("comments", {}).items()}
    product_dates_base = {tuple(json.loads(k)): v for k, v in base.get("product_dates", {}).items()}
    assembly_dates_base= {tuple(json.loads(k)): v for k, v in base.get("assembly_dates", {}).items()}
    storage_dates_base = {tuple(json.loads(k)): v for k, v in base.get("storage_dates", {}).items()}
    assembly_years_base= {tuple(json.loads(k)): v for k, v in base.get("assembly_years", {}).items()}

    marked_statuses = {tuple(json.loads(k)): v for k, v in base.get("marked_statuses", {}).items()}
    up112_hints     = set(tuple(x) for x in base.get("up112_hints", []))
    repair_blocks   = set((tuple(json.loads(k)), b) for (k, b) in base.get("repair_blocks", []))
    locked_storage  = set(tuple(x) for x in base.get("locked_storage", []))

    # --- архив ---
    assembled_products  = set(tuple(x) for x in arch.get("assembled", []))
    assembly_dates_arch = {tuple(json.loads(k)): v for k, v in arch.get("assembly_dates", {}).items()}
    storage_dates_arch  = {tuple(json.loads(k)): v for k, v in arch.get("storage_dates", {}).items()}
    assembly_years_arch = {tuple(json.loads(k)): v for k, v in arch.get("assembly_years", {}).items()}
    comments_arch       = {tuple(json.loads(k)): v for k, v in arch.get("comments", {}).items()}
    product_dates_arch  = {tuple(json.loads(k)): v for k, v in arch.get("product_dates", {}).items()}

    # --- склейка (в память храним как раньше — единые словари) ---
    comments       = {**comments_base, **comments_arch}
    product_dates  = {**product_dates_base, **product_dates_arch}
    assembly_dates = {**assembly_dates_base, **assembly_dates_arch}
    storage_dates  = {**storage_dates_base, **storage_dates_arch}
    assembly_years = {**assembly_years_base, **assembly_years_arch}

    # настройки
    settings = base.get("settings", {})
    global xl_max_rows
    try:
        v = int(settings.get("xl_max_rows", xl_max_rows))
        xl_max_rows = max(1, min(XL_MAX_ROWS_LIMIT, v))
    except Exception:
        xl_max_rows = 120
    global variant_overrides, vs7_112_meta
    global display_full_serials
    global vs6d_only_mode
    global hide_added_date
    hide_added_date = _as_bool(settings.get("hide_added_date", False))
    vs6d_only_mode = _as_bool(settings.get("vs6d_only_mode", False))
    display_full_serials = _as_bool(settings.get("display_full_serials", display_full_serials))
    display_full_serials_storage = _as_bool(settings.get("display_full_serials_storage", display_full_serials_storage))
    vs6d_only_mode = _as_bool(settings.get("vs6d_only_mode", vs6d_only_mode))

    factory_order       = settings.get("factory_order", ["ВЕКТОР", "ИНТЕГРАЛ", "РЗП", "КНИИТМУ", "СВТ", "СИГНАЛ"])
    sort_order          = settings.get("sort_order", "new_first")
    archive_view_mode   = settings.get("archive_view_mode", "journal")
    data_path           = SERP_BASE_DIR
    scaling_factor      = settings.get("scaling_factor", 2.0)
    show_draft_group    = settings.get("show_draft_group", False)
    enable_vs13_mode    = settings.get("enable_vs13_mode", False)
    last_selected_year  = settings.get("last_year", f"{datetime.now().year % 100:02d}")
    variant_overrides = {_s2k(k): v for k, v in settings.get("variant_overrides", {}).items()}
    vs7_112_meta = {_s2k(k): v for k, v in settings.get("vs7_112_meta", {}).items()}

    # Миграции форматов ключей (2->3->4) поверх объединённых данных
    def _migrate_2to3(dct_like=None, set_like=None):
        changed = False
        if dct_like is not None:
            for k in list(dct_like.keys()):
                if isinstance(k, tuple) and len(k) == 2:
                    v = dct_like.pop(k)
                    dct_like[(k[0], k[1], "ВС6Д")] = v
                    changed = True
        if set_like is not None:
            old = list(set_like)
            set_like.clear()
            for k in old:
                if isinstance(k, tuple) and len(k) == 2:
                    set_like.add((k[0], k[1], "ВС6Д"))
                    changed = True
                else:
                    set_like.add(k)
        return changed

    changed_any = False
    for d in (products, comments, product_dates, assembly_dates, storage_dates, assembly_years, marked_statuses):
        changed_any |= _migrate_2to3(dct_like=d)
    for s in (assembled_products, storage_products, draft_products, redy_products, locked_storage):
        changed_any |= _migrate_2to3(set_like=s)

    # Привести repair_blocks к 4-ключам как у вас раньше
    if repair_blocks:
        by3_to_year2 = {}
        for k in products.keys():
            if isinstance(k, tuple) and len(k) >= 4:
                by3_to_year2[(k[0], k[1], k[2])] = k[3]
        for k, v in assembly_years.items():
            k3 = key2to3(k)
            if k3 not in by3_to_year2 and isinstance(v, str):
                by3_to_year2[k3] = (v or "??")[:2]

        def _to4(k3or4):
            if isinstance(k3or4, tuple) and len(k3or4) >= 4:
                return k3or4
            k3 = key2to3(k3or4)
            yy = by3_to_year2.get(k3, "??")
            return (k3[0], k3[1], k3[2], yy)

        repair_blocks = set((_to4(k), b) for (k, b) in repair_blocks)


    if changed_any:
        save_data(False)
    # --- XL подсветка: после любой загрузки данных подхватываем кэш из top_ssb.jason ---
    try:
        _load_xl_cache(silent=True)   # читает XL_JSON_PATH и заполняет xl_hints_112/114/161
    except Exception:
        pass




def _conflicts_for_number_factory(num4: str, factory_name: str, variant: str = None, year2: str = None):
    """
    Ищем конфликты: теперь учитываем ГОД.
    Совпадение считается конфликтом только если совпали номер+завод+вариант+год.
    """
    conflicts = []
    all_keys = set(products.keys()) | set(storage_products) | set(assembled_products)
    for k in all_keys:
        if not (isinstance(k, tuple) and len(k) >= 3):
            continue
        same_num_factory = k[0] == num4 and k[1] == factory_name
        same_variant = (variant is None) or (get_variant(k) == variant)
        same_year = (year2 is None) or (get_year2(k) == year2)
        if same_num_factory and same_variant and same_year:
            if k in assembled_products:
                conflicts.append(f"№{serial_for_key(k)} — уже в архиве")
            elif k in storage_products:
                conflicts.append(f"№{serial_for_key(k)} — собран и ждёт отправки на склад")
            elif k in products:
                conflicts.append(f"№{serial_for_key(k)} — уже добавлен")
    return conflicts

def block_is_on(key, block_type: str) -> bool:
    """Возвращает True, если блок сейчас включен (зелёный) для данного изделия."""
    st = marked_statuses.get(key)
    bt = str(block_type).strip()
    if isinstance(st, dict):
        return bool(st.get(bt))
    if isinstance(st, (set, list, tuple)):
        return bt in st
    return False


def _create_blank_blocks_for_variant(variant: str) -> dict:
    return {t: False for t in BLOCKS_BY_VARIANT.get(variant, BLOCKS_BY_VARIANT["ВС6Д"])}


def process_serial():
    MAX_BATCH = 300
    raw = entry_var.get().strip()
    year = year_var.get().strip()
    if not (year.isdigit() and len(year) == 2):
        messagebox.showerror("Ошибка", "Некорректный год (нужно ровно 2 цифры).")
        return
    variants_to_add = ["ВС6Д"] if vs6d_only_mode else [product_mode]
    if "-" in raw:
        s = raw.replace(" ", "")
        try:
            left, right = s.split("-", 1)
        except ValueError:
            messagebox.showerror("Ошибка", "Неверный формат диапазона. Пример: 45-5066")
            return
        if not (left.isdigit() and right.isdigit()) or len(right) < 2:
            messagebox.showerror(
                "Ошибка",
                "После тире нужна вторая часть и код завода в конце (пример: 45-5066).",
            )
            return
        start_num = int(left)
        factory_code = right[-1]
        end_num_str = right[:-1]
        if factory_code not in factory_mapping:
            messagebox.showerror(
                "Ошибка", "Код завода должен быть 1–6 (последняя цифра)."
            )
            return
        if end_num_str == "":
            messagebox.showerror(
                "Ошибка", "После тире должен быть номер и код завода. Пример: 45-5066."
            )
            return
        end_num = int(end_num_str)
        if not (0 <= start_num <= 9999 and 0 <= end_num <= 9999):
            messagebox.showerror("Ошибка", "Номера должны быть 0..9999.")
            return
        if end_num < start_num:
            messagebox.showerror("Ошибка", "Конец диапазона меньше начала.")
            return
        count_numbers = end_num - start_num + 1
        total_to_add = count_numbers * len(variants_to_add)
        if total_to_add > MAX_BATCH:
            messagebox.showerror(
                "Слишком много", f"Диапазон * варианты = {total_to_add} > {MAX_BATCH}."
            )
            return
        factory_name = factory_mapping[factory_code]
        conflicts = []
        for n in range(start_num, end_num + 1):
            num4 = f"{n:04d}"
            for variant in variants_to_add:
                conflicts.extend(
                    _conflicts_for_number_factory(num4, factory_name, variant, year)
                )
        if conflicts:
            lines = "\n".join(conflicts[:50])
            more = "" if len(conflicts) <= 50 else f"\n…и ещё {len(conflicts) - 50}"
            messagebox.showerror(
                "Дубликаты в диапазоне", f"Ничего не добавлено:\n{lines}{more}"
            )
            return
        for n in range(start_num, end_num + 1):
            num4 = f"{n:04d}"
            for variant in variants_to_add:
                key4 = (num4, factory_name, variant, year)
                products[key4] = _create_blank_blocks_for_variant(variant)
                product_dates[key4] = datetime.now().strftime("%d.%m.%y")
                assembly_years[key4] = year
                storage_products.discard(key4)
                assembled_products.discard(key4)
                draft_products.discard(key4)
                redy_products.discard(key4)
        save_data()
        update_product_list()
        entry_var.set("")
        if len(variants_to_add) == 1:
            messagebox.showinfo(
                "Готово",
                f"Добавлено изделий: {total_to_add}\n{start_num:04d}–{end_num:04d} ({factory_name}, {variants_to_add[0]})",
            )
        else:
            messagebox.showinfo(
                "Готово",
                f"Добавлено изделий: {total_to_add}\n{start_num:04d}–{end_num:04d} ({factory_name}, {' + '.join(variants_to_add)})",
            )
        return
    raw_number = raw
    if not raw_number.isdigit() or len(raw_number) < 2 or int(raw_number) > 99999:
        messagebox.showerror(
            "Ошибка", "Введите от 2 до 5 цифр (номер+код завода в конце)."
        )
        return
    factory_code = raw_number[-1]
    if factory_code not in factory_mapping:
        messagebox.showerror("Ошибка", "Код завода должен быть 1–6 (последняя цифра).")
        return
    num4 = raw_number[:-1].zfill(4)
    factory_name = factory_mapping[factory_code]
    conflicts = _conflicts_for_number_factory(num4, factory_name, product_mode, year)
    if conflicts:
        messagebox.showwarning("Ошибка", "\n".join(conflicts))
        entry_var.set("")
        return
    for variant in variants_to_add:
        key4 = (num4, factory_name, variant, year)
        products[key4] = _create_blank_blocks_for_variant(variant)
        product_dates[key4] = datetime.now().strftime("%d.%m.%y")
        assembly_years[key4] = year
        storage_products.discard(key4)
        assembled_products.discard(key4)
        draft_products.discard(key4)
        redy_products.discard(key4)
    save_data()
    update_product_list()


    entry_var.set("")

def toggle_block(key, block_type):
    """
    Добавлено: если после переключения блоков у изделия не остаётся ни одного установленного SSB,
    статус ЗАГОТОВКА автоматически снимается (становится НЕ ГОТОВО).
    """
    if key in products:
        products[key][block_type] = not products[key][block_type]

        # если 112-й поставили, убираем подсказку up_112 (храним по 3-ключу)
        if block_type == "ССБ 112" and products[key][block_type]:
            try:
                up112_hints.discard(key2to3(key))
            except Exception:
                pass

        # если все блоки стали серыми — снимаем "ЗАГОТОВКА"
        needed = block_types_for(key)
        if key in draft_products and not any(
            products[key].get(t, False) for t in needed
        ):
            draft_products.discard(key)

        save_data()
        update_row_widgets(key)


def is_block_112_on(key) -> bool:
    """
    True, если для key сейчас включён именно ССБ 112 (зелёная кнопка).
    """
    try:
        st = marked_statuses.get(key)
        if isinstance(st, (set, list, tuple)):
            return "ССБ 112" in st or "ССБ112" in st or "SSB 112" in st or "SSB112" in st
        if isinstance(st, dict):
            return any(st.get(name, False) for name in ("ССБ 112", "ССБ112", "SSB 112", "SSB112"))
    except Exception:
        pass
    # запасной путь через products
    try:
        b = products.get(key, {}) or {}
        return any(b.get(name, False) for name in ("ССБ 112", "ССБ112", "SSB 112", "SSB112"))
    except Exception:
        return False


def on_block_left_click(key, block_type):
    def _is_vs7_like(v):
        s = str(v).strip()
        return s in ("ВС7", "ВС13")

    def _is_bt_112(bt):
        s = str(bt).replace(" ", "").upper()
        return s in ("ССБ112", "SSB112")

    # --- спец-логика: ВС7/ВС13 + ССБ 112 ---
    try:
        if _is_vs7_like(get_variant(key)) and _is_bt_112(block_type):
            # уже включён? → снять и очистить метаданные
            if products.get(key, {}).get("ССБ 112", False):
                products[key]["ССБ 112"] = False
                vs7_112_meta.pop(key, None)
                save_data(False)
                _load_block112_index()
                update_row_widgets(key)
                return

            # диалог ввода последних 4 цифр (без автоподстановки)
            prefill = ""
            while True:
                val = _ask_vs7_112_last4(key, prefill=prefill)
                if val is None or str(val).strip() == "":
                    return  # отмена
                last4 = str(val).strip()
                if not (last4.isdigit() and len(last4) == 4):
                    messagebox.showwarning("ССБ 112", "Нужно ввести ровно 4 цифры.")
                    prefill = last4
                    continue

                # жёсткая проверка конфликта по (yy, z, last4)
                _load_block112_index()  # актуализируем на всякий случай
                t, other = _112_conflict_for(key, last4)
                if other and other != key:
                    messagebox.showwarning(
                        "Дубликат 112",
                        f"Блок 12{t[0]}{t[1]}{t[2]} уже установлен на {format_key_long(other)}."
                    )
                    prefill = last4
                    continue

                # ок — включаем
                vs7_112_meta[key] = {"from_vs6d": False, "last4": last4}
                products.setdefault(key, {}).update({"ССБ 112": True})
                save_data(False)
                _load_block112_index()
                update_row_widgets(key)
                return
    except Exception:
        # если что-то пойдёт не так — мягко fallback на универсальный путь
        pass

    # --- универсальная логика для прочих блоков ---
    if (key, block_type) in repair_blocks:
        repair_blocks.discard((key, block_type))
        save_data(False)
        update_row_widgets(key)
    else:
        toggle_block(key, block_type)



def toggle_block_repair(key, block_type):
    if (key, block_type) in repair_blocks:
        repair_blocks.discard((key, block_type))
    else:
        repair_blocks.add((key, block_type))
    save_data()
    update_row_widgets(key)

def delete_serp(key):
    if not messagebox.askyesno("Подтверждение", f"Удалить {format_key_long(key)}?"):
        return

    changed = set()
    if key in storage_products:
        changed.add("storage")
    else:
        changed.add("work6" if get_variant(key) == "ВС6Д" else "work7")

    products.pop(key, None)
    comments.pop(key, None)
    product_dates.pop(key, None)
    draft_products.discard(key)
    redy_products.discard(key)
    storage_products.discard(key)
    storage_dates.pop(key, None)
    locked_storage.discard(key)
    assembled_products.discard(key)
    assembly_dates.pop(key, None)

    save_data()

    update_product_list(
        preserve_scroll=True,
        changed_canvases=tuple(changed),
        reset_others_to_top=False,
        repack=False,
        regroup=False  # <<< важно
    )


def save_all_comments():
    for area in [work_frame, storage_frame]:
        for widget in area.winfo_children():
            if isinstance(widget, tb.Frame) and hasattr(widget, "key"):
                for child in widget.winfo_children():
                    if isinstance(child, tb.Entry) and hasattr(child, "key_reference"):
                        comments[child.key_reference] = child.get()
    save_data(False)

def _record_vs7_112_meta_on_storage(key):
    """
    Вызывается, когда изделие ВС7 переводят в хранилище.
    Сохраняет метаданные про '112': взяли ли его из ВС6Д и/или последние 4 цифры.
    """
    try:
        if get_variant(key) != "ВС7":
            return
    except Exception:
        return

    # Спросить: 112 взят из ВС6Д?
    ans = messagebox.askyesno("ССБ 112 для ВС7", "112 взят из ВС6Д?", parent=root)
    from_vs6d = bool(ans)
    last4 = ""
    if not from_vs6d:
        # Если свой 112 — можно ввести последние 4 цифры (необязательно)
        s = simpledialog.askstring("ССБ 112 для ВС7", "Последние 4 цифры 112 (можно пусто):", parent=root)
        s = (s or "").strip()
        if s and (not s.isdigit() or len(s) != 4):
            messagebox.showwarning("Формат", "Нужно ввести ровно 4 цифры или оставить пусто.")
            s = ""
        last4 = s

    # Сохраняем метаданные
    vs7_112_meta[key] = {"from_vs6d": from_vs6d, "last4": last4}
    save_data(False)



def mark_storage(key):
    save_all_comments()

    # служебный комментарий убираем из "в работе", как было
    if _is_storage_comment(comments.get(key, "")):
        comments.pop(key, None)
        row = work6_widgets.get(key) or work7_widgets.get(key)
        if row and getattr(row, "comment_entry", None):
            try:
                row.comment_entry.delete(0, tk.END)
            except Exception:
                pass

    storage_products.add(key)
    storage_dates[key] = datetime.now().strftime("%d.%m.%y")
    assembled_products.discard(key)
    draft_products.discard(key)

    save_data()

    target = "work6" if get_variant(key) == "ВС6Д" else "work7"
    update_product_list(
        preserve_scroll=True,
        changed_canvases=(target, "storage"),
        reset_others_to_top=False,
        repack=False,
        regroup=False   # <<< важно: без перегруппировки
    )

def mark_assembled(key):
    if key in storage_products and key in locked_storage:
        messagebox.showinfo("Заблокировано", " Снимите замок, чтобы отправить на склад.")
        return

    save_all_comments()

    assembled_products.add(key)
    assembly_dates[key] = datetime.now().strftime("%d.%m.%y")
    storage_products.discard(key)
    draft_products.discard(key)
    locked_storage.discard(key)

    save_data()
    update_assembly_archive()

    update_product_list(
        preserve_scroll=True,
        changed_canvases=("storage",),
        reset_others_to_top=False,
        regroup=False  # <<< важно
    )
def return_to_work(key):
    save_all_comments()

    # служебный комментарий
    comments[key] = "С хранилища"
    row = storage_widgets.get(key)
    if row and getattr(row, "comment_entry", None):
        try:
            row.comment_entry.delete(0, tk.END)
            row.comment_entry.insert(0, "С хранилища")
        except Exception:
            pass

    # перенос из хранилища обратно "в работу"
    if key in storage_products:
        was_draft = key in draft_products
        storage_products.remove(key)
        storage_dates.pop(key, None)
        if was_draft and key not in draft_products:
            draft_products.add(key)

    save_data()

    target = "work6" if get_variant(key) == "ВС6Д" else "work7"

    # ВАЖНО: здесь нам нужна нормальная сортировка,
    # поэтому включаем regroup=True и repack=True
    update_product_list(
        preserve_scroll=True,
        changed_canvases=("storage", target),
        reset_others_to_top=False,
        repack=True,
        regroup=True
    )


def toggle_draft_status(key):
    """
    ЗАГОТОВКА включается только при наличии хотя бы одного установленного блока ССБ.
    Если блоков нет — ничего не меняем. Снять ЗАГОТОВКА можно всегда.
    """
    save_all_comments()

    needed = block_types_for(key)
    blocks = products.get(key, {})
    has_any_block = any(blocks.get(t, False) for t in needed)

    if key in draft_products:
        draft_products.remove(key)
    else:
        if not has_any_block:
            return
        draft_products.add(key)

    save_data()
    update_row_widgets(key)


def back_to_draft(key):
    save_all_comments()
    if key in redy_products:
        redy_products.remove(key)
        draft_products.add(key)
    save_data()
    update_row_widgets(key)


def toggle_marked_status(key):
    return


def format_date_genitive(date_str):
    if not date_str or not isinstance(date_str, str):
        return date_str
    date_str = date_str.rstrip(".")
    parts = date_str.split(".")
    if len(parts) >= 2:
        try:
            day = int(parts[0])
            month = int(parts[1])
            if 1 <= month <= 12:
                return f"{day} {MONTHS_GENITIVE[month - 1]}"
        except ValueError:
            pass
    return date_str


def format_ddmm(date_str: str) -> str:
    """Вернуть 'дд.мм' из 'дд.мм' или 'дд.мм.гг'."""
    if not isinstance(date_str, str):
        return "??.??"
    parts = date_str.split(".")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return f"{int(parts[0]):02d}.{int(parts[1]):02d}"
    return "??.??"


def send_all_to_storage():
    """
    Пакетная отправка из хранилища:
      - отправляем ТОЛЬКО те, у кого нет замка;
      - делаем одно сохранение/перерисовку;
      - замки у отправленных снимаем.
    """
    # кандидаты к отправке: всё, что в хранилище и ещё не в архиве
    to_process = [k for k in storage_products if k not in assembled_products]
    if not to_process:
        messagebox.showinfo("Информация", "Нет изделий для отправки на склад")
        return

    # разделяем на без замка / с замком
    unlocked = [k for k in to_process if k not in locked_storage]
    locked   = [k for k in to_process if k in locked_storage]

    if not unlocked:
        messagebox.showinfo("Информация", "Нет изделий для отправки на склад.")
        return

    # Подтверждение с количеством
    if not messagebox.askyesno(
        "Подтверждение",
        f"Отправить на склад {len(unlocked)} шт."
    ):
        return

    # Пакетная операция (без дерганья UI на каждую строку)
    save_all_comments()
    save_scroll_positions()

    today = datetime.now().strftime("%d.%m.%y")
    for key in unlocked:
        assembled_products.add(key)
        assembly_dates[key] = today
        storage_products.discard(key)
        draft_products.discard(key)
        # снятие замка у успешно отправленных
        locked_storage.discard(key)

    save_data(False)
    update_assembly_archive()
    update_product_list(preserve_scroll=True)
    restore_scroll_positions()

    # Сообщение пользователю
    if locked:
        # На этом шаге только констатация факта; красивую формулировку «Кроме №…»
        # добавим на шаге 4.
        messagebox.showinfo(
            "Готово",
            f"Отправлено: {len(unlocked)} шт."
        )
    else:
        messagebox.showinfo("Успех", "Все изделия отправлены на склад")

def bulk_mark_assembled(keys):
    """
    Пакетная отправка на склад без N-кратной перерисовки.
    Возвращает количество реально перенесённых изделий.
    """
    if not keys:
        return 0

    save_all_comments()

    now_str = datetime.now().strftime("%d.%m.%y")
    moved = 0
    for key in keys:
        if key in assembled_products:
            continue
        assembled_products.add(key)
        assembly_dates[key] = now_str
        storage_products.discard(key)
        draft_products.discard(key)
        moved += 1

    save_data()
    update_assembly_archive()

    # Меняется только хранилище — перерисуем ТОЛЬКО его.
    # Остальные НЕ трогаем (reset_others_to_top=False).
    update_product_list(
        preserve_scroll=True,
        changed_canvases=("storage",),
        reset_others_to_top=False
    )

    return moved


def parse_date_str(date_str, key=None):
    try:
        parts = date_str.split(".")
        if len(parts) == 2:
            day, month = parts
            if key and key in assembly_years:
                year = (
                    int("20" + assembly_years[key])
                    if len(assembly_years[key]) == 2
                    else int(assembly_years[key])
                )
            else:
                year = datetime.now().year
            return datetime(year, int(month), int(day))
        elif len(parts) == 3:
            day, month, year = parts
            year_full = int("20" + year) if len(year) == 2 else int(year)
            return datetime(year_full, int(month), int(day))
        else:
            return datetime.min
    except Exception:
        return datetime.min


def format_short_serp(key):
    number, factory_name = key
    factory_code = factory_reverse_mapping.get(factory_name, "?")
    return f"{get_product_prefix()} {number} [{factory_code}]"

def update_storage_sort_button_visual():
    """Подпись/цвет кнопки сортировки склада в зависимости от режима."""
    try:
        if storage_sort_mode == "dates":
            storage_sort_button.configure(text="Сортировка", bootstyle="secondary")
        else:
            storage_sort_button.configure(text="Сортировка", bootstyle="info")
    except Exception:
        pass

def cycle_storage_sort_mode():
    """Цикл: По датам <-> По моделям, пересборка правой колонки без сброса скроллов в работе."""
    global storage_sort_mode

    # 1. Запоминаем текущие скроллы для ВС6Д и ВС7 (левая и правая колонка "в работе")
    w6_pos = None
    w7_pos = None

    try:
        if "work_canvas" in globals() and work_canvas.winfo_exists():
            w6_pos = work_canvas.yview()[0]
    except Exception:
        w6_pos = None

    try:
        if "work7_canvas" in globals() and work7_canvas.winfo_exists():
            w7_pos = work7_canvas.yview()[0]
    except Exception:
        w7_pos = None

    # 2. Переключаем режим сортировки хранилища
    storage_sort_mode = "models" if storage_sort_mode == "dates" else "dates"
    save_data(False)
    update_storage_sort_button_visual()

    # 3. Перестраиваем списки:
    #    preserve_scroll=False -> все канвасы уйдут в начало,
    #    НО ниже мы вернём скроллы работы обратно.
    update_product_list(preserve_scroll=False)

    # 4. Восстанавливаем скроллы "в работе", хранилище остаётся в начале
    try:
        if w6_pos is not None and "work_canvas" in globals() and work_canvas.winfo_exists():
            work_canvas.yview_moveto(w6_pos)
    except Exception:
        pass

    try:
        if w7_pos is not None and "work7_canvas" in globals() and work7_canvas.winfo_exists():
            work7_canvas.yview_moveto(w7_pos)
    except Exception:
        pass

    # 5. В режиме «по моделям» обновляем формат подписи строк хранилища
    if storage_sort_mode == "models":
        refresh_storage_labels_in_place()




def update_storage_row_lock_state(key):
    """Перекрасить и (де)активировать кнопки строки хранилища в зависимости от замка."""
    row = storage_widgets.get(key)
    if not row or not row.winfo_exists():
        return
    is_locked = key in locked_storage

    # Кнопки, которые мы сохранили на строке
    btn_return = getattr(row, "btn_return_to_work", None)
    btn_sklad  = getattr(row, "btn_sklad", None)
    btn_lock   = getattr(row, "btn_lock", None)

    if btn_lock:
        btn_lock.configure(bootstyle=("info" if is_locked else "secondary"))

    def _apply(btn):
        if not btn:
            return
        if is_locked:
            btn.configure(state="disabled", bootstyle="light")
        else:
            btn.configure(state="normal",  bootstyle="info" if btn is btn_sklad else "secondary")

    _apply(btn_return)
    _apply(btn_sklad)

def toggle_storage_panel():
    """Показ/скрытие правой панели хранилища (в обычном режиме)."""
    if vs6d_only_mode:
        return "break"  # в режиме ВС6Д хранилище всегда открыто и стрелка не кликается
    global storage_visible
    storage_visible = not storage_visible
    try:
        if storage_visible:
            storage_frame_container.grid()
            toggle_storage_button.configure(text="▶", width=2)
        else:
            storage_frame_container.grid_remove()
            lbl = f"◀ ({storage_count_cache})"
            toggle_storage_button.configure(text=lbl, width=len(lbl))
    except Exception:
        pass
    return "break"


def hide_storage_if_open():
    """Спрятать хранилище, если оно открыто (для случаев после добавления и т.п.)."""
    global storage_visible
    if not storage_visible:
        return
    storage_visible = False
    try:
        storage_frame_container.grid_remove()
        lbl = f"◀ ({storage_count_cache})"
        toggle_storage_button.configure(text=lbl, width=len(lbl))
    except Exception:
        pass


def toggle_storage_lock(key):
    """Переключить замок и обновить строку хранилища."""
    # В VS6D-режиме панель хранилища держим открытой,
    # но сам замок должен переключаться.
    if vs6d_only_mode:
        try:
            storage_frame_container.grid()
        except Exception:
            pass
        # НЕ выходим — продолжаем переключение

    if key in locked_storage:
        locked_storage.remove(key)
    else:
        locked_storage.add(key)

    save_data(False)
    update_storage_row_lock_state(key)



def create_product_row(
    parent, key, blocks, current_comments, in_storage, show_factory_code=False
):
    row = tb.Frame(parent, padding=2, style="Main.TFrame")
    row.pack(fill=X, pady=PADY_PRODUCT)
    row.key = key

    number, factory_name, _ = _split_key_safe(key)
    factory_code = factory_reverse_mapping.get(factory_name, "?")
    if in_storage:
        # --- Надпись слева в хранилище ---
        number, factory_name, _ = _split_key_safe(key)
        z = factory_reverse_mapping.get(factory_name, "?")

        if display_full_serials_storage:
            # Полный серийный номер: №yyzxxxx (без пробела после №)
            label_text = f"№{serial_for_key(key)}"
        else:
            pref = get_product_prefix_from_key(key, compact=True)  # "ВС6Д" / "ВС7"
            if storage_sort_mode == "models":
                # Режим "по моделям": "<модель> №xxxx [z]"
                label_text = f"{pref} №{number} [{z}]"
            else:
                # Режим "по датам/по заводам": "<модель> №xxxx" — без [z]
                label_text = f"{pref} №{number}"

        serial_lbl = tb.Label(
            row,
            text=label_text,
            font=FONT_PRIMARY,
            style="Main.TLabel"
        )
        serial_lbl.pack(side="left", fill=X, expand=True)
        row.serial_label = serial_lbl

        # Подсказка "112: xxxx" для ВС7 в хранилище
        try:
            for ev in ("<Enter>", "<Leave>", "<Motion>"):
                try:
                    serial_lbl.unbind(ev)
                except Exception:
                    pass

            if str(get_variant(key)).strip() in ("ВС7", "ВС13"):
                attach_storage_112_tooltip(serial_lbl, key)
        except Exception:
            pass

        # Обёртка с тонкой рамкой вокруг Entry
        comment_box = _wrap_with_border(row)
        comment_box.pack(side=LEFT, padx=5, fill=X, expand=True)

        comment_entry = tb.Entry(
            comment_box, width=15, font=FONT_PRIMARY, bootstyle="light"
        )
        comment_entry.pack(fill=X, padx=4, pady=3)
        comment_entry.key_reference = key
        comment_entry.insert(0, current_comments.get(key, comments.get(key, "")))
        row.comment_entry = comment_entry
        _prepare_comment_entry(comment_entry, key)

        # Кнопки действий в хранилище
        btn_return = tb.Button(
            row,
            text="В работу",
            width=10,
            bootstyle="secondary",
            command=lambda k=key: return_to_work(k),
        )
        btn_return.pack(side=LEFT, padx=2)

        btn_sklad = tb.Button(
            row,
            text="Склад",
            width=8,
            bootstyle="info",
            command=lambda k=key: mark_assembled(k),
        )
        btn_sklad.pack(side=LEFT, padx=2)

        # NEW: Кнопка-замок
        btn_lock = tb.Button(
            row,
            text="🔒",
            width=3,
            bootstyle=("info" if key in locked_storage else "secondary"),
            command=lambda k=key: toggle_storage_lock(k),
        )
        btn_lock.pack(side=LEFT, padx=2)

        tb.Button(
            row,
            text="✕",
            width=3,
            bootstyle="secondary",
            command=lambda k=key: delete_serp(k),
        ).pack(side=LEFT, padx=2)

        # Сохраняем ссылки для дальнейших обновлений
        # Сохраняем ссылки для дальнейших обновлений
        row.btn_return_to_work = btn_return
        row.btn_sklad = btn_sklad
        row.btn_lock = btn_lock

        # ВАЖНО: здесь НЕ вызываем update_storage_row_lock_state,
        # потому что row ещё не лежит в storage_widgets.
        return row

    # ---------- рабочие списки (ВС6Д/ВС7)
    is_draft = key in draft_products
    needed_blocks = block_types_for(key)
    all_blocks_effective = all(
        blocks.get(t, False) and (key, t) not in repair_blocks
        for t in needed_blocks
    )

    has_any_block = any(blocks.get(t, False) for t in needed_blocks)

    # Кнопка статуса: ЗАГОТОВКА/НЕ ГОТОВО
    # Кнопка статуса: ЗАГОТОВКА/НЕ ГОТОВО
    status_text, status_style = (
        ("ЗАГОТОВКА", "warning") if is_draft else ("НЕ СОБРАН", "danger")
    )
    status_btn = tb.Button(
        row,
        text=status_text,
        width=12,
        bootstyle=status_style,
        command=lambda k=key: toggle_draft_status(k),
        state="normal",  # <-- ВСЕГДА активно, без "disabled"
    )
    status_btn.pack(side=LEFT, padx=(0, 4))

#############

    # ---------- рабочие списки (ВС6Д/ВС7)
    # подпись и ширина берём из единого хелпера
    # Подпись изделия — ширину не задаём, делаем фиксированный пиксельный отступ справа
    label_text, _w_ignored = format_work_label_and_width(key, show_factory_code)

    num_lbl = tb.Label(
        row,
        text=label_text,
        font=FONT_PRIMARY,
        style="Main.TLabel",
        anchor="w",
    )
    num_lbl.pack(side=LEFT, padx=(0, 0))  # без внешнего отступа

    # Гуттер 12px между номером и первой кнопкой ССБ
    tk.Frame(
        row,
        width=GAP_AFTER_NUMBER_PX,
        height=1,  # >0, чтобы pack зарезервировал место
        bg=BG_MAIN,
        highlightthickness=0,
        bd=0,
    ).pack(side=LEFT)

    #######

    block_btns = []
    for block_type in needed_blocks:
        installed = blocks.get(block_type, False)
        in_repair = (key, block_type) in repair_blocks
        if in_repair:
            btn_text = REPAIR_TEXT
            btn_style = REPAIR_BOOTSTYLE
        else:
            btn_text = block_type
            if all_blocks_effective or installed:
                btn_style = "success"
            else:
                # up_112-подсказка только для ВС6Д
                if (
                        block_type == "ССБ 112"
                        and get_variant(key) == "ВС6Д"
                        and key2to3(key) in up112_hints
                ):
                    btn_style = "info"
                else:
                    btn_style = "secondary"
        block_btn = tb.Button(
            row,
            text=btn_text,
            width=8,
            bootstyle=btn_style,
            command=lambda k=key, b=block_type: on_block_left_click(k, b),
        )
        block_btn.pack(side=LEFT, padx=1)
        block_btn.bind(
            "<Button-3>", lambda e, k=key, b=block_type: toggle_block_repair(k, b)
        )
        block_btns.append(block_btn)
        try:
            if str(block_type).strip() in ("ССБ 112", "ССБ112", "SSB 112", "SSB112") \
                    and str(get_variant(key)).strip() in ("ВС7", "ВС13"):
                attach_vs7_112_tooltip(block_btn, key)
        except Exception:
            pass
    row.block_btns = block_btns

    # Обёртка с тонкой рамкой вокруг Entry
    comment_box = _wrap_with_border(row)
    comment_box.pack(side=LEFT, padx=10)

    comment_entry = tb.Entry(
        comment_box, width=12, font=FONT_PRIMARY, bootstyle="light"
    )
    comment_entry.pack(fill=X, padx=4, pady=3)  # небольшой внутренний отступ
    comment_entry.key_reference = key
    comment_entry.insert(0, current_comments.get(key, comments.get(key, "")))
    row.comment_entry = comment_entry
    # общая подготовка: серый текст + показывать начало строки
    _prepare_comment_entry(comment_entry, key)

    if all_blocks_effective:
        check_btn = tb.Button(
            row,
            text="✓",
            width=3,
            bootstyle="info",
            command=lambda k=key: mark_storage(k),
        )
    else:
        check_btn = tb.Button(
            row, text="✓", width=3, bootstyle="light", state="disabled"
        )
        check_btn.bind(
            "<ButtonPress-1>",
            lambda e, k=key, bbs=block_btns, cb=check_btn: start_long_press(
                e, k, bbs, cb
            ),
        )
        check_btn.bind("<ButtonRelease-1>", lambda e, k=key: cancel_long_press(e, k))
    check_btn.pack(side=LEFT, padx=2)
    row.check_btn = check_btn

    tb.Button(
        row,
        text="✕",
        width=3,
        bootstyle="secondary",
        command=lambda k=key: delete_serp(k),
    ).pack(side=LEFT, padx=2)

    return row


def start_long_press(event, key, block_btns, check_btn):
    global long_press_timers
    if key in long_press_timers:
        root.after_cancel(long_press_timers[key])
    long_press_timers[key] = root.after(
        750, lambda k=key, bbs=block_btns, cb=check_btn: complete_long_press(k, bbs, cb)
    )


def cancel_long_press(event, key):
    global long_press_timers
    if key in long_press_timers:
        root.after_cancel(long_press_timers[key])
        del long_press_timers[key]


def complete_long_press(key, block_btns, check_btn):
    try:
        if get_variant(key) == "ВС7":
            meta = vs7_112_meta.get(key) or {}
            last4 = meta.get("last4", "")
            if not (last4.isdigit() and len(last4) == 4):
                val = _ask_vs7_112_last4(key, prefill=last4)
                if not (val and val.isdigit() and len(val) == 4):
                    return  # без номера 112 ничего не ставим
                vs7_112_meta[key] = {"from_vs6d": False, "last4": val}
                save_data(False)
    except Exception:
        pass
    global long_press_timers
    if key in long_press_timers:
        del long_press_timers[key]
    for block_type in block_types_for(key):
        products[key][block_type] = True
    for btn in block_btns:
        btn.configure(bootstyle="success")
    check_btn.configure(
        bootstyle="info", state="normal", command=lambda k=key: mark_storage(k)
    )
    check_btn.unbind("<ButtonPress-1>")
    check_btn.unbind("<ButtonRelease-1>")
    save_data()
    update_row_widgets(key)


def save_scroll_positions():
    global work_scroll_position, work7_scroll_position, storage_scroll_position
    global work_top_element, work7_top_element, storage_top_element

    if globals().get("work_canvas") and work_canvas.winfo_exists():
        work_scroll_position = work_canvas.yview()[0]
    if globals().get("work7_canvas") and work7_canvas.winfo_exists():
        work7_scroll_position = work7_canvas.yview()[0]
    if globals().get("storage_canvas") and storage_canvas.winfo_exists():
        storage_scroll_position = storage_canvas.yview()[0]

    work_top_element = (
        get_top_visible_element(work_frame) if globals().get("work_frame") else None
    )
    work7_top_element = (
        get_top_visible_element(work7_frame) if globals().get("work7_frame") else None
    )
    storage_top_element = (
        get_top_visible_element(storage_frame)
        if globals().get("storage_frame")
        else None
    )


def get_top_visible_element(parent_frame):
    """Возвращает key верхнего видимого элемента для заданного фрейма,
    либо None, если фрейм пуст или не относится к известным канвасам."""
    if not parent_frame or not parent_frame.winfo_children():
        return None

    # Выбираем корректный canvas
    if parent_frame is work_frame:
        canvas = work_canvas
    elif parent_frame is work7_frame:
        canvas = work7_canvas
    elif parent_frame is storage_frame:
        canvas = storage_canvas
    else:
        # для чужих фреймов (например, архивного списка) ничего не делаем
        return None

    y_top, y_bottom = canvas.canvasy(0), canvas.canvasy(canvas.winfo_height())
    for widget in parent_frame.winfo_children():
        if isinstance(widget, tb.Frame) and hasattr(widget, "key"):
            widget_y, widget_h = widget.winfo_y(), widget.winfo_height()
            if widget_y + widget_h > y_top and widget_y < y_bottom:
                return widget.key
    return None


def restore_scroll_positions():
    if globals().get("work_canvas") and work_canvas.winfo_exists():
        work_canvas.yview_moveto(work_scroll_position)
    if globals().get("work7_canvas") and work7_canvas.winfo_exists():
        work7_canvas.yview_moveto(work7_scroll_position)
    if globals().get("storage_canvas") and storage_canvas.winfo_exists():
        storage_canvas.yview_moveto(storage_scroll_position)

    if globals().get("work_top_element") and work_top_element:
        scroll_to_element(work_frame, work_top_element)
    if globals().get("work7_top_element") and work7_top_element:
        scroll_to_element(work7_frame, work7_top_element)
    if globals().get("storage_top_element") and storage_top_element:
        scroll_to_element(storage_frame, storage_top_element)


def scroll_to_element(parent_frame, element_key):
    """Прокрутка до строки с указанным key в нужном канвасе."""
    if not parent_frame or element_key is None:
        return

    # Выбираем корректный canvas
    if parent_frame is work_frame:
        canvas = work_canvas
    elif parent_frame is work7_frame:
        canvas = work7_canvas
    elif parent_frame is storage_frame:
        canvas = storage_canvas
    else:
        return  # нечего скроллить

    parent_frame.update_idletasks()
    for widget in parent_frame.winfo_children():
        if (
            isinstance(widget, tb.Frame)
            and hasattr(widget, "key")
            and widget.key == element_key
        ):
            try:
                canvas.yview_moveto(
                    widget.winfo_y() / max(1, parent_frame.winfo_height())
                )
            except Exception:
                pass
            break


def update_row_widgets(key):
    # обновляем строку в том канвасе, где она есть (в левом и/или правом)
    rows = []
    if key in work6_widgets:
        rows.append(work6_widgets[key])
    if key in work7_widgets:
        rows.append(work7_widgets[key])

    if not rows:
        return

    blocks = products.get(key, {})
    needed_blocks = block_types_for(key)

    for row in rows:
        for i, block_type in enumerate(needed_blocks):
            installed = blocks.get(block_type, False)
            in_repair = (key, block_type) in repair_blocks

            if in_repair:
                btn_text = REPAIR_TEXT
                btn_style = REPAIR_BOOTSTYLE
            else:
                btn_text = block_type
                # up_112-подсветка — только для ВС6Д, только когда блок ещё не установлен
                # подсветка info: старый up112 (только 112) + новый XL (112/114/161)
                # подсветка info: старый up112 + новый XL (112/114/161)
                # подсветка: XL (info/danger) + старый up112 (только info для 112)
                info_tint = False
                try:
                    info_tint = (
                            block_type == "ССБ 112"
                            and get_variant(key) == "ВС6Д"
                            and not installed
                            and key2to3(key) in up112_hints
                    )
                except Exception:
                    info_tint = False

                # XL-подсказки
                if _xl_hint_on(key, block_type, installed):
                    info_tint = True

                btn_style = ("success" if installed else ("info" if info_tint else "secondary"))

            btn = row.block_btns[i]
            btn.configure(
                text=btn_text,
                bootstyle=btn_style,
                command=lambda k=key, b=block_type: on_block_left_click(k, b),
            )
            try:
                btn.unbind("<Button-3>")
            except Exception:
                pass
            btn.bind(
                "<Button-3>", lambda e, k=key, b=block_type: toggle_block_repair(k, b)
            )
        try:
            if str(block_type).strip() in ("ССБ 112", "ССБ112", "SSB 112", "SSB112") \
                    and str(get_variant(key)).strip() in ("ВС7", "ВС13"):
                attach_vs7_112_tooltip(btn, key)
        except Exception:
            pass
        all_blocks_effective = all(
            blocks.get(t, False) and (key, t) not in repair_blocks
            for t in needed_blocks
        )
        check_btn = row.check_btn
        if all_blocks_effective:
            check_btn.configure(
                bootstyle="info", state="normal", command=lambda k=key: mark_storage(k)
            )
            check_btn.unbind("<ButtonPress-1>")
            check_btn.unbind("<ButtonRelease-1>")
        else:
            check_btn.configure(bootstyle="light", state="disabled", command=None)
            check_btn.unbind("<ButtonPress-1>")
            check_btn.unbind("<ButtonRelease-1>")
            check_btn.bind(
                "<ButtonPress-1>",
                lambda e, k=key, bbs=row.block_btns, cb=check_btn: start_long_press(
                    e, k, bbs, cb
                ),
            )
            check_btn.bind("<ButtonRelease-1>", lambda e, k=key: cancel_long_press(e, k))

        status_btn = row.winfo_children()[0]

        # есть ли хотя бы один установленный блок
        has_any = any(blocks.get(t, False) for t in needed_blocks)

        try:
            status_btn.unbind("<Button-3>")
        except Exception:
            pass

        if key in draft_products:
            # ЗАГОТОВКА — можно снять всегда
            status_btn.configure(
                text="ЗАГОТОВКА",
                bootstyle="warning",
                command=lambda k=key: toggle_draft_status(k),
                state="normal",
            )
        else:
            # НЕ ГОТОВО — ВСЕГДА красная и активная (клик при 0 блоков ничего не меняет)
            status_btn.configure(
                text="НЕ СОБРАН",
                bootstyle="danger",
                command=lambda k=key: toggle_draft_status(k),
                state="normal",  # <-- убрали условие по has_any
            )


def _get_current_comments():
    current = {}
    frames = []
    wf = globals().get("work_frame")   # ВС6Д (левый)
    wf7 = globals().get("work7_frame") # ВС7 (правый)
    sf = globals().get("storage_frame")# хранилище (когда открыто)

    if wf is not None:
        frames.append(wf)
    if wf7 is not None:
        frames.append(wf7)
    if sf is not None:
        frames.append(sf)

    for frame in frames:
        try:
            for widget in frame.winfo_children():
                if isinstance(widget, tb.Frame) and hasattr(widget, "key"):
                    for child in widget.winfo_children():
                        if isinstance(child, tb.Entry) and hasattr(child, "key_reference"):
                            k = child.key_reference
                            # ВАЖНО: если изделие уже в storage_products,
                            # то игнорируем текст из рабочих колонок (wf, wf7),
                            # иначе «С хранилища» прилипнет обратно.
                            if k in storage_products and frame in (wf, wf7):
                                continue
                            current[k] = child.get()
        except Exception:
            pass
    return current


def _ensure_header(parent, hid, text, style, repack: bool = False):
    w = header_widgets.get(hid)
    if w is None or not w.winfo_exists():
        w = tb.Label(parent, text=text, style=style)
        w._vid = hid
        w.is_header = True
        header_widgets[hid] = w
        if "FactoryHeader" in str(style):
            w.pack(anchor="w", pady=(10, 0))
        else:
            w.pack(anchor="w", padx=10, pady=(10, 5))
    else:
        if str(w.cget("text")) != text:
            w.config(text=text)
        try:
            if str(w.cget("style")) != style:
                w.config(style=style)
        except Exception:
            pass
        if repack:
            try:
                w.pack_forget()
            except Exception:
                pass
            try:
                style_name = str(w.cget("style"))
            except Exception:
                style_name = ""
            if "FactoryHeader" in style_name:
                w.pack(anchor="w", pady=(10, 0))
            else:
                w.pack(anchor="w", padx=10, pady=(10, 5))
    return w


def _ensure_row(
    parent,
    key,
    blocks,
    current_comments,
    in_storage,
    show_factory_code=False,
    which="work6",
    repack=True,
):
    # выбираем словарь для виджетов
    if in_storage:
        src_dict = storage_widgets
        other_dicts = [work6_widgets, work7_widgets]
    else:
        if which == "work7":
            src_dict = work7_widgets
            other_dicts = [work6_widgets, storage_widgets]
        else:
            src_dict = work6_widgets
            other_dicts = [work7_widgets, storage_widgets]

    ##
    v = get_variant(key)
    y2 = get_year2(key)

    FS = 1 if (not in_storage and display_full_serials) else 0
    FSs = 1 if (in_storage and display_full_serials_storage) else 0

    # ✅ важно: при переключении hide_added_date хотим пересоздавать строки (иначе порядок "залипает")
    H = 1 if (not in_storage and globals().get("hide_added_date", False)) else 0

    desired_vid = (
        f"R:{'S' if in_storage else which}:{key[0]}:{key[1]}:{v}:{y2}"
        f":F{1 if show_factory_code else 0}:FS{FS}:FSs{FSs}:H{H}"
    )

    ###
    row = src_dict.get(key)

    # удалить клоны в других словарях
    for od in other_dicts:
        if key in od:
            try:
                od[key].destroy()
            except Exception:
                pass
            od.pop(key, None)

    # если есть, но «старого вида» — пересоздаём
    if row is not None and row.winfo_exists():
        if getattr(row, "_vid", None) != desired_vid:
            try:
                row.destroy()
            except Exception:
                pass
            src_dict.pop(key, None)
            row = None

    created_now = False
    if row is None or not row.winfo_exists():
        row = create_product_row(
            parent, key, blocks, current_comments, in_storage, show_factory_code
        )
        row._vid = desired_vid
        src_dict[key] = row
        created_now = True
    else:
        update_row_widgets(key)

    # ВАЖНО: только здесь row уже лежит в storage_widgets,
    # поэтому теперь можно корректно применить состояние замка.
    if in_storage:
        try:
            update_storage_row_lock_state(key)
        except Exception:
            pass

    # КЛЮЧЕВОЕ: если мы НЕ хотим глобальной перегруппировки...
    if created_now and not repack:
        _pack_row_after_factory(parent, row, key, which if not in_storage else "storage")

    return row

def apply_vs6d_only_mode_ui():
    apply_vs6d_mode_visual()

def refresh_sort_buttons():
    """Подсветка активной кнопки сортировки."""
    try:
        for mode, btn in sort_buttons.items():
            btn.configure(
                bootstyle=("primary" if work_sort_mode == mode else "secondary")
            )
    except Exception:
        pass

sort_mode_button = None  # глобальный хэндл новой кнопки (серой)


def update_sort_mode_button_visual():
    """
    Кнопка всегда называется 'Сортировка'.
    Цвет показывает режим:
      - factories -> secondary (серая)
      - drafts   -> warning   (жёлтая)
    """
    if not sort_mode_button or not sort_mode_button.winfo_exists():
        return

    try:
        # режим определяем по work_sort_mode
        if work_sort_mode == "drafts":
            sort_mode_button.configure(text="Сортировка", bootstyle="info")
        else:
            sort_mode_button.configure(text="Сортировка", bootstyle="secondary")
    except Exception:
        pass


def cycle_sort_mode():
    """
    Переключаем только два режима:
      factories <-> drafts
    Режим 'blocks' убран.
    """
    if work_sort_mode == "drafts":
        set_sort_mode("factories")
    else:
        set_sort_mode("drafts")
    try:
        print("SORT:", work_sort_mode,
              "xl112=", len(xl_hints_112),
              "xl114=", len(xl_hints_114),
              "xl161=", len(xl_hints_161))
    except Exception as e:
        print("SORT debug error:", e)

    # На всякий случай обновим цвет кнопки сразу
    update_sort_mode_button_visual()

def refresh_lists():
    save_all_comments()
    update_product_list(preserve_scroll=False, regroup=True)

def set_sort_mode(mode: str):
    """
    Устанавливаем режим сортировки для рабочих колонок.
    Допустимые режимы:
      - 'factories' : завод -> дата
      - 'drafts'    : заготовки/ремонт/остальные

    Режим 'blocks' удалён. blockcount_mode принудительно выключаем,
    чтобы кнопка 'Сортировка' не попадала в этот режим.
    """
    global work_sort_mode, show_draft_group, blockcount_mode

    if mode not in ("drafts", "factories"):
        return

    work_sort_mode = mode

    # Ваша логика в _desired_work_sequence_variant использует show_draft_group / blockcount_mode
    show_draft_group = (mode == "drafts")
    blockcount_mode = False   # <-- важно: кнопка сортировки больше не включает режим "по блокам"

    # Если у вас есть 🧩-кнопка — пусть её подсветка тоже будет синхронизирована
    try:
        set_puzzle_button_visual()
    except Exception:
        pass

    # Если остались старые кнопки сортировки (на всякий случай)
    refresh_sort_buttons()

    save_data(False)
    update_product_list(preserve_scroll=False)

    # НОВОЕ: обновляем внешний вид кнопки "Сортировка"
    update_sort_mode_button_visual()


def _hide_model_buttons(hide: bool):
    """Скрыть/вернуть кнопки выбора модели ВС6Д/ВС7, сохранив их pack-параметры."""
    try:
        for btn in (vs6d_btn, vs13_btn):
            if not btn:
                continue
            if hide:
                if btn.winfo_manager():
                    btn._old_pack_info = btn.pack_info()
                    btn.pack_forget()
            else:
                if hasattr(btn, "_old_pack_info"):
                    btn.pack(**btn._old_pack_info)
    except Exception:
        pass


def _block_space_handler(e):
    """Глобальная блокировка ПРОБЕЛА вне полей ввода."""
    try:
        # В Entry/Text пробел оставляем для набора текста
        if isinstance(e.widget, (tk.Entry, tb.Entry, tk.Text)):
            return
    except Exception:
        pass
    return "break"


def _apply_space_blocking(on: bool):
    """Включить/выключить блокировку пробела."""
    try:
        if on:
            root.bind_all("<space>", _block_space_handler, add="+")
        else:
            # снимаем наш хэндлер (если у вас были другие бинды на пробел — они останутся)
            root.unbind_all("<space>")
    except Exception:
        pass


def open_storage_panel():
    """Принудительно открыть панель хранилища и обновить её кнопку-стрелку."""
    global storage_visible
    storage_visible = True
    try:
        # показать контейнер хранилища (у вас grid)
        storage_frame_container.grid()
    except Exception:
        pass
    try:
        # при открытом хранилище у вас текст стрелки "▶"
        toggle_storage_button.configure(text="▶", width=2)
    except Exception:
        pass

def apply_vs6d_mode_visual():
    if vs6d_only_mode:
        # принудительно ВС6Д
        try:
            set_product_mode("ВС6Д")
        except Exception:
            global product_mode
            product_mode = "ВС6Д"

        # прячем переключатели модели
        _hide_model_buttons(True)

        # открыть хранилище и заблокировать стрелку, блокируем пробел и т.д.
        open_storage_panel()
        try:
            if toggle_storage_button and toggle_storage_button.winfo_exists():
                toggle_storage_button.configure(state="disabled")
        except Exception:
            pass
        _apply_space_blocking(True)

    else:
        # вернуть переключатели модели
        _hide_model_buttons(False)

        # включить стрелку, вернуть поведение пробела и т.п.
        try:
            if toggle_storage_button and toggle_storage_button.winfo_exists():
                toggle_storage_button.configure(state="normal")
        except Exception:
            pass
        _apply_space_blocking(False)
        try:
            root.bind("<space>", _on_space_toggle_storage)
        except Exception:
            pass

    # без сдвига прокрутки обновим списки
    try:
        update_product_list(preserve_scroll=True)
    except Exception:
        pass



def on_toggle_vs6d_mode():
    global vs6d_only_mode
    vs6d_only_mode = bool(vs6d_only_mode_var.get())
    save_data(False)
    apply_vs6d_mode_visual()   # <-- вместо apply_vs6d_only_mode_ui()



# --- коммент: оформление и поведение поля комментария ---
COMMENT_FG = "#444444"        # тёмно-серый текст
COMMENT_BORDER = "#cfd4da"    # цвет тонкой рамки

def _wrap_with_border(parent):
    """
    Возвращает tk.Frame-обёртку с тонкой серой рамкой под Entry.
    """
    box = tk.Frame(parent,
                   highlightthickness=1,
                   highlightbackground=COMMENT_BORDER,
                   bd=0,
                   bg=BG_WORK_AREA)
    return box



COMMENT_FG = "#444444"        # тёмно-серый текст
COMMENT_BORDER = "#cfd4da"    # цвет тонкой рамки (если используете _wrap_with_border)

def _prepare_comment_entry(entry, key):
    """
    Настройки для поля комментария:
    - тёмно-серый текст;
    - нормальная печать во время ввода;
    - после завершения ввода (Enter/потеря фокуса) прокрутка вида к началу.
    """
    try:
        entry.configure(foreground=COMMENT_FG)
    except Exception:
        pass

    def _snap_to_start():
        try:
            # показываем левый край текста
            entry.xview_moveto(0)
            # НЕ трогаем каретку во время ввода (icursor(0) не ставим!)
        except Exception:
            pass

    # Сохраняем и показываем начало ТОЛЬКО при завершении ввода:
    entry.bind("<FocusOut>", lambda e, k=key: (comments.update({k: entry.get()}), _snap_to_start()), add="+")
    entry.bind("<Return>",   lambda e, k=key: (comments.update({k: entry.get()}), root.focus(), _snap_to_start()), add="+")

    # При создании — сразу показываем начало (для длинных уже сохранённых комментариев)
    entry.after_idle(_snap_to_start)

def get_year2(key) -> str:
    """Вернуть 2-значный год для изделия: из ключа (4-й элемент) или из assembly_years."""
    try:
        if isinstance(key, tuple) and len(key) >= 4 and isinstance(key[3], str):
            return key[3]
    except Exception:
        pass
    return (assembly_years.get(key) or "??")[:2]

def key2to3(key):
    """3-элементный вариант ключа (номер, завод, вариант) — полезно для up112_hints."""
    if isinstance(key, tuple) and len(key) >= 3:
        return (key[0], key[1], key[2])
    return key

def serial_for_key(key) -> str:
    """Полный серийник: yy + factory_code + 4-значный номер."""
    number, factory_name, _ = _split_key_safe(key)
    factory_code = factory_reverse_mapping.get(factory_name, "?")
    return f"{get_year2(key)}{factory_code}{number.zfill(4)}"

def serial_parts_from_key(key):
    """Вернёт ('гг', 'z', 'xxxx') для ключа."""
    number, factory_name, _ = _split_key_safe(key)
    yy = get_year2(key)
    z  = factory_reverse_mapping.get(factory_name, "?")
    num4 = number.zfill(4)
    return yy, z, num4

# Возвращает "12yyzXXXX" для ВС7, где XXXX берём из vs7_112_meta (или ????)
def _vs7_112_from_meta(key):
    yy, z, _ = serial_parts_from_key(key)

    # пытаемся достать по tuple-ключу и (на всякий случай) по строковому
    meta = {}
    try:
        meta = (vs7_112_meta.get(key)
                or vs7_112_meta.get("|".join(map(str, key)))  # если ключи сериализуются строкой
                or {})
    except Exception:
        pass

    last4 = meta.get("last4", "")
    if not (isinstance(last4, str) and last4.isdigit() and len(last4) == 4):
        last4 = "????"
    return f"12{yy}{z}{last4}"

def ssb_serial_auto(key, block_type):
    """
    Вернёт автосерийник ССБ по серийному номеру СЕРПа.
    ВС6Д: 112=12yyzxxxx, 114=14yyzxxxx, 161=61yyzxxxx
    ВС7 : 116=16yyzxxxx, 112=12yyzXXXX (XXXX берём из vs7_112_meta или '????')
    """
    yy, z, num4 = serial_parts_from_key(key)  # 'yy', 'z', 'xxxx'
    variant = get_variant(key)
    bt = str(block_type).strip()

    if variant == "ВС6Д":
        if bt == "ССБ 112": return f"12{yy}{z}{num4}"
        if bt == "ССБ 114": return f"14{yy}{z}{num4}"
        if bt == "ССБ 161": return f"61{yy}{z}{num4}"
        return ""
    elif variant == "ВС7":
        if bt == "ССБ 116": return f"16{yy}{z}{num4}"
        if bt == "ССБ 112": return _vs7_112_from_meta(key)  # <-- вот тут главное изменение
        return ""
    return ""



def serial_display(key, sep=" "):
    """Вернёт строку серийника с разделителем между частями: 'гг z xxxx' при sep=' '."""
    yy, z, num4 = serial_parts_from_key(key)
    return f"{yy}{sep}{z}{sep}{num4}"


def format_work_label_and_width(key, show_factory_code: bool):
    """
    Подпись изделия в колонках ВС6Д/ВС7 (НЕ хранилище) + целевая ширина label.
    Возвращает (text, width).
      - display_full_serials=True -> "№ ггzxxxx", width = -1 (НЕ задаём фикс. ширину)
      - иначе:
          * режим "Заводы": "<модель> № xxxx"
          * режим "Заготовки/Блоки": "<модель> xxxx [z]"
    """
    variant = get_variant(key)
    number, factory_name, _ = _split_key_safe(key)
    factory_code = factory_reverse_mapping.get(factory_name, "?")

    if display_full_serials:
        # компактная подпись «полный серийник» с разбивкой: "гг z xxxx"
        text = f"№ {serial_display(key)}"
        # ширина для лейбла теперь игнорируется в create_product_row, но возвращаем что-то нейтральное
        return text, -1

    # --- старый (смешанный) режим ---
    pref = get_product_prefix_from_key(key, compact=True)  # "ВС6Д" / "ВС7"

    if work_sort_mode == "factories":
        # "<модель> № xxxx"
        text = f"{pref} №{number}"
        WIDTH_NO_CODE = {"ВС7": 11, "ВС6Д": 12}
        WIDTH_WITH_CODE = {"ВС7": 13, "ВС6Д": 14}
        w = (WIDTH_WITH_CODE if show_factory_code else WIDTH_NO_CODE).get(variant, 12)
        return text, w

    # "Заготовки"/"Блоки": "<модель> xxxx [z]"
    text = f"{pref} {number} [{factory_code}]"
    WIDTH_BLOCKS = {"ВС7": 13, "ВС6Д": 14}
    return text, WIDTH_BLOCKS.get(variant, 14)

def on_year_change(event=None):
    """Сохраняем выбранный год в settings."""
    global last_selected_year
    try:
        last_selected_year = year_var.get().strip()
    except Exception:
        return
    save_data(False)

def migrate_keys_to_4():
    """
    Мягкая миграция: добавляем год в ключ (номер, завод, вариант, yy).
    Старые данные с ключами длины 2/3 переводим в 4.
    Без изменения остальной логики.
    """
    global products, assembled_products, storage_products
    global draft_products, redy_products, comments, product_dates
    global assembly_dates, storage_dates, assembly_years, marked_statuses
    global up112_hints, repair_blocks
    global locked_storage

    # уже мигрировано?
    all_prod4 = all(isinstance(k, tuple) and len(k) >= 4 for k in products.keys())
    all_arch4 = all(isinstance(k, tuple) and len(k) >= 4 for k in assembled_products)
    all_stor4 = all(isinstance(k, tuple) and len(k) >= 4 for k in storage_products)
    all_rep4 = (not repair_blocks) or all(isinstance(k, tuple) and len(k) >= 4 for (k, _) in repair_blocks)

    if all_prod4 and all_arch4 and all_stor4 and all_rep4:
        return

    # снимок годов по 3-ключам
    ay3 = {}
    for k, v in assembly_years.items():
        kk = key2to3(k)
        ay3[kk] = (v or "??")[:2]

    def add_year(k):
        if isinstance(k, tuple) and len(k) >= 4:
            return k
        if isinstance(k, tuple) and len(k) >= 3:
            yy = ay3.get((k[0], k[1], k[2])) or "??"
            return (k[0], k[1], k[2], yy)
        if isinstance(k, tuple) and len(k) == 2:
            # по умолчанию ВС6Д, как и в миграции 2->3
            base3 = (k[0], k[1], "ВС6Д")
            yy = ay3.get(base3) or "??"
            return (k[0], k[1], "ВС6Д", yy)
        return k

    def map_dict(d):
        out = {}
        for k, v in d.items():
            out[add_year(k)] = v
        return out

    def map_set(s):
        return set(add_year(k) for k in s)

    products         = map_dict(products)
    comments         = map_dict(comments)
    product_dates    = map_dict(product_dates)
    assembly_dates   = map_dict(assembly_dates)
    storage_dates    = map_dict(storage_dates)
    assembly_years   = {add_year(k): (v or "??")[:2] for k, v in assembly_years.items()}
    marked_statuses  = map_dict(marked_statuses)

    assembled_products = map_set(assembled_products)
    storage_products   = map_set(storage_products)
    draft_products     = map_set(draft_products)
    redy_products      = map_set(redy_products)
    locked_storage   = map_set(locked_storage)


    # подсказки up112 храним по 3-ключам — оставляем как есть

    # ремонтные отметки: (key, block_type)
    if repair_blocks:
        repair_blocks = set((add_year(k), b) for (k, b) in repair_blocks)

    # финально сохраним в новом формате
    save_data(False)




def _ru_plural(n: int, forms=("блок", "блока", "блоков")) -> str:
    """
    Возвращает корректную форму слова 'блок' для русского языка:
    1 блок, 2 блока, 5 блоков и т.п.
    """
    n = abs(n) % 100
    n1 = n % 10
    if 11 <= n <= 14:
        return forms[2]
    if n1 == 1:
        return forms[0]
    if 2 <= n1 <= 4:
        return forms[1]
    return forms[2]


def _header_text_for_count(count: int, need: int) -> str:
    """
    Заголовок группы в режиме 'по количеству блоков'.
    Если count == need -> 'Готовы к сборке'
    Иначе -> '<count> блок/блока/блоков в наличии'
    """
    if count == need:
        return "Готовы к сборке"
    return f"{count} {_ru_plural(count)} в наличии"


def _desired_work_sequence_variant(current_comments, variant: str):
    """
    Унифицированная версия для ВС6Д и ВС7.
    Режимы:
      - blockcount_mode: группы по числу установленных необходимых блоков (0..need)
      - show_draft_group: «Заготовки» / «В ремонте» / «Остальные»
      - базовый режим: завод -> дата
    Везде считаем только те блоки, которые требуются для данного варианта (block_types_for).
    """
    seq = []

    # Собираем только этот вариант и только те, что «в работе»
    work_grouped = {factory: [] for factory in factory_order}
    for key, blocks in products.items():
        if get_variant(key) != variant:
            continue
        if key in assembled_products or key in storage_products:
            continue
        work_grouped[key[1]].append((key, blocks))

    sort_reverse = sort_order == "new_first"

    # === ВАЖНО: ремонт только по блокам, которые реальны для ЭТОГО варианта ===
    # (иначе ключ окажется «в ремонте», хотя видимых ремонтных кнопок нет)
    repair_keys = set()
    for rk, rb in repair_blocks:
        if get_variant(rk) != variant:
            continue
        if rb in block_types_for(rk):
            repair_keys.add(rk)

    # ------------------------ 1) Режим «по количеству блоков» ------------------------
    if blockcount_mode:
        flat_items = [
            item for factory in factory_order for item in work_grouped[factory]
        ]

        # Максимум «нужных» блоков для этого варианта (обычно ВС6Д=3, ВС7=2)
        max_needed = 0
        for key, blocks in flat_items:
            _, need = installed_count_for(key, blocks)
            if need > max_needed:
                max_needed = need

        if max_needed == 0:
            return seq  # на всякий случай

        # Группы по count установленным нужным блокам
        groups = {i: [] for i in range(0, max_needed + 1)}
        for key, blocks in flat_items:
            cnt, _need = installed_count_for(key, blocks)
            groups[cnt].append((key, blocks))

        # Выводим от полностью готовых к нулю
        for count in range(max_needed, -1, -1):
            items = groups[count]
            if not items:
                continue

            # внутри группы: сначала заготовки, затем дата, затем номер
            items.sort(
                key=lambda item: (
                    0 if item[0] in draft_products else 1,
                    parse_date_str(product_dates.get(item[0], "01.01.00"), item[0]),
                    item[0][0],
                )
            )

            header_text = (
                "Готовы к сборке"
                if count == max_needed
                else f"{count} " + _ru_plural(count) + " в наличии"
            )

            seq.append(
                {
                    "kind": "header",
                    "id": f"W{variant}:grp:{count}",
                    "text": f"{header_text} ({len(items)})",
                    "style": "Header3.TLabel",
                }
            )

            for key, blocks in items:
                seq.append(
                    {
                        "kind": "row",
                        "key": key,
                        "blocks": blocks,
                        "in_storage": False,
                        "show_factory_code": True,
                    }
                )

        return seq

    # ----------- 2) Режим «Заготовки / Ремонт / Остальные» -----------
    if show_draft_group:
        # --- Заготовки (только если НЕ в ремонте по «нужным» блокам) ---
        drafts = []

        for factory in factory_order:
            for key, blocks in work_grouped[factory]:
                if key in draft_products and key not in repair_keys:
                    drafts.append((key, blocks))

        if drafts:
            seq.append(
                {"kind": "header",
                 "id": f"W{variant}:hdr:drafts",
                 "text": f"Заготовки ({len(drafts)})",
                 "style": "SectionHeader.TLabel"}
            )
            # Группировка заготовок по числу установленных нужных блоков
            max_needed = 0
            for key, blocks in drafts:
                _, need = installed_count_for(key, blocks)
                max_needed = max(max_needed, need)

            draft_groups = {i: [] for i in range(0, max_needed + 1)}
            for key, blocks in drafts:
                cnt, _need = installed_count_for(key, blocks)
                draft_groups[cnt].append((key, blocks))

            for cnt in range(max_needed, -1, -1):
                bucket = draft_groups[cnt]
                if not bucket:
                    continue
                header_text = "Готовы к сборке" if cnt == max_needed else f"{cnt} " + _ru_plural(cnt) + " в наличии"
                seq.append(
                    {"kind": "header",
                     "id": f"W{variant}:hdr:drafts:{cnt}",
                     "text": f"{header_text} ({len(bucket)})",
                     "style": "Header3.TLabel"}
                )
                bucket.sort(
                    key=lambda item: (
                        parse_date_str(product_dates.get(item[0], "01.01.00"), item[0]),
                        item[0][0],
                    )
                )
                for key, blocks in bucket:
                    seq.append(
                        {"kind": "row", "key": key, "blocks": blocks,
                         "in_storage": False, "show_factory_code": True}
                    )

        # --- Остальные: группировка по числу установленных нужных блоков ---
        others = []
        for factory in factory_order:
            for key, blocks in work_grouped[factory]:
                if key not in draft_products and key not in repair_keys:
                    others.append((key, blocks))

        if others:
            seq.append(
                {"kind": "header",
                 "id": f"W{variant}:hdr:others",
                 "text": f"Остальные ({len(others)})",
                 "style": "SectionHeader.TLabel"}
            )

            # сколько блоков «надо» для варианта (ВС6Д=3, ВС7=2) берём из данных
            max_needed = 0
            for key, blocks in others:
                _, need = installed_count_for(key, blocks)
                if need > max_needed:
                    max_needed = need

            groups = {i: [] for i in range(0, max_needed + 1)}
            for key, blocks in others:
                cnt, _need = installed_count_for(key, blocks)
                groups[cnt].append((key, blocks))

            # вывод: от готовых к нулю
            for cnt in range(max_needed, -1, -1):
                bucket = groups[cnt]
                if not bucket:
                    continue

                header_text = "Готовы к сборке" if cnt == max_needed else f"{cnt} " + _ru_plural(cnt) + " в наличии"
                seq.append(
                    {"kind": "header",
                     "id": f"W{variant}:hdr:others:{cnt}",
                     "text": f"{header_text} ({len(bucket)})",
                     "style": "Header3.TLabel"}
                )

                # внутри группы: по дате добавления, затем по номеру
                bucket.sort(
                    key=lambda item: (
                        parse_date_str(product_dates.get(item[0], "01.01.00"), item[0]),
                        item[0][0],
                    )
                )
                for key, blocks in bucket:
                    seq.append(
                        {"kind": "row",
                         "key": key,
                         "blocks": blocks,
                         "in_storage": False,
                         "show_factory_code": True}  # <- формат xxxx [z]
                    )


        # --- В ремонте (ВСЕГДА в самом конце) ---
        repair_items = []
        for factory in factory_order:
            for key, blocks in work_grouped[factory]:
                if key in repair_keys:
                    repair_items.append((key, blocks))

        if repair_items:
            seq.append(
                {"kind": "header",
                 "id": f"W{variant}:hdr:repair",
                 "text": f"В ремонте ({len(repair_items)})",
                 "style": "SectionHeader.TLabel"}
            )
            repair_items.sort(
                key=lambda item: (
                    parse_date_str(product_dates.get(item[0], "01.01.00"), item[0]),
                    item[0][0],
                )
            )
            for key, blocks in repair_items:
                seq.append(
                    {"kind": "row", "key": key, "blocks": blocks,
                     "in_storage": False, "show_factory_code": True}
                )
        return seq

    # ------------------------ 3) Базовый режим: завод -> дата ------------------------
    # ------------------------ 3) Базовый режим: завод -> дата ------------------------
    for factory in factory_order:
        items = work_grouped[factory]
        if not items:
            continue

        # Заголовок завода
        seq.append(
            {
                "kind": "header",
                "id": f"W{variant}:hdr:factory:{factory}",
                "text": f"Завод: {factory}",
                "style": "FactoryHeader.TLabel",
            }
        )

        # ✅ НОВОЕ: если скрываем даты — просто выводим изделия по номерам (возрастание)
        if globals().get("hide_added_date", False):
            for key, blocks in sorted(items, key=lambda x: x[0][0]):  # x[0][0] = num4 ('0001'..)
                seq.append(
                    {
                        "kind": "row",
                        "key": key,
                        "blocks": blocks,
                        "in_storage": False,
                        "show_factory_code": False,
                    }
                )
            continue  # следующий завод

        # --- СТАРОЕ ПОВЕДЕНИЕ: группировка по датам добавления ---
        by_date = {}
        for key, blocks in items:
            d = product_dates.get(key, "??.??.??")
            by_date.setdefault(d, []).append((key, blocks))

        for d in sorted(
                by_date.keys(),
                key=lambda dd: parse_date_str(dd, by_date[dd][0][0]),
                reverse=sort_reverse,
        ):
            seq.append(
                {
                    "kind": "header",
                    "id": f"W{variant}:hdr:date:{factory}:{d}",
                    "text": format_date_genitive(d),
                    "style": "Header3.TLabel",
                }
            )
            for key, blocks in sorted(by_date[d], key=lambda x: x[0][0]):
                seq.append(
                    {
                        "kind": "row",
                        "key": key,
                        "blocks": blocks,
                        "in_storage": False,
                        "show_factory_code": False,
                    }
                )

    return seq


def normalize_repair_blocks():
    """Оставляем отметки ремонта только по 'нужным' для варианта блокам."""
    global repair_blocks
    cleaned = set()
    for k, b in repair_blocks:
        if b in block_types_for(k):
            cleaned.add((k, b))
    if cleaned != repair_blocks:
        repair_blocks = cleaned
        save_data(False)  # тихо сохраняем

def _desired_storage_sequence(current_comments):
    """Строим список для правой колонки «Ждут отправки на склад» с 2 режимами."""
    seq = []
    sort_reverse = (sort_order == "new_first")

    # ----- РЕЖИМ 1: По датам (как раньше) -----
    if storage_sort_mode == "dates":
        by_factory = {factory: [] for factory in factory_order}
        for key in storage_products:
            if key in assembled_products:
                continue
            by_factory[key[1]].append(key)

        for factory in factory_order:
            items = by_factory.get(factory, [])
            if not items:
                continue

            # Заголовок завода
            seq.append(
                {
                    "kind": "header",
                    "id": f"S:hdr:factory:{factory}",
                    "text": f"Завод: {factory}",
                    "style": "FactoryHeader.TLabel",
                }
            )

            # Внутри — группировка по дате сборки (или дате добавления)
            by_date = {}
            for key in items:
                date = storage_dates.get(key) or product_dates.get(key, "??.??.??")
                by_date.setdefault(date, []).append(key)

            for date in sorted(by_date.keys(), key=lambda d: parse_date_str(d), reverse=sort_reverse):
                seq.append(
                    {
                        "kind": "header",
                        "id": f"S:hdr:date:{factory}:{date}",
                        "text": f"Дата сборки: {format_ddmm(date)}",
                        "style": "Header3.TLabel",
                    }
                )
                for key in sorted(by_date[date], key=lambda k: k[0]):
                    seq.append(
                        {
                            "kind": "row",
                            "key": key,
                            "blocks": products.get(key, {}),
                            "in_storage": True,
                            "show_factory_code": False,
                        }
                    )
        return seq

    # ----- РЕЖИМ 2: По моделям -----
    # Собираем всё, что в хранилище и ещё не отправлено
    items = [k for k in storage_products if k not in assembled_products]

    # Группы по модели (нормализуем ВС13 -> ВС7)
    groups = {"ВС6Д": [], "ВС7": []}
    for key in items:
        v = str(get_variant(key)).strip()
        v_norm = "ВС7" if v in ("ВС7", "ВС13") else "ВС6Д"
        groups[v_norm].append(key)

    def _num(k):
        try:
            return int(_split_key_safe(k)[0])
        except Exception:
            return 0

    # Порядок секций: ВС6Д, затем ВС7 (как просили две шапки)
    for variant in ("ВС6Д", "ВС7"):
        bucket = groups[variant]
        if not bucket:
            continue

        seq.append({
            "kind": "header",
            "id": f"S:hdr:variant:{variant}",
            "text": f"{variant} ({len(bucket)})",
            "style": "SectionHeader.TLabel",   # крупнее и без «Завод/Дата»
        })

        # внутри — сортировка по номеру изделия возрастающе
        for key in sorted(bucket, key=_num):
            seq.append({
                "kind": "row",
                "key": key,
                "blocks": products.get(key, {}),
                "in_storage": True,
                "show_factory_code": True,     # нам нужен [z], но подпись выставим в refresh_storage_labels_in_place()
            })

    return seq


def _apply_sequence(frame, sequence, which, regroup=True, repack=True):
    current_comments = _get_current_comments()
    widgets_in_order, seen_vids = [], set()

    for item in sequence:
        if item["kind"] == "header":
            w = _ensure_header(frame, item["id"], item["text"], item["style"])
        else:
            w = _ensure_row(
                parent=frame,
                key=item["key"],
                blocks=item["blocks"],
                current_comments=current_comments,
                in_storage=item["in_storage"],
                show_factory_code=item.get("show_factory_code", False),
                which=which,
                repack=repack,
            )
        widgets_in_order.append(w)
        seen_vids.add(getattr(w, "_vid", None))

    # подчистка лишнего (один раз)
    for child in list(frame.winfo_children()):
        vid = getattr(child, "_vid", None)
        if vid and vid not in seen_vids:
            try:
                child.destroy()
            except Exception:
                pass

    # ВАЖНО:
    # - Для обычных списков уважаем флаг regroup (как было).
    # - Для ХРАНИЛИЩА принудительно выполняем перегруппировку,
    #   т.к. группы по датам без неё "залипают" под старым заголовком.
    if which != "storage" and not regroup:
        return

    # --- далее перестановка секций (как было) ---
    def is_section_header(vid: str):
        if not vid:
            return False
        if ":hdr:drafts" in vid:  return True
        if ":hdr:repair" in vid:  return True
        if ":hdr:others" in vid:  return True
        if vid.startswith("S:hdr:factory:"): return True     # секция хранилища по заводу
        if ":hdr:factory:" in vid: return True               # секции работы по заводу
        return False

    desired_groups = []
    cur = []
    for w in widgets_in_order:
        vid = getattr(w, "_vid", "")
        if getattr(w, "is_header", False) and is_section_header(vid):
            if cur:
                desired_groups.append(cur)
            cur = [w]
        else:
            (cur or (cur := [])).append(w)
    if cur:
        desired_groups.append(cur)

    def _pack(w, before=None):
        try:
            w.pack_forget()
        except Exception:
            pass
        if getattr(w, "is_header", False):
            style_name = ""
            try:
                style_name = str(w.cget("style"))
            except Exception:
                pass
            if "FactoryHeader" in style_name:
                kwargs = dict(anchor="w", pady=(10, 0))
            else:
                kwargs = dict(anchor="w", padx=10, pady=(10, 5))
        else:
            kwargs = dict(fill=X, pady=PADY_PRODUCT)
        if before is not None:
            w.pack(before=before, **kwargs)
        else:
            w.pack(**kwargs)

    def current_order():
        return [w for w in frame.winfo_children() if getattr(w, "_vid", None) in seen_vids]

    for i in range(len(desired_groups)):
        cur_order = current_order()
        try:
            start = cur_order.index(desired_groups[i][0])
            ok = all(
                start + j < len(cur_order)
                and cur_order[start + j] is desired_groups[i][j]
                for j in range(len(desired_groups[i]))
            )
        except ValueError:
            ok = False
        if ok:
            continue
        later_firsts = [g[0] for g in desired_groups[i + 1 :] if g and g[0] in cur_order]
        anchor = (min(later_firsts, key=lambda w: cur_order.index(w)) if later_firsts else None)
        for w in reversed(desired_groups[i]):
            _pack(w, before=anchor)
            anchor = w


def _capture_view(frame, canvas):
    """
    Снимок вида: (top_key, intra_offset, alt_key).
    Если сверху заголовок, берём ПЕРВУЮ строку ниже верхней кромки.
    """
    try:
        frame.update_idletasks()
        y_top = canvas.canvasy(0)

        # берём только РЯДЫ (Frame со свойством .key), заголовки игнорим
        rows = [w for w in frame.winfo_children()
                if isinstance(w, tb.Frame) and hasattr(w, "key")]
        if not rows:
            return (None, 0, None)

        # перехват: первая строка, которая хотя бы частично находится НИЖЕ верхней кромки
        for i, w in enumerate(rows):
            wy, wh = w.winfo_y(), w.winfo_height()
            if wy + wh > y_top:
                intra = max(0, int(y_top - wy))
                alt_key = rows[i+1].key if i+1 < len(rows) else (rows[i-1].key if i-1 >= 0 else None)
                return (w.key, intra, alt_key)

        # fallback: последняя строка (низ списка)
        last = rows[-1]
        wy, wh = last.winfo_y(), last.winfo_height()
        return (last.key, max(0, wh - 1), None)
    except Exception:
        return (None, 0, None)


def _restore_view(frame, canvas, top_key, intra_offset, alt_key=None):
    """
    Восстановление: если исходная строка исчезла — пробуем alt; если и её нет — прижимаемся к низу.
    """
    try:
        frame.update_idletasks()
        rows = [w for w in frame.winfo_children()
                if isinstance(w, tb.Frame) and hasattr(w, "key")]
        if not rows:
            canvas.yview_moveto(0.0)
            return

        keys = [w.key for w in rows]
        anchor = top_key if top_key in keys else (alt_key if alt_key in keys else rows[-1].key)

        target_y = None
        for w in rows:
            if w.key == anchor:
                target_y = w.winfo_y()
                break

        if target_y is None:
            canvas.yview_moveto(0.0)
        else:
            canvas.yview_moveto((target_y + max(0, int(intra_offset))) / max(1, frame.winfo_height()))
    except Exception:
        pass


def _capture_view(frame, canvas):
    """
    Снимок вида: (top_key, intra_offset, alt_key)
      top_key      — key верхней видимой строки
      intra_offset — смещение внутри этой строки (в пикселях)
      alt_key      — следующая видимая строка (на случай, если top_key исчезнет)
    """
    try:
        frame.update_idletasks()
        y_top = canvas.canvasy(0)
        children = [w for w in frame.winfo_children() if isinstance(w, tb.Frame)]
        top_key, alt_key = None, None
        intra = 0
        for i, w in enumerate(children):
            if hasattr(w, "key"):
                wy, wh = w.winfo_y(), w.winfo_height()
                if wy <= y_top < wy + wh:
                    top_key = w.key
                    intra = y_top - wy
                    # alt — следующая видимая вниз
                    for j in range(i + 1, len(children)):
                        w2 = children[j]
                        if hasattr(w2, "key"):
                            wy2, wh2 = w2.winfo_y(), w2.winfo_height()
                            if wy2 < y_top + canvas.winfo_height():
                                alt_key = w2.key
                                break
                    break
        return (top_key, int(intra), alt_key)
    except Exception:
        return (None, 0, None)


def _restore_view(frame, canvas, top_key, intra_offset, alt_key=None):
    """
    Восстановить вид по top_key+intra_offset; если top_key исчез — пробуем alt_key;
    если и он не найден — в начало.
    """
    try:
        frame.update_idletasks()
        target_key = top_key
        # если исходная верхняя строка уже исчезла — fallback
        keys = [getattr(w, "key", None) for w in frame.winfo_children()]
        if target_key not in keys and alt_key in keys:
            target_key = alt_key

        if target_key is None:
            canvas.yview_moveto(0.0)
            return

        target_y = None
        for w in frame.winfo_children():
            if isinstance(w, tb.Frame) and getattr(w, "key", None) == target_key:
                target_y = w.winfo_y()
                break

        if target_y is None:
            canvas.yview_moveto(0.0)
        else:
            canvas.yview_moveto((target_y + max(0, int(intra_offset))) / max(1, frame.winfo_height()))
    except Exception:
        pass

def _pack_row_after_factory(parent, row, key, which):
    """
    Вставить row в конец секции конкретного завода (перед следующим заголовком завода),
    не задевая глобальную перегруппировку.
    """
    factory = key[1]
    variant = get_variant(key)

    if which == "storage":
        factory_hid = f"S:hdr:factory:{factory}"
        is_factory_hdr = lambda w: getattr(w, "is_header", False) and ":hdr:factory:" in (getattr(w, "_vid", "") or "")
    else:
        factory_hid = f"W{variant}:hdr:factory:{factory}"
        is_factory_hdr = lambda w: getattr(w, "is_header", False) and ":hdr:factory:" in (getattr(w, "_vid", "") or "")

    children = list(parent.winfo_children())

    # наш заголовок завода
    hdr_idx = None
    for i, w in enumerate(children):
        if getattr(w, "_vid", None) == factory_hid:
            hdr_idx = i
            break
    if hdr_idx is None:
        return  # шапки ещё нет — оставим как есть, выровняется при «Обновить»

    # следующий именно ЗАГОЛОВОК ЗАВОДА (граница секции)
    next_factory_widget = None
    for j in range(hdr_idx + 1, len(children)):
        w2 = children[j]
        if is_factory_hdr(w2):
            next_factory_widget = w2
            break

    try:
        row.pack_forget()
    except Exception:
        pass

    kwargs = dict(fill=X, pady=PADY_PRODUCT)
    if next_factory_widget is not None:
        # вставляем перед следующим заводом => в конец текущего завода
        row.pack(before=next_factory_widget, **kwargs)
    else:
        # наш завод — последний, кладём в самый низ
        row.pack(**kwargs)


def _yset(cvs, pos: float):
    try:
        if cvs and cvs.winfo_exists():
            cvs.yview_moveto(pos)
    except Exception:
        pass


def update_product_list(
        preserve_scroll=True,
        regroup=True,
        reset_others_to_top=False,
        changed_canvases=None,
        repack=True,                         # флаг "умной" перепаковки
    ):
    global storage_count_cache
    repack_work = False
    # проверка наличия фреймов
    if (
        globals().get("work_frame") is None
        or globals().get("work7_frame") is None
        or globals().get("storage_frame") is None
    ):
        return
    # XL подсветка: перед любым построением списков подхватываем кэш


    # ---- сохранить вид до перерисовки
    if preserve_scroll:
        try:
            w6_view = _capture_view(work_frame, work_canvas)           # (top, off, alt)
        except Exception:
            w6_view = (None, 0, None)
        try:
            w7_view = _capture_view(work7_frame, work7_canvas)
        except Exception:
            w7_view = (None, 0, None)
        try:
            st_view = _capture_view(storage_frame, storage_canvas)
        except Exception:
            st_view = (None, 0, None)
    else:
        # при явном сбросе — всегда в начало
        for cvs in (globals().get("work_canvas"),
                    globals().get("work7_canvas"),
                    globals().get("storage_canvas")):
            try:
                if cvs and cvs.winfo_exists():
                    cvs.yview_moveto(0.0)
            except Exception:
                pass

    # ---- подсчёты
    work6_count = sum(
        1
        for k in products
        if get_variant(k) == "ВС6Д"
        and k not in assembled_products
        and k not in storage_products
    )
    work7_count = sum(
        1
        for k in products
        if get_variant(k) == "ВС7"
        and k not in assembled_products
        and k not in storage_products
    )
    storage_count = sum(1 for k in storage_products if k not in assembled_products)

    # ---- заголовки
    try:
        work_header_label.config(text=f"ВС6Д в работе ({work6_count})")
    except Exception:
        pass
    try:
        work7_header_label.config(text=f"ВС7 в работе ({work7_count})")
    except Exception:
        pass
    try:
        storage_header_label.config(text=f"{TEXT_STORAGE_TITLE} ({storage_count})")
    except Exception:
        pass

    # кнопка-стрелка всегда в шапке ВС7
    storage_count_cache = storage_count
    try:
        if not storage_visible:
            lbl = f"◀ ({storage_count_cache})"
            toggle_storage_button.config(text=lbl, width=len(lbl))
        else:
            toggle_storage_button.config(text="▶", width=2)
    except Exception:
        pass

    # ---- последовательности
    cc = _get_current_comments()
    work6_seq = _desired_work_sequence_variant(cc, "ВС6Д")
    work7_seq = _desired_work_sequence_variant(cc, "ВС7")
    storage_seq = _desired_storage_sequence(cc)

    # --- КЛЮЧЕВОЕ: для "Заготовки"/"Блоки" ВСЕГДА делаем полноценную перепаковку ---
    # В этих режимах появляются/исчезают группы ("Готовы к сборке", "0 блоков" и т.п.),
    # и новые заголовки иначе всегда будут рисоваться внизу.
    if work_sort_mode == "factories" and globals().get("hide_added_date", False):
        repack_work = True

    _apply_sequence(work_frame,  work6_seq, which="work6",  regroup=regroup, repack=repack_work)
    _apply_sequence(work7_frame, work7_seq, which="work7",  regroup=regroup, repack=repack_work)
    _apply_sequence(storage_frame, storage_seq, which="storage", regroup=regroup, repack=repack)
    # --- XL / up112: после пересборки списков принудительно обновим стили кнопок ---
    # Это нужно, потому что при смене сортировки часть строк пересоздаётся
    # и может пройти через путь, где подсветка (info) не применяется.
    try:
        # если есть что подсвечивать — пробегаемся и применяем стили
        if (globals().get("up112_hints") is not None) or \
           (globals().get("xl_hints_112") is not None) or \
           (globals().get("xl_hints_114") is not None) or \
           (globals().get("xl_hints_161") is not None):

            # work6_widgets / work7_widgets должны быть уже заполнены после _apply_sequence
            for k in list(work6_widgets.keys()):
                try:
                    update_row_widgets(k)
                except Exception:
                    pass

            for k in list(work7_widgets.keys()):
                try:
                    update_row_widgets(k)
                except Exception:
                    pass
    except Exception:
        pass

    # ---- восстановление видов
    if preserve_scroll:
        changed = set(changed_canvases or [])

        def _restore(view, frame, canvas):
            top, off, alt = (view + (None,))[:3] if len(view) == 2 else view
            _restore_view(frame, canvas, top, off, alt)

        # если задан reset_others_to_top и указаны "менявшиеся" канвасы —
        # восстанавливаем их, а остальные уводим в начало
        if reset_others_to_top and changed:
            if "work6" in changed:
                _restore(w6_view, work_frame, work_canvas)
            else:
                try:
                    work_canvas.yview_moveto(0.0)
                except Exception:
                    pass

            if "work7" in changed:
                _restore(w7_view, work7_frame, work7_canvas)
            else:
                try:
                    work7_canvas.yview_moveto(0.0)
                except Exception:
                    pass

            if "storage" in changed:
                _restore(st_view, storage_frame, storage_canvas)
            else:
                try:
                    storage_canvas.yview_moveto(0.0)
                except Exception:
                    pass
        else:
            # обычный случай — просто вернуть как было
            _restore(w6_view, work_frame, work_canvas)
            _restore(w7_view, work7_frame, work7_canvas)
            _restore(st_view, storage_frame, storage_canvas)

        try:
            if notebook.tab(notebook.select(), "text") == "Запросить блоки":
                update_request_table()
        except Exception:
            pass

def _vs7_112_text_for_search(key) -> str:
    """
    Возвращает текст для строки поиска по ССБ112 у ВС7/ВС13.
    Если данных нет — возвращает пустую строку (ничего не показываем).
    """
    meta = None

    # 1) пробуем по точному ключу
    try:
        meta = vs7_112_meta.get(key)
    except Exception:
        meta = None

    # 2) если не нашли — пробуем по 3-ключу (на случай старого формата хранения)
    if not isinstance(meta, dict):
        try:
            meta = vs7_112_meta.get(key2to3(key))
        except Exception:
            meta = None

    if not isinstance(meta, dict):
        return ""   # <- ВАЖНО: ничего не показываем

    from_vs6d = bool(meta.get("from_vs6d", False))
    last4 = str(meta.get("last4", "") or "").strip()

    if from_vs6d:
        return "взят из ВС6Д"

    if last4:
        return f"№ {last4}"

    return ""  # <- ВАЖНО: если свой, но номер не введён — тоже не показываем



def search_product():
    search_text = search_var.get().strip()

    # валидация ввода
    if not search_text.isdigit() or len(search_text) < 2 or len(search_text) > 5:
        messagebox.showerror("Ошибка", "Введите от 2 до 5 цифр (номер и завод)")
        return

    number_part = search_text[:-1].zfill(4)
    factory_code = search_text[-1]
    if factory_code not in factory_mapping:
        messagebox.showerror("Ошибка", "Некорректный код завода (1-6)")
        return

    factory_name = factory_mapping[factory_code]

    # собираем ВСЕ ключи, соответствующие номеру+заводу (оба варианта: ВС6Д и ВС7)
    all_keys = list(
        set(products.keys()) | set(storage_products) | set(assembled_products)
    )
    candidate_keys = [
        k
        for k in all_keys
        if len(k) >= 2 and k[0] == number_part and k[1] == factory_name
    ]

    if not candidate_keys:
        messagebox.showinfo("Результат поиска", "СЕРП'ы с такими номерами не найдены.")
        return

    # хотим стабильный порядок: сначала ВС6Д, затем ВС7
    def _variant_rank(key):
        v = get_variant(key)
        return 0 if v == "ВС6Д" else 1

    candidate_keys.sort(key=_variant_rank)

    # формируем отчёт по каждому найденному изделию
    sections = []
    for key in candidate_keys:
        name_long = format_key_long(key)  # "СЕРП ВС6Д №yyz0001" / "СЕРП ВС7 №yyz0001"
        lines = [name_long]
        # Для ВС7/ВС13 показываем информацию по ССБ112 из vs7_112_meta (только если она реально есть)
        try:
            v = str(get_variant(key)).strip()
            if v in ("ВС7", "ВС13"):
                txt112 = _vs7_112_text_for_search(key)
                if txt112:
                    lines.append(f"ССБ 112: {txt112}")
        except Exception:
            pass

        # ✅ ДОБАВЛЕНО: для ВС7 показываем привязанный серийник ССБ112
        try:
            if str(get_variant(key)).strip() == "ВС7":
                ssb112 = _get_vs7_ssb112_serial(key)

        except Exception:
            pass

        if (
            (key in products)
            and (key not in storage_products)
            and (key not in assembled_products)
        ):
            lines.append("— ожидает сборки")
            lines.append(f"Дата добавления: {product_dates.get(key, '??.??')}")
            blocks = products.get(key, {})
            for t in block_types_for(key):
                lines.append(f"{t.split()[-1]}: {'✓' if blocks.get(t, False) else '✕'}")

        elif key in storage_products and key not in assembled_products:
            ds = storage_dates.get(key, "??.??.??")
            lines.append("— собран, ждёт отправки на склад")
            lines.append(f"Дата сборки: {ds}")

        elif key in assembled_products:
            date_ship = assembly_dates.get(key, "??.??.??")
            date_asm = storage_dates.get(key, date_ship)
            lines.append("— отправлен на склад")
            lines.append(f"Дата отправки: {date_ship}")
            lines.append(f"Дата сборки: {date_asm}")

        else:
            lines.append("— состояние: неизвестно")

        sections.append("\n".join(lines))

    messagebox.showinfo("Результат поиска", "\n\n".join(sections))


def format_date_display(date_str):
    return format_date_genitive(date_str)

def _is_storage_comment(txt) -> bool:
    if not isinstance(txt, str):
        return False
    norm = " ".join(txt.strip().split()).casefold()
    return norm == "с хранилища"

# =========================
# POPUP ПОИСКА (вместо строки поиска)
# =========================

search_popup = None
search_popup_var = None
search_popup_entry = None

def close_search_popup():
    global search_popup, search_popup_var, search_popup_entry
    try:
        if search_popup is not None and search_popup.winfo_exists():
            search_popup.destroy()
    except Exception:
        pass
    search_popup = None
    search_popup_var = None
    search_popup_entry = None


def run_search_popup():
    """Запуск поиска из popup-окна."""
    global search_popup_var, search_popup_entry

    try:
        txt = (search_popup_var.get() or "").strip()
    except Exception:
        txt = ""

    if not txt:
        return

    # search_product() берёт текст из search_var -> кладём туда
    try:
        search_var.set(txt)
    except Exception:
        pass

    try:
        search_product()
    finally:
        # очистка поля в popup
        try:
            search_var.set("")
        except Exception:
            pass
        try:
            search_popup_var.set("")
        except Exception:
            pass
        try:
            if search_popup_entry and search_popup_entry.winfo_exists():
                search_popup_entry.xview_moveto(0)
                search_popup_entry.focus_set()
        except Exception:
            pass


def open_search_popup():
    """Открыть окно поиска (если уже открыто — просто поднять и сфокусировать)."""
    global search_popup, search_popup_var, search_popup_entry

    try:
        if search_popup is not None and search_popup.winfo_exists():
            search_popup.deiconify()
            search_popup.lift()
            if search_popup_entry and search_popup_entry.winfo_exists():
                search_popup_entry.focus_set()
            return
    except Exception:
        pass

    win = tb.Toplevel(root)
    search_popup = win

    try:
        win.title("Поиск")
    except Exception:
        pass

    try:
        win.configure(background=BG_MAIN)
    except Exception:
        pass

    try:
        win.transient(root)
    except Exception:
        pass

    try:
        win.resizable(False, False)
    except Exception:
        pass

    # --- контейнер ---
    container = tb.Frame(win, style="Main.TFrame", padding=10)
    container.pack(fill="both", expand=True)
    # --- строка ввода ---
    # --- строка ввода ---
    row = tb.Frame(container, style="Main.TFrame")
    row.pack(fill="x")

    # 1) значения по умолчанию
    yy_now = f"{datetime.now().year % 100:02d}"  # текущий год (2 цифры)
    variant_default = "ВС6Д"  # по умолчанию ВС6Д

    # 2) переменные (ВАЖНО: master=win)
    search_popup_year_var = tk.StringVar(master=win, value=yy_now)
    search_popup_variant_var = tk.StringVar(master=win, value=variant_default)
    search_popup_var = tk.StringVar(master=win, value="")

    # 3) списки значений
    year_values = [f"{y:02d}" for y in range(24, 100)]  # как у тебя в основной форме
    variant_values = ["ВС6Д", "ВС7"]

    # 4) комбобокс года
    year_box = _wrap_with_border(row)
    year_box.pack(side="left", padx=(0, 6))

    year_combo = tb.Combobox(
        year_box,
        textvariable=search_popup_year_var,
        state="readonly",
        width=4,
        bootstyle="light",
        font=FONT_PRIMARY,
    )
    year_combo.configure(values=year_values)
    year_combo.pack(fill="x", padx=4, pady=3)

    # ПРИНУДИТЕЛЬНО выбираем текущий год (иначе readonly может показать пусто)
    try:
        year_combo.current(year_values.index(yy_now))
    except Exception:
        year_combo.current(0)
        search_popup_year_var.set(year_values[0])

    # 5) комбобокс варианта
    variant_box = _wrap_with_border(row)
    variant_box.pack(side="left", padx=(0, 6))

    variant_combo = tb.Combobox(
        variant_box,
        textvariable=search_popup_variant_var,
        state="readonly",
        width=6,
        bootstyle="light",
        font=FONT_PRIMARY,
    )
    variant_combo.configure(values=variant_values)
    variant_combo.pack(fill="x", padx=4, pady=3)

    # ПРИНУДИТЕЛЬНО выбираем ВС6Д
    try:
        variant_combo.current(0)  # "ВС6Д"
    except Exception:
        search_popup_variant_var.set("ВС6Д")

    # если режим VS6D-only — блокируем выбор ВС7
    try:
        if vs6d_only_mode:
            search_popup_variant_var.set("ВС6Д")
            variant_combo.current(0)
            variant_combo.configure(state="disabled")
    except Exception:
        pass

    # --- FIX: combobox'ы могут быть пустыми до выбора мышью, поэтому ставим значения ПОСЛЕ показа окна ---
    def _force_defaults_after_open():
        try:
            # год
            year_combo.configure(state="normal")
            year_combo.set(yy_now)  # например "25"
            year_combo.configure(state="readonly")
        except Exception:
            pass

        try:
            # вариант
            variant_combo.configure(state="normal")
            variant_combo.set("ВС6Д")
            variant_combo.configure(state="disabled" if vs6d_only_mode else "readonly")
        except Exception:
            pass

        try:
            # чтобы сразу было удобно печатать номер
            ent.focus_set()
        except Exception:
            pass

    # важно: сделать ПОСЛЕ отображения окна
    win.after(50, _force_defaults_after_open)
    win.after(150, _force_defaults_after_open)  # дубль (иногда первый тик не успевает)

    # 6) поле ввода (как у тебя было)
    box = _wrap_with_border(row)  # ваш же стиль рамки
    box.pack(side="left", fill="x", expand=True)

    ent = tb.Entry(
        box,
        textvariable=search_popup_var,
        font=FONT_PRIMARY,
        bootstyle="light",
        width=22,
    )
    ent.pack(fill="x", padx=4, pady=3)

    # 7) кнопка поиска
    tb.Button(
        row,
        text="🔍",
        width=4,
        bootstyle="secondary",
        command=run_search_popup,
    ).pack(side="left", padx=(8, 0))

    # бинды
    ent.bind("<Return>", lambda e: run_search_popup())
    win.bind("<Escape>", lambda e: close_search_popup())
    win.protocol("WM_DELETE_WINDOW", close_search_popup)

    # позиционирование по центру главного окна
    try:
        win.update_idletasks()
        x = root.winfo_rootx() + (root.winfo_width() // 2) - (win.winfo_reqwidth() // 2)
        y = root.winfo_rooty() + 80
        win.geometry(f"+{max(0, x)}+{max(0, y)}")
    except Exception:
        pass

    try:
        ent.focus_set()
    except Exception:
        pass


def _is_file_locked(path: str) -> bool:
    """True, если файл заблокирован (например, открыт в Excel)."""
    if not os.path.exists(path):
        return False
    try:
        os.rename(path, path)  # на Windows бросит PermissionError, если файл залочен
        return False
    except Exception:
        return True

def export_to_excel():
    r"""
    Создаёт/обновляет отчёт C:\serp_base\Отчёт по сборке.xlsx

    Структура (всегда 14 листов):
      • "ВС6Д (Отчёт)", "ВС7 (Отчёт)"
      • По заводам: "<Завод> ВС6Д" и "<Завод> ВС7" — для каждого из 6 заводов.

    Колонки:
      • Общие: Модель СЕРП, Серийный номер, Завод, Дата отправки, Дата сборки, Дата добавления, Комментарий
      • Для ВС6Д (перед «Комментарий»): ССБ 112, ССБ 161, ССБ 114
        ─ 112 = "12" + yy + z + xxxx
        ─ 161 = "61" + yy + z + xxxx
        ─ 114 = "14" + yy + z + xxxx
      • Для ВС7 (перед «Комментарий»): ССБ 112, ССБ 116
        ─ 116 = "16" + yy + z + xxxx
        ─ 112 = "12" + yy + z + "????"  (пока неизвестны последние 4 цифры)
    """
    # ---------- deps ----------
    try:
        import pandas as pd
    except ImportError:
        messagebox.showerror("Ошибка", "Для отчёта нужен пакет pandas: pip install pandas")
        return

    engine_name = "xlsxwriter"
    try:
        import xlsxwriter  # noqa: F401
    except Exception:
        engine_name = "openpyxl"

    # ---------- helpers ----------


    def _vs7_112_for(key) -> str:
        """
        12yyzXXXX для ВС7:
          - если meta['from_vs6d'] == True -> XXXX = номер СЕРПа (num4)
          - иначе если meta['last4'] заданы корректно -> XXXX = last4
          - иначе XXXX = '????'
        """
        yy, z, num4 = serial_parts_from_key(key)
        meta = vs7_112_meta.get(key) or {}
        if meta.get("from_vs6d") is True:
            last4 = num4
        else:
            last4 = meta.get("last4") or "????"
            if not (last4.isdigit() and len(last4) == 4):
                last4 = "????"
        return f"12{yy}{z}{last4}"

    def _normalize_variant(v):
        # Исторически «ВС7» мог записываться как «ВС13». Для отчёта считаем это «ВС7».
        return "ВС7" if str(v).strip() in ("ВС7", "ВС13") else str(v).strip()

    def _is_file_locked(path: str) -> bool:
        if not os.path.exists(path):
            return False
        try:
            os.rename(path, path)
            return False
        except Exception:
            return True

    def _norm_ddmmyy(s: str, fallback_year2: str) -> str:
        if not isinstance(s, str) or not s.strip():
            return ""
        parts = s.strip().split(".")
        try:
            if len(parts) == 2:
                d, m = int(parts[0]), int(parts[1])
                y2 = (fallback_year2 or f"{datetime.now().year % 100:02d}")[:2]
                return f"{d:02d}.{m:02d}.{y2}"
            if len(parts) >= 3:
                d, m = int(parts[0]), int(parts[1])
                y = str(parts[2]).strip()
                y2 = y[-2:] if len(y) >= 2 and y[-2:].isdigit() else f"{datetime.now().year % 100:02d}"
                return f"{d:02d}.{m:02d}.{y2}"
        except Exception:
            pass
        return ""

    def _excel_col_letter(idx1: int) -> str:
        s = ""
        while idx1:
            idx1, r = divmod(idx1 - 1, 26)
            s = chr(65 + r) + s
        return s

    def _calc_widths(df_out, order_cols, date_cols=()):
        tmp = df_out.copy()
        for c in date_cols:
            if c in tmp.columns:
                tmp[c] = pd.to_datetime(tmp[c], errors="coerce").dt.strftime("%d.%m.%y").fillna("")
        max_lens = {}
        for c in order_cols:
            if c not in tmp.columns:
                continue
            series_str = tmp[c].fillna("").astype(str).tolist()
            max_len = max([len(c)] + [len(x) for x in series_str])
            max_lens[c] = max_len
        return max_lens

    def _write_sheet_xlsxwriter(writer, df_out, sheet_name, order_cols, remove_cols=()):
        df_to_write = df_out.drop(columns=list(remove_cols), errors="ignore")
        present_cols = [c for c in order_cols if c in df_to_write.columns]
        df_to_write = df_to_write.reindex(columns=present_cols)
        df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = writer.sheets[sheet_name]

        base_fmt = wb.add_format(
            {"font_name": "Times New Roman", "font_size": 12, "align": "center", "valign": "vcenter"}
        )
        header_fmt = wb.add_format(
            {"font_name": "Times New Roman", "font_size": 12, "bold": True, "align": "center", "valign": "vcenter"}
        )
        date_fmt = wb.add_format(
            {"font_name": "Times New Roman", "font_size": 12, "align": "center", "valign": "vcenter",
             "num_format": "dd.mm.yy"}
        )
        text_fmt = wb.add_format(
            {"font_name": "Times New Roman", "font_size": 12, "align": "center", "valign": "vcenter",
             "num_format": "@"}
        )

        ws.set_row(0, None, header_fmt)

        PADDING = 6
        MIN_WIDTHS = {
            "Завод": 22,
            "Серийный номер": 18,
            "Модель СЕРП": 14,
            "Комментарий": 14,
            "Дата отправки": 12,
            "Дата сборки": 12,
            "Дата добавления": 12,
            "ССБ 112": 18,
            "ССБ 114": 18,
            "ССБ 116": 18,
            "ССБ 161": 18,
        }
        max_lens = _calc_widths(
            df_to_write,
            present_cols,
            date_cols=("Дата отправки", "Дата сборки", "Дата добавления"),
        )

        col_idx = {name: i + 1 for i, name in enumerate(present_cols)}
        for name in present_cols:
            col_letter = _excel_col_letter(col_idx[name])
            width = max(max_lens.get(name, 10) + PADDING, MIN_WIDTHS.get(name, 10))
            if name in ("Дата отправки", "Дата сборки", "Дата добавления"):
                ws.set_column(f"{col_letter}:{col_letter}", width, date_fmt)
            elif name in ("Серийный номер", "ССБ 112", "ССБ 114", "ССБ 116", "ССБ 161"):
                ws.set_column(f"{col_letter}:{col_letter}", width, text_fmt)
                try:
                    ws.ignore_errors({"number_stored_as_text": f"{col_letter}2:{col_letter}{len(df_to_write)+1}"})
                except Exception:
                    pass
            else:
                ws.set_column(f"{col_letter}:{col_letter}", width, base_fmt)

        ws.freeze_panes(1, 0)
        try:
            ws.protect("serp_base")
        except Exception:
            pass

    def _write_sheet_openpyxl(writer, df_out, sheet_name, order_cols, remove_cols=()):
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        df_to_write = df_out.drop(columns=list(remove_cols), errors="ignore")
        present_cols = [c for c in order_cols if c in df_to_write.columns]
        df_to_write = df_to_write.reindex(columns=present_cols)
        df_to_write.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        font = Font(name="Times New Roman", size=12)
        font_bold = Font(name="Times New Roman", size=12, bold=True)
        align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = font_bold
            cell.alignment = align

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font
                cell.alignment = align

        date_cols_names = [c for c in ("Дата отправки", "Дата сборки", "Дата добавления") if c in present_cols]
        date_cols_idx = [present_cols.index(c)+1 for c in date_cols_names]
        for col in date_cols_idx:
            for (cell,) in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                cell.number_format = "DD.MM.YY"

        PADDING = 6
        MIN_WIDTHS = {
            "Завод": 22,
            "Серийный номер": 18,
            "Модель СЕРП": 14,
            "Комментарий": 14,
            "Дата отправки": 12,
            "Дата сборки": 12,
            "Дата добавления": 12,
            "ССБ 112": 18,
            "ССБ 114": 18,
            "ССБ 116": 18,
            "ССБ 161": 18,
        }
        max_lens = _calc_widths(
            df_to_write,
            present_cols,
            date_cols=("Дата отправки", "Дата сборки", "Дата добавления"),
        )
        for i, name in enumerate(present_cols, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(max_lens.get(name, 10) + PADDING,
                                                                   MIN_WIDTHS.get(name, 10))

        ws.freeze_panes = "A2"

        try:
            from openpyxl.worksheet.ignored_errors import IgnoredErrors, IgnoredError
            ie = IgnoredErrors()
            last_row = ws.max_row
            for name in ("Серийный номер", "ССБ 112", "ССБ 114", "ССБ 116", "ССБ 161"):
                if name in present_cols:
                    cidx = present_cols.index(name) + 1
                    col_letter = get_column_letter(cidx)
                    for (cell,) in ws.iter_rows(min_row=2, max_row=last_row, min_col=cidx, max_col=cidx):
                        cell.number_format = "@"
                    ie.append(IgnoredError(range=f"{col_letter}2:{col_letter}{last_row}", numberStoredAsText=True))
            ws.ignored_errors = ie
        except Exception:
            pass

        try:
            ws.protection.set_password("serp_base")
            ws.protection.sheet = True
        except Exception:
            try:
                ws.protection.password = "serp_base"
                ws.protection.sheet = True
            except Exception:
                pass

    def _block_serials_for_key(key, variant_norm: str):
        yy, z, num4 = serial_parts_from_key(key)
        base = f"{yy}{z}{num4}"
        if variant_norm == "ВС6Д":
            return {"ССБ 112": f"12{base}", "ССБ 161": f"61{base}", "ССБ 114": f"14{base}", "ССБ 116": ""}
        else:
            return {"ССБ 112": _vs7_112_for(key), "ССБ 161": "", "ССБ 114": "", "ССБ 116": f"16{base}"}

    # ---------- данные ----------
    # ---------- данные ----------
    rows = []
    for key in assembled_products:
        # number берём как раньше, но ВАРИАНТ — только через get_variant + нормализацию
        number, factory, _ = _split_key_safe(key)
        variant_raw = get_variant(key)
        variant = "ВС7" if str(variant_raw).strip() in ("ВС7", "ВС13") else "ВС6Д"

        y2 = (assembly_years.get(key) or f"{datetime.now().year % 100:02d}")[:2]
        ship_dt = _norm_ddmmyy(assembly_dates.get(key, ""), y2)
        asm_dt = _norm_ddmmyy(storage_dates.get(key, assembly_dates.get(key, "")), y2)
        add_dt = _norm_ddmmyy(product_dates.get(key, ""), y2)

        serp_serial = serial_for_key(key)
        comment = comments.get(key, "")
        comment = comment if isinstance(comment, str) and comment.strip() else "нет"

        # КЛЮЧЕВОЕ: блок-серийники считаем общей функцией
        blk_112 = ssb_serial_auto(key, "ССБ 112")
        blk_161 = ssb_serial_auto(key, "ССБ 161")
        blk_114 = ssb_serial_auto(key, "ССБ 114")
        blk_116 = ssb_serial_auto(key, "ССБ 116")

        rows.append(
            {
                "Модель СЕРП": variant,  # теперь ВС7 не превратится в ВС6Д
                "Серийный номер": str(serp_serial),
                "Завод": factory,
                "Дата отправки": ship_dt,
                "Дата сборки": asm_dt,
                "Дата добавления": add_dt,
                "ССБ 112": blk_112,  # для ВС7 тут будет 12yyzXXXX из vs7_112_meta
                "ССБ 161": blk_161 if variant == "ВС6Д" else "",  # у ВС7 пусто
                "ССБ 114": blk_114 if variant == "ВС6Д" else "",  # у ВС7 пусто
                "ССБ 116": blk_116 if variant == "ВС7" else "",  # у ВС6Д пусто
                "Комментарий": comment,
            }
        )

    all_cols = [
        "Модель СЕРП", "Серийный номер", "Завод",
        "Дата отправки", "Дата сборки", "Дата добавления",
        "ССБ 112", "ССБ 161", "ССБ 114", "ССБ 116",
        "Комментарий",
    ]
    df = pd.DataFrame(rows, columns=all_cols)

    for c in ("Дата отправки", "Дата сборки", "Дата добавления"):
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], format="%d.%m.%y", errors="coerce")

    df_vs6d = df[df["Модель СЕРП"] == "ВС6Д"].sort_values(by=["Дата отправки"], ascending=False)
    df_vs7  = df[df["Модель СЕРП"] == "ВС7" ].sort_values(by=["Дата отправки"], ascending=False)

    base_common = ["Модель СЕРП", "Серийный номер", "Завод", "Дата отправки", "Дата сборки", "Дата добавления"]
    cols_vs6d   = base_common + ["ССБ 112", "ССБ 161", "ССБ 114", "Комментарий"]
    cols_vs7    = base_common + ["ССБ 112", "ССБ 116", "Комментарий"]

    # ---------- путь и проверка "файл открыт" ----------
    report_dir = r"C:\serp_base"
    os.makedirs(report_dir, exist_ok=True)
    report_path = os.path.join(report_dir, "Отчёт по сборке.xlsx")

    while _is_file_locked(report_path):
        if not messagebox.askretrycancel(
            "Файл открыт",
            f"Отчёт уже открыт в Excel:\n{report_path}\n\nЗакройте файл и нажмите «Повторить».",
        ):
            return

    factories = ["ВЕКТОР", "ИНТЕГРАЛ", "РЗП", "КНИИТМУ", "СВТ", "СИГНАЛ"]

    # ---------- запись: 14 листов (всегда) ----------
    while True:
        try:
            if engine_name == "xlsxwriter":
                with pd.ExcelWriter(report_path, engine="xlsxwriter", datetime_format="dd.mm.yy") as writer:
                    _write_sheet_xlsxwriter(writer, df_vs6d, "ВС6Д (Отчёт)", cols_vs6d)
                    _write_sheet_xlsxwriter(writer, df_vs7,  "ВС7 (Отчёт)",  cols_vs7)

                    for fac in factories:
                        dff6 = df_vs6d[df_vs6d["Завод"] == fac].copy()
                        dff7 = df_vs7 [df_vs7 ["Завод"] == fac].copy()

                        for dff, cols in ((dff6, cols_vs6d), (dff7, cols_vs7)):
                            dff["_y2"]  = pd.to_numeric(dff["Серийный номер"].str.slice(0, 2), errors="coerce")
                            dff["_num"] = pd.to_numeric(dff["Серийный номер"].str.slice(-4), errors="coerce")
                            dff.sort_values(by=["_y2", "_num"], ascending=[True, True], inplace=True)
                            dff.drop(columns=["_y2", "_num"], inplace=True)
                            name = f"{fac} {'ВС6Д' if cols is cols_vs6d else 'ВС7'}"
                            _write_sheet_xlsxwriter(writer, dff, name, cols, remove_cols=("Завод","Модель СЕРП"))

                    try:
                        writer.book.protect("serp_base")
                    except Exception:
                        pass

            else:
                with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
                    _write_sheet_openpyxl(writer, df_vs6d, "ВС6Д (Отчёт)", cols_vs6d)
                    _write_sheet_openpyxl(writer, df_vs7,  "ВС7 (Отчёт)",  cols_vs7)

                    for fac in factories:
                        dff6 = df_vs6d[df_vs6d["Завод"] == fac].copy()
                        dff7 = df_vs7 [df_vs7 ["Завод"] == fac].copy()

                        for dff, cols in ((dff6, cols_vs6d), (dff7, cols_vs7)):
                            dff["_y2"]  = pd.to_numeric(dff["Серийный номер"].str.slice(0, 2), errors="coerce")
                            dff["_num"] = pd.to_numeric(dff["Серийный номер"].str.slice(-4), errors="coerce")
                            dff.sort_values(by=["_y2", "_num"], ascending=[True, True], inplace=True)
                            dff.drop(columns=["_y2", "_num"], inplace=True)
                            name = f"{fac} {'ВС6Д' if cols is cols_vs6d else 'ВС7'}"
                            _write_sheet_openpyxl(writer, dff, name, cols, remove_cols=("Завод","Модель СЕРП"))

                    try:
                        from openpyxl.workbook.protection import WorkbookProtection
                        wb = writer.book
                        wb.security = WorkbookProtection(workbookPassword="serp_base", lockStructure=True)
                    except Exception:
                        pass

            break

        except PermissionError:
            if not messagebox.askretrycancel(
                "Файл открыт",
                f"Excel блокирует перезапись:\n{report_path}\n\nЗакройте файл и нажмите «Повторить».",
            ):
                return
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{e}")
            return

    try:
        os.startfile(report_path)
    except Exception:
        pass



def move_factory_to_top(factory):
  return


def reset_factory_order():
    return

def toggle_sort_order():
    global sort_order
    sort_order = "old_first" if sort_order == "new_first" else "new_first"
    sort_button.config(text="⏫" if sort_order == "new_first" else "⏬")
    save_data()
    save_scroll_positions()
    update_product_list()
    restore_scroll_positions()


def set_product_mode(mode: str):
    global product_mode
    if vs6d_only_mode:
        # в режиме ВС6Д-только игнорируем любые попытки выбрать ВС7
        mode = "ВС6Д"
    if mode not in ("ВС6Д", "ВС7"):
        return
    product_mode = mode
    try:
        if vs6d_btn:
            vs6d_btn.configure(
                bootstyle="primary" if product_mode == "ВС6Д" else "secondary"
            )
        if vs13_btn:
            vs13_btn.configure(  # теперь это кнопка ВС7
                bootstyle="primary" if product_mode == "ВС7" else "secondary"
            )

    except Exception:
        pass
    update_product_list()
    if notebook.tab(notebook.select(), "text") == "Архив сборки":
        update_assembly_archive()

def update_vs13_button_state():
    global product_mode
    if not (vs6d_btn and vs13_btn):
        return
    try:
        vs6d_btn.configure(
            text="ВС6Д",
            bootstyle="primary" if product_mode == "ВС6Д" else "secondary",
            state="normal",
        )
        vs13_btn.configure(
            text="ВС7",
            bootstyle="primary" if product_mode == "ВС7" else "secondary",
            state="normal",
        )
    except Exception:
        pass

    # ВАЖНО: заголовок фиксированный
    try:
        if entry_header_label and entry_header_label.cget("text") != "Добавить СЕРП:":
            entry_header_label.config(text="Добавить СЕРП:")
    except Exception:
        pass

    update_product_list()
    if notebook.tab(notebook.select(), "text") == "Архив сборки":
        update_assembly_archive()




def on_draft_puzzle_left_click(event=None):
    # ЛКМ: включить/выключить режим «заготовки/ремонт».
    # Если был активен режим по количеству — отключаем его.
    global show_draft_group, blockcount_mode
    if blockcount_mode:
        blockcount_mode = False
        show_draft_group = True  # явный переход в заготовки
    else:
        show_draft_group = not show_draft_group
    set_puzzle_button_visual()
    save_data()
    update_product_list(preserve_scroll=False)
    return "break"


def toggle_blockcount_mode(event=None):
    # ПКМ по 🧩 или ПКМ по стрелке даты: режим "по количеству блоков"
    global blockcount_mode, show_draft_group
    blockcount_mode = not blockcount_mode
    if blockcount_mode:
        show_draft_group = False  # режимы взаимоисключающие
    set_puzzle_button_visual()
    save_data()
    update_product_list(preserve_scroll=False)
    return "break"

def on_toggle_full_serials_work():
    global display_full_serials
    display_full_serials = display_full_serials_var.get()
    save_data(False)
    # Полный серийник влияет на разметку в рабочих колонках → пересоздать списки
    update_product_list(preserve_scroll=False, regroup=True)

def on_toggle_hide_added_date():
    global hide_added_date
    try:
        hide_added_date = bool(hide_added_date_var.get())
    except Exception:
        hide_added_date = False

    save_data(False)
    # дата в подписи влияет на разметку => лучше перегруппировать
    update_product_list(preserve_scroll=True, regroup=True)


def refresh_storage_labels_in_place():
    """Переобновить подписи в хранилище без пересборки и без сдвига скролла."""
    for key, row in storage_widgets.items():
        if not row or not row.winfo_exists():
            continue

        number, factory_name, _ = _split_key_safe(key)
        z = factory_reverse_mapping.get(factory_name, "?")

        if display_full_serials_storage:
            # Полный серийник: №yyzxxxx
            txt = f"№{serial_for_key(key)}"
        else:
            pref = get_product_prefix_from_key(key, compact=True)
            if storage_sort_mode == "models":
                # Модельный режим: "<модель> №xxxx [z]"
                txt = f"{pref} №{number} [{z}]"
            else:
                # Режим "по датам/по заводам": "<модель> №xxxx"
                txt = f"{pref} №{number}"

        try:
            row.serial_label.config(text=txt)
        except Exception:
            pass

def on_toggle_full_serials_storage():
    global display_full_serials_storage
    display_full_serials_storage = display_full_serials_storage_var.get()
    save_data(False)
    refresh_storage_labels_in_place()  # ни пересборки, ни прокрутки





def save_settings():
    save_data()
    messagebox.showinfo("Настройки", "Настройки успешно сохранены!")
    update_product_list()


def update_button_colors():
    return


def set_puzzle_button_visual():

    try:
        if blockcount_mode:
            draft_group_button.configure(text="🧩", bootstyle="danger")
        elif show_draft_group:
            draft_group_button.configure(text="🧩", bootstyle="warning")
        else:
            draft_group_button.configure(text="🧩", bootstyle="secondary")
    except Exception:
        pass


def get_week_range():
    """
    Возвращает границы недели как date (без времени):
    понедельник .. воскресенье включительно.
    """
    today = datetime.now().date()
    start = today - timedelta(days=today.weekday())  # понедельник
    end = start + timedelta(days=6)                  # воскресенье
    return start, end


def count_assembled_this_month():
    now = datetime.now()
    count = 0
    for key in assembled_products:
        date_str = assembly_dates.get(key, "")
        try:
            if "." in date_str:
                parts = date_str.split(".")
                if len(parts) == 2:
                    day, month = int(parts[0]), int(parts[1])
                    year = now.year
                elif len(parts) == 3:
                    day, month, year_part = int(parts[0]), int(parts[1]), int(parts[2])
                    year = 2000 + year_part if year_part < 100 else year_part
                else:
                    continue
                if month == now.month and year == now.year:
                    count += 1
        except:
            continue
    return count


def count_assembled_this_week():
    start_date, end_date = get_week_range()
    count = 0

    for key in assembled_products:
        date_str = assembly_dates.get(key, "")
        try:
            parts = date_str.split(".")
            if len(parts) == 2:
                day, month = int(parts[0]), int(parts[1])
                year = datetime.now().year
            elif len(parts) == 3:
                day, month = int(parts[0]), int(parts[1])
                yp = int(parts[2])
                year = 2000 + yp if yp < 100 else yp
            else:
                continue

            d = datetime(year, month, day).date()  # <-- СРАВНИВАЕМ КАК date
            if start_date <= d <= end_date:
                count += 1
        except Exception:
            continue

    return count

def clear_archive():
    global archive_scroll_position
    if not messagebox.askyesno(
        "Подтверждение", "Вы уверены, что хотите полностью очистить архив сборки?"
    ):
        return
    save_scroll_positions()
    archive_keys = set(assembled_products)
    assembled_products.clear()
    for key in archive_keys:
        products.pop(key, None)
        comments.pop(key, None)
        product_dates.pop(key, None)
        assembly_dates.pop(key, None)
        assembly_years.pop(key, None)
    update_product_list()
    update_assembly_archive()
    save_data()


def delete_from_archive(key):
    global archive_scroll_position
    if not messagebox.askyesno(
        "Подтверждение", f"Удалить {format_key_long(key)} из архива?"
    ):
        return
    save_scroll_positions()
    if key in assembled_products:
        assembled_products.remove(key)
    assembly_dates.pop(key, None)
    assembly_years.pop(key, None)
    comments.pop(key, None)
    products.pop(key, None)
    update_assembly_archive()
    update_product_list()
    save_data()


def set_archive_mode(mode):
    global archive_view_mode, archive_scroll_position
    save_scroll_positions()
    archive_view_mode = mode
    save_data()
    update_assembly_archive()


def update_assembly_archive():
    for widget in archive_frame.winfo_children():
        widget.destroy()

    main_container = tb.Frame(archive_frame, style="Main.TFrame")
    main_container.pack(fill=BOTH, expand=True, padx=10, pady=10)
    main_container.configure(style="Main.TFrame")

    # левая часть – список
    left_frame = tb.Frame(main_container, style="Main.TFrame")
    left_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
    left_frame.configure(style="Main.TFrame")

    # правая часть – статистика и действия
    right_frame = tb.Frame(main_container, style="Main.TFrame", width=300)
    right_frame.pack(side=RIGHT, fill=Y, padx=(10, 0))
    right_frame.pack_propagate(False)
    right_frame.configure(style="Main.TFrame")

    # ------ режим просмотра (слева, над списком) ------
    mode_frame = tb.Frame(left_frame, style="Main.TFrame")
    mode_frame.pack(fill=X, pady=(0, 10))
    tb.Label(
        mode_frame,
        text="Режим просмотра:",
        font=("Segoe UI", 12, "bold"),
        style="Main.TLabel",
    ).pack(side=LEFT, padx=10, pady=5)

    tb.Button(
        mode_frame,
        text="Журнал",
        width=12,
        bootstyle="primary" if archive_view_mode == "journal" else "secondary",
        command=lambda: set_archive_mode("journal"),
    ).pack(side=LEFT, padx=5, pady=5)

    tb.Button(
        mode_frame,
        text="История",
        width=12,
        bootstyle="primary" if archive_view_mode == "history" else "secondary",
        command=lambda: set_archive_mode("history"),
    ).pack(side=LEFT, padx=5, pady=5)

    # прокрутка для списка
    list_container = tb.Frame(left_frame, style="Main.TFrame")
    list_container.pack(fill=BOTH, expand=True)
    list_container.configure(style="Main.TFrame")

    list_canvas = tb.Canvas(list_container, background=BG_MAIN, highlightthickness=0)
    list_scrollbar = tb.Scrollbar(
        list_container,
        orient=VERTICAL,
        command=list_canvas.yview,
        bootstyle="light-round",
    )
    list_content = tb.Frame(list_canvas, style="Main.TFrame")
    list_content.configure(style="Main.TFrame")
    list_canvas.create_window((0, 0), window=list_content, anchor=NW)

    def on_canvas_configure(event):
        list_canvas.itemconfig("all", width=event.width)

    def on_frame_configure(event):
        list_canvas.configure(scrollregion=list_canvas.bbox("all"))

    list_canvas.bind("<Configure>", on_canvas_configure)
    list_content.bind("<Configure>", on_frame_configure)
    list_canvas.configure(yscrollcommand=list_scrollbar.set)
    list_scrollbar.pack(side=RIGHT, fill=Y)
    list_canvas.pack(side=LEFT, fill=BOTH, expand=True)

    # ---------- правая колонка: статистика + действия ----------
    stats_header = tb.Label(
        right_frame, text="Статистика архива", font=("Arial", 14), style="Main.TLabel"
    )
    stats_header.pack(pady=(0, 10))

    stats_container = tb.Frame(right_frame, style="Main.TFrame")
    stats_container.pack(fill=X, pady=(0, 15), padx=5)
    stats_container.configure(style="Main.TFrame")

    week_count = count_assembled_this_week()
    month_count = count_assembled_this_month()
    total_count = len(assembled_products)
    months_full = [
        "январь",
        "февраль",
        "март",
        "апрель",
        "май",
        "июнь",
        "июль",
        "август",
        "сентябрь",
        "октябрь",
        "ноябрь",
        "декабрь",
    ]
    month_name = months_full[datetime.now().month - 1]

    for label, value in [
        ("За неделю:", f"{week_count} шт."),
        (f"За {month_name}:", f"{month_count} шт."),
        ("В архиве:", f"{total_count} шт."),
    ]:
        row = tb.Frame(stats_container, style="Main.TFrame")
        row.pack(fill=X, pady=4, padx=10)
        row.configure(style="Main.TFrame")
        tb.Label(row, text=label, font=("Segoe UI", 11), style="Main.TLabel").pack(
            side=LEFT
        )
        tb.Label(row, text=value, font=("Segoe UI", 11), style="Main.TLabel").pack(
            side=RIGHT
        )

    clear_btn = tb.Button(
        right_frame,
        text="Очистить архив",
        width=20,
        bootstyle="secondary",
        command=clear_archive,
    )
    clear_btn.pack(pady=10)

    export_btn = tb.Button(
        right_frame,
        text="Отчёт Excel",
        width=20,
        bootstyle="success",
        command=export_to_excel,
    )
    export_btn.pack(pady=10)

    # ---------- helpers ----------
    def _added_ddmmyy(_key):
        """
        product_dates хранит 'дд.мм' или 'дд.мм.гг'; assembly_years — 2 цифры года добавления.
        Возвращаем 'дд.мм.гг'
        """
        d = product_dates.get(_key)
        y2 = assembly_years.get(_key)
        if isinstance(d, str):
            parts = d.split(".")
            if len(parts) == 2:
                # есть только дд.мм, год возьмем из assembly_years (2 цифры) или из текущего
                y2s = y2 if y2 and len(y2) == 2 else str(datetime.now().year)[-2:]
                return f"{parts[0]}.{parts[1]}.{y2s}"
            elif len(parts) == 3:
                return f"{parts[0]}.{parts[1]}.{parts[2][-2:]}"
        return "??.??.??"

    def _header_human(date_str):
        """'дд.мм.гг' -> 'дд месяц(в родительном)' без года"""
        try:
            parts = (date_str or "").split(".")
            if len(parts) >= 2:
                day = int(parts[0])
                month = int(parts[1])
                if 1 <= month <= 12:
                    return f"{day} {MONTHS_GENITIVE[month - 1]}"
        except Exception:
            pass
        return date_str or "??.??"

    # ---------- наполнение левой части ----------
    all_items = list(assembled_products)
    if archive_view_mode == "history":
        # группировка по дате отправки (assembly_dates), сортировка по дате убыв.
        all_items.sort(
            key=lambda k: parse_date_str(assembly_dates.get(k, "01.01.00")),
            reverse=True,
        )
        by_date = {}
        for key in all_items:
            d_ship = assembly_dates.get(key, "??.??.??")
            by_date.setdefault(d_ship, []).append(key)

        sorted_dates = sorted(
            by_date.keys(), key=lambda d: parse_date_str(d), reverse=True
        )

        for date_str in sorted_dates:
            keys_for_day = by_date[date_str]
            n_rows = len(keys_for_day)

            # Заголовок даты: "12 октября (n)"
            date_header = tb.Frame(list_content, style="Main.TFrame")
            date_header.pack(fill=X, padx=10, pady=(15, 5))
            tb.Label(
                date_header,
                text=f"{_header_human(date_str)} ({n_rows})",
                font=ARCHIVE_HEADER_FONT,
                style="Main.TLabel",
            ).pack(anchor="w")

            # Контейнер со списком строк (на основном сером фоне)
            lines = tb.Frame(list_content, style="Main.TFrame")
            lines.pack(fill=X, padx=30, pady=(0, 8))

            for idx, key in enumerate(sorted(keys_for_day, key=lambda k: int(_split_key_safe(k)[0])), start=1):
                variant = get_variant(key)  # "ВС6Д" или "ВС7"
                serial = serial_for_key(key)  # "ггzxxxx"
                asm_ddmm = format_ddmm(storage_dates.get(key, assembly_dates.get(key, "")))  # "дд.мм"

                text = f"{idx}) {variant} №{serial} (дата сборки: {asm_ddmm})"

                row = tb.Frame(lines, style="Main.TFrame")
                row.pack(fill=X, pady=ARCHIVE_ROW_PADY)

                lbl = tb.Label(row, text=text, font=ARCHIVE_FONT, style="Main.TLabel")
                lbl.pack(anchor="w")
                lbl._archive_key = key

                menu = tk.Menu(lbl, tearoff=False)
                menu.add_command(label="Удалить", command=lambda k=key: delete_from_archive(k))

                def _rclick(e, m=menu):
                    try:
                        m.tk_popup(e.x_root, e.y_root)
                    finally:
                        m.grab_release()

                lbl.bind("<Button-3>", _rclick)
                lbl.bind("<Control-Button-1>", _rclick)

        # прокрутка в начало
        list_canvas.update_idletasks()
        list_canvas.yview_moveto(0)

        # фон всего блока (на всякий случай)
        def update_bg(widget):
            try:
                widget.configure(background=BG_MAIN)
            except:
                pass
            for child in widget.winfo_children():
                update_bg(child)

        update_bg(main_container)
    elif archive_view_mode == "journal":
        # Вкладки по заводам. В каждой: список от меньшего номера к большему.
        tabs = tb.Notebook(list_content, bootstyle="secondary")  # или "primary" - что используешь
        tabs.pack(fill=BOTH, expand=True)

        def pad_tab_titles(nb, left=2, right=2):
            em = "\u2003"  # em-space
            for tid in nb.tabs():
                t = nb.tab(tid, "text")
                nb.tab(tid, text=f"{em * left}{t}{em * right}")

        # Шире вкладки, шрифт не трогаем
        pad_tab_titles(tabs, left=2, right=2)
        # тонкая линия под вкладками (даёт чёткий край на белом фоне)
        tb.Separator(list_content, bootstyle="secondary").pack(fill=X, pady=(0, 8))


        for factory in factory_order:  # ["ВЕКТОР", "ИНТЕГРАЛ", ...]
            tab = tb.Frame(tabs, style="Main.TFrame")
            tabs.add(tab, text=factory)

            # Берём только изделия этого завода из архива
            keys_for_factory = [k for k in assembled_products if _split_key_safe(k)[1] == factory]
            # сортировка по номеру изделия ↑
            def _sint(x, default=0):
                try:
                    return int(x)
                except:
                    return default

            keys_for_factory.sort(
                key=lambda k: (_sint(get_year2(k)), _sint(_split_key_safe(k)[0])),
                reverse=True
            )
            inner = tb.Frame(tab, style="Main.TFrame")
            inner.pack(fill=BOTH, expand=True, padx=10, pady=10)

            for key in keys_for_factory:
                prefix = get_product_prefix_from_key(key)  # "СЕРП ВС6Д"/"СЕРП ВС7"
                serial = serial_for_key(key)  # "ггzxxxx"
                asm_ddmm = format_ddmm(storage_dates.get(key, assembly_dates.get(key, "")))
                text = f"{prefix} №{serial} (дата сборки: {asm_ddmm})"

                row = tb.Frame(inner, style="Main.TFrame")
                row.pack(fill=X, pady=ARCHIVE_ROW_PADY)
                lbl = tb.Label(row, text=text, font=ARCHIVE_FONT, style="Main.TLabel")
                lbl.pack(anchor="w")

                # ПКМ «Удалить» как и в истории
                menu = tk.Menu(lbl, tearoff=False)
                menu.add_command(label="Удалить", command=lambda k=key: delete_from_archive(k))

                def _rclick(e, m=menu):
                    try:
                        m.tk_popup(e.x_root, e.y_root)
                    finally:
                        m.grab_release()

                lbl.bind("<Button-3>", _rclick)
                lbl.bind("<Control-Button-1>", _rclick)


    list_canvas.update_idletasks()
    list_canvas.yview_moveto(0)

    # фон
    def update_bg(widget):
        try:
            widget.configure(background=BG_MAIN)
        except:
            pass
        for child in widget.winfo_children():
            update_bg(child)

    update_bg(main_container)


def apply_request_filter():
    global request_filter
    request_filter = filter_var.get()
    update_request_table()


def update_request_table():
    if request_table is None:
        return

    # очистить таблицу
    request_table.delete(*request_table.get_children())

    # четыре колонки
    missing_112, missing_161, missing_114, missing_116 = [], [], [], []

    for key, blocks in products.items():
        # не показываем то, что уже собрано/отправляется
        if key in assembled_products or key in storage_products:
            continue

        number, factory_name, _variant = _split_key_safe(key)

        # применить выбранный фильтр
        if request_filter != "Все":
            if request_filter == "Заготовки":
                if key not in draft_products:
                    continue
            elif request_filter in factory_mapping.values():
                if factory_name != request_filter:
                    continue

        factory_code = factory_reverse_mapping.get(factory_name, "?")
        display = f"{number} [{factory_code}]"

        # учитывать только блоки, которые НУЖНЫ для данного варианта
        needed = set(block_types_for(key))

        # --- НЕ показывать ССБ 112 для ВС7/ВС13 ---
        is_vs7 = str(get_variant(key)).strip() in ("ВС7", "ВС13")

        if "ССБ 112" in needed and not blocks.get("ССБ 112", False):
            if not is_vs7:  # ← ВС7 пропускаем
                missing_112.append(display)

        if "ССБ 161" in needed and not blocks.get("ССБ 161", False):
            missing_161.append(display)
        if "ССБ 114" in needed and not blocks.get("ССБ 114", False):
            missing_114.append(display)
        if "ССБ 116" in needed and not blocks.get("ССБ 116", False):
            missing_116.append(display)

    # сортировка внутри колонок по коду завода
    def sort_by_factory(x: str):
        try:
            return int(x.split("[")[-1].split("]")[0])
        except Exception:
            return 999

    for lst in (missing_112, missing_161, missing_114, missing_116):
        lst.sort(key=sort_by_factory)

    # выводим строки в 4 колонки
    max_len = max(len(missing_112), len(missing_161), len(missing_114), len(missing_116), 0)
    for i in range(max_len):
        request_table.insert(
            "",
            "end",
            values=(
                missing_112[i] if i < len(missing_112) else "",
                missing_161[i] if i < len(missing_161) else "",
                missing_114[i] if i < len(missing_114) else "",
                missing_116[i] if i < len(missing_116) else "",
            ),
        )



def on_tab_change(event):
    save_scroll_positions()
    selected_tab = notebook.tab(notebook.select(), "text")
    if selected_tab == "Запросить блоки":
        update_request_table()
    elif selected_tab == "Архив сборки":
        update_assembly_archive()


def _on_mousewheel(event):
    x, y = root.winfo_pointerx(), root.winfo_pointery()
    widget = root.winfo_containing(x, y)
    while widget:
        if isinstance(widget, tb.Canvas):
            step = 0
            if event.delta:
                step = int(-1 * (event.delta / 120))
            else:
                if event.num == 4:
                    step = -1
                elif event.num == 5:
                    step = 1
            y0, y1 = widget.yview()
            if (step < 0 and y0 <= 0.0) or (step > 0 and y1 >= 1.0):
                return "break"
            widget.yview_scroll(step, "units")
            return "break"
        widget = widget.master


def toggle_draft_group():
    global show_draft_group
    has_drafts = len(draft_products) > 0
    if not has_drafts:
        if show_draft_group:
            show_draft_group = False
            draft_group_button.configure(bootstyle="secondary")
            save_data()
            update_product_list()
        else:
            messagebox.showinfo("Информация", "Нет заготовки для отображения")
        return
    show_draft_group = not show_draft_group
    draft_group_button.configure(
        bootstyle="warning" if show_draft_group else "secondary"
    )
    save_data()
    update_product_list()


def collapse_storage_on_start():
    global storage_visible
    if vs6d_only_mode:
        storage_visible = True
        try:
            storage_frame_container.grid()       # показать колонку
            toggle_storage_button.configure(state="disabled")  # кнопку делаем неактивной
        except Exception:
            pass
        return

    storage_visible = False
    storage_frame_container.grid_remove()
    try:
        toggle_storage_button.configure(state="normal")
    except Exception:
        pass



def toggle_storage_visibility():
    global storage_visible, storage_count_cache

    if vs6d_only_mode:
        try:
            storage_frame_container.grid()
            if toggle_storage_button and toggle_storage_button.winfo_exists():
                toggle_storage_button.configure(state="disabled")
        except Exception:
            pass
        return "break"

    if storage_visible:
        # СКРЫВАЕМ хранилище
        storage_frame_container.grid_remove()
        work7_frame_container.grid()
        storage_visible = False

        if toggle_storage_button and toggle_storage_button.winfo_exists():
            label = f"◀ ({storage_count_cache})"
            toggle_storage_button.config(text=label, width=len(label), state="normal")  # ← добавили state
        if storage_toggle_button and storage_toggle_button.winfo_exists():
            storage_toggle_button.config(text="▶", width=2, state="normal")            # ← добавили state
    else:
        # ПОКАЗЫВАЕМ хранилище
        work7_frame_container.grid_remove()
        storage_frame_container.grid()
        storage_visible = True

        if storage_toggle_button and storage_toggle_button.winfo_exists():
            storage_toggle_button.config(text="▶", width=2, state="normal")            # ← добавили state
        if toggle_storage_button and toggle_storage_button.winfo_exists():
            label = f"◀ ({storage_count_cache})"
            toggle_storage_button.config(text=label, width=len(label), state="normal")  # ← добавили state


def hide_storage_if_open():
    if vs6d_only_mode:
        return
    global storage_visible
    if storage_visible:
        storage_visible = False
        storage_frame_container.grid_remove()
        try:
            toggle_storage_button.configure(text="▶", width=2, state="normal")
        except Exception:
            pass

def _show_xl_updated_info(event=None):
    """
    ПКМ по 📄: показывает, когда обновлялся XL-кэш (top_ssb.jason)
    и когда менялся файл xl.xlsx.
    """
    try:
        lines = []

        # 1) meta из top_ssb.jason
        json_path = globals().get("XL_JSON_PATH")
        if json_path and os.path.isfile(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                meta = data.get("meta", {}) or {}
                updated_at = meta.get("updated_at", "неизвестно")
                src = meta.get("source", "")

                if src:
                    lines.append(f"Источник: {src}")
            except Exception as e:
                lines.append(f"Не удалось прочитать top_ssb: {e}")
        else:
            lines.append("Кэш подсветки (top_ssb) ещё не создан. Нажми 📄 (ЛКМ).")

        # 2) mtime самого xl.xlsx
        xlsx_path = globals().get("XL_XLSX_PATH")
        if xlsx_path and os.path.isfile(xlsx_path):
            try:
                ts = datetime.fromtimestamp(os.path.getmtime(xlsx_path)).strftime("%d.%m.%Y %H:%M:%S")
                lines.append(f"Файл изменён: {ts}")
            except Exception:
                pass

        messagebox.showinfo("XL", "\n".join(lines), parent=globals().get("root"))
    except Exception as e:
        # если тут что-то сломалось — покажем явно
        messagebox.showerror("XL", f"Ошибка обработчика ПКМ:\n{e}", parent=globals().get("root"))

    return "break"



def _is_text_input_widget(w):
    import tkinter as tk
    from tkinter import ttk
    try:
        import ttkbootstrap as tb
        tb_entry = tb.Entry
        tb_combo = tb.Combobox
    except Exception:
        tb_entry = tb_combo = object  # запасной вариант, чтобы isinstance не упал

    # не перехватываем пробел, когда печатаешь в полях ввода/комбо
    return isinstance(
        w,
        (tk.Entry, ttk.Entry, tk.Text, ttk.Combobox, tb_entry, tb_combo),
    )

def on_space_toggle(event=None):
    # если фокус в поле ввода — даём поставить пробел и выходим
    focus = root.focus_get()
    if focus and _is_text_input_widget(focus):
        return
    toggle_storage_visibility()
    return "break"  # чтобы пробел больше ничего не делал



def center_window(initial=False):
    try:
        if initial:
            w = max(100, root.winfo_reqwidth())
            h = max(100, root.winfo_reqheight())
            if w < 200 or h < 200:
                geom = root.geometry()
                if "x" in geom:
                    wh = geom.split("+")[0]
                    w2, h2 = [int(x) for x in wh.split("x")]
                    w, h = max(w, w2), max(h, h2)
        else:
            root.update_idletasks()
            w = max(100, root.winfo_width())
            h = max(100, root.winfo_height())
            if w < 200 or h < 200:
                w = max(w, root.winfo_reqwidth())
                h = max(h, root.winfo_reqheight())
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        root.geometry(f"{w}x{h}+{x}+{y}")
    except Exception as e:
        print("center_window() error:", e)


load_data()
_load_block112_index()
migrate_keys_to_4()
work_sort_mode = "factories"
show_draft_group = False
blockcount_mode = False
root = tb.Window(themename="litera")
root.tk.call("tk", "scaling", scaling_factor)
root.title("serp-base")
root.withdraw()
root.geometry("600x600")
root.minsize(600, 600)
root.configure(background=BG_MAIN)

display_full_serials_var = tk.BooleanVar(master=root, value=display_full_serials)                  # Рабочая зона
display_full_serials_storage_var = tk.BooleanVar(master=root, value=display_full_serials_storage)

try:
    root.iconbitmap(resource_path("new.ico"))
except:
    print("Иконка не найдена")
style = tb.Style()
TAB_BG          = "#f4f6f9"   # фон обычной вкладки
TAB_BG_ACTIVE   = "#ecf2ff"   # при наведении
TAB_BG_SELECTED = "#dfe7f7"   # выбранная
TAB_FG          = "#1f2937"   # тёмный текст

style.configure("Wide.TNotebook",
                background=BG_MAIN,    # фон области вокруг вкладок
                tabmargins=(14, 8, 14, 0))

style.configure("Wide.TNotebook.Tab",
                padding=(40, 20),      # ширина вкладок
                background=TAB_BG,
                foreground=TAB_FG)


# динамика: выбранная/активная
style.map("Wide.TNotebook.Tab",
          background=[("selected", TAB_BG_SELECTED),
                      ("active",   TAB_BG_ACTIVE)],
          foreground=[("selected", TAB_FG), ("active", TAB_FG)])

style.configure("Main.TFrame", background=BG_MAIN)
style.configure("WorkArea.TFrame", background=BG_WORK_AREA)
style.configure(
    "Main.TLabelframe",
    background=BG_MAIN,
    borderwidth=1,   # тонкая рамка
)

style.configure(
    "Main.TLabelframe.Label",
    background=BG_MAIN,
    foreground=TEXT_MAIN,
    font=FONT_PRIMARY,
)

style.configure("Header.TFrame", background=BG_HEADER)
style.configure("Settings.TFrame", background=BG_SETTINGS)
style.configure("Wide.Vertical.TScrollbar", width=14)
style.configure("Save.TButton", font=FONT_SAVE_BUTTON)
style.configure("Status.TButton", font=FONT_STATUS_BUTTON)
style.configure("Block.TButton", font=FONT_BLOCK)
style.configure(
    "Header1.TLabel", font=FONT_HEADER1, background=BG_MAIN, foreground=TEXT_HEADER
)
style.configure("Settings.TCheckbutton", background=BG_SETTINGS, foreground=TEXT_MAIN)
style.configure(
    "Header2.TLabel", font=FONT_HEADER2, background=BG_MAIN, foreground=TEXT_HEADER
)
style.configure(
    "Header3.TLabel", font=FONT_HEADER3, background=BG_MAIN, foreground=TEXT_DATE
)
style.configure(
    "FactoryHeader.TLabel",
    font=FONT_FACTORY,
    background=BG_MAIN,
    foreground=TEXT_FACTORY,
)
style.configure(
    "Main.TLabel", background=BG_MAIN, foreground=TEXT_MAIN, font=FONT_PRIMARY
)
style.configure(
    "WorkArea.TLabel", background=BG_WORK_AREA, foreground=TEXT_MAIN, font=FONT_PRIMARY
)
style.configure(
    "Bold.TLabel", font=FONT_SETTINGS, background=BG_MAIN, foreground=TEXT_HEADER
)
style.configure(
    "BlockCounter.TLabel",
    font=FONT_COUNTER,
    foreground=TEXT_COUNTER,
    background=BG_MAIN,
)
notebook = tb.Notebook(root, bootstyle="primary")
notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
frame_main = tb.Frame(notebook, style="Main.TFrame")
frame_assembly = tb.Frame(notebook, style="Main.TFrame")
frame_request = tb.Frame(notebook, style="Main.TFrame")
frame_settings = tb.Frame(notebook, style="Settings.TFrame")
notebook.add(frame_main, text="     Основной учёт     ")
notebook.add(frame_assembly, text="    Архив сборки    ")
notebook.add(frame_request, text="     Запросить блоки     ")
notebook.add(frame_settings, text="     Настройки     ")
request_panel = tb.Frame(frame_request, style="Main.TFrame")
request_panel.pack(fill=BOTH, expand=True, padx=10, pady=10)
filter_frame = tb.Frame(request_panel, style="Main.TFrame")
filter_frame.pack(fill=X, pady=(0, 10))
tb.Label(filter_frame, text="Фильтр:", style="Main.TLabel").pack(side=LEFT, padx=5)
filter_var = tb.StringVar(value="Все")
filter_combo = tb.Combobox(
    filter_frame,
    textvariable=filter_var,
    values=[
        "Все",
        "Заготовки",
        "ВЕКТОР",
        "ИНТЕГРАЛ",
        "РЗП",
        "КНИИТМУ",
        "СВТ",
        "СИГНАЛ",
    ],
    state="readonly",
    width=15,
)
filter_combo.pack(side=LEFT, padx=5)
apply_btn = tb.Button(
    filter_frame, text="Применить", bootstyle="secondary", command=apply_request_filter
)
apply_btn.pack(side=LEFT, padx=5)
columns = ("Блок питания ССБ112", "Детектор ССБ161", "Колпак ССБ114", "Колпак ССБ116")

request_table = tb.Treeview(
    request_panel, columns=columns, show="headings", height=25, bootstyle="light"
)

style = tb.Style()
style.configure(
    "Treeview",
    background=BG_WORK_AREA,
    foreground=TEXT_MAIN,
    fieldbackground=BG_WORK_AREA,
    font=FONT_PRIMARY,
    rowheight=60,
)
style.configure(
    "Treeview.Heading", background=BG_HEADER, foreground=TEXT_HEADER, font=FONT_BOLD
)
style.map("Treeview", background=[("selected", BG_HEADER)])

# подберите ширины при необходимости
request_table.heading("Блок питания ССБ112", text="Блок питания ССБ112")
request_table.heading("Детектор ССБ161", text="Детектор ССБ161")
request_table.heading("Колпак ССБ114", text="Колпак ССБ114")
request_table.heading("Колпак ССБ116", text="Колпак ССБ116")

request_table.column("Блок питания ССБ112", width=200, anchor=CENTER)
request_table.column("Детектор ССБ161", width=200, anchor=CENTER)
request_table.column("Колпак ССБ114", width=200, anchor=CENTER)
request_table.column("Колпак ССБ116", width=200, anchor=CENTER)

request_table.pack(fill=BOTH, expand=True)

try:
    # сразу заполнить таблицу текущим фильтром ("Все")
    apply_request_filter()
except Exception:
    pass

def apply_up112():
    base_dir = r"C:\ssb_data\global_data"
    candidates = [
        os.path.join(base_dir, "up_112"),
        os.path.join(base_dir, "up_112.json"),
        os.path.join(base_dir, "up_112.txt"),
    ]
    path = next((p for p in candidates if os.path.isfile(p)), None)
    if path is None:
        messagebox.showerror(
            "Файл не найден",
            "Не найден файл с номерами 112. Создайте один из вариантов:\n"
            + "\n".join(candidates)
            + '\n\nФорматы: JSON вида {"nums":["10012","10023"]} или текст по строке.',
        )
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read().strip()
        if raw.startswith("{"):
            data = json.loads(raw)
            nums = data.get("nums", [])
        else:
            nums = [line.strip() for line in raw.splitlines() if line.strip()]
    except Exception as e:
        messagebox.showerror("Ошибка чтения", f"Не удалось прочитать {path}:\n{e}")
        return
    touched = 0
    for s in nums:
        if not s.isdigit() or len(s) < 5:
            continue
        number_part = s[:4]
        factory_code = s[4]
        if factory_code not in factory_mapping:
            continue
        factory_name = factory_mapping[factory_code]
        matches = [
            k for k in products
            if len(k) >= 2 and k[0] == number_part and k[1] == factory_name
               and k not in storage_products and k not in assembled_products
               and not products.get(k, {}).get("ССБ 112", False)
        ]
        for mk in matches:
            up112_hints.add(key2to3(mk))  # подсказка по 3-ключу (для всех годов)
        touched += len(matches)

    if touched == 0:
        messagebox.showinfo("UP 112", "Нет новых блоков ССБ 112.")
    update_product_list()

request_scrollbar = tb.Scrollbar(
    request_panel,
    orient=VERTICAL,
    command=request_table.yview,
    style="Wide.Vertical.TScrollbar",
)
request_table.configure(yscrollcommand=request_scrollbar.set)
request_scrollbar.pack(side=RIGHT, fill=Y)

entry_frame_container = tb.Frame(frame_main, style="Main.TFrame")
entry_frame_container.pack(fill=X, pady=10, padx=10)

entry_header = tb.Frame(entry_frame_container, style="Header.TFrame")
entry_header.pack(fill="x", pady=(0, PADY_HEADER))

# короткий заголовок
entry_header_label = tb.Label(
    entry_header,
    text="Добавить СЕРП:",
    style="Header2.TLabel",
)
entry_header_label.pack(side="left", padx=10, pady=5)

# переключатели модели прямо в этой строке

vs6d_btn = tb.Button(
    entry_header,
    text="ВС6",
    width=5,
    bootstyle="primary" if product_mode == "ВС6Д" else "secondary",
    command=lambda: set_product_mode("ВС6Д"),
)
vs6d_btn.pack(side="left", padx=(8, 2), pady=5)

vs13_btn = tb.Button(
    entry_header,
    text="ВС7",
    width=5,
    bootstyle="primary" if product_mode == "ВС7" else "secondary",
    command=lambda: set_product_mode("ВС7"),
)
vs13_btn.pack(side="left", padx=(2, 10), pady=5)

# действия справа (как и было): [ Сохранить ] и [12]
# действия справа: 3 кнопки в ряд (Excel / Поиск / Сохранить) — одинаковая ширина и отступы
header_actions = tb.Frame(entry_header, style="Header.TFrame")
header_actions.pack(side="right", padx=10, pady=5)

# 3 одинаковые колонки под кнопки
for col in range(3):
    header_actions.grid_columnconfigure(col, weight=1, uniform="hdrbtn", minsize=48)
header_actions.grid_rowconfigure(0, weight=1)

BTN_PAD = (10, 8)   # внутренние отступы кнопки (можешь чуть подстроить)
BTN_GAP = 4         # внешний отступ между кнопками

# 1) Excel (📄) — слева
xl_button_top = tb.Button(
    header_actions,
    text="📄",
    width=3,
    bootstyle="success",
    padding=BTN_PAD,
    command=_safe_apply_xl,   # ЛКМ как и было
)
xl_button_top.grid(row=0, column=0, padx=(0, BTN_GAP), sticky="nsew")
try:
    xl_button_top.unbind("<Button-3>")
except Exception:
    pass

# ПКМ -> окно настроек чтения Excel
xl_button_top.bind("<Button-3>", _open_xl_settings_popup)

# обновим цвет по давности файла
update_xl_button_visual()



# 2) Поиск (🔍) — по центру
search_button_top = tb.Button(
    header_actions,
    text="🔍",
    width=3,
    bootstyle="secondary",
    padding=BTN_PAD,
    command=open_search_popup,  # ВАЖНО: поставь сюда свою функцию кнопки поиска
)
search_button_top.grid(row=0, column=1, padx=BTN_GAP, sticky="nsew")

# 3) Сохранить (💾) — справа
btn_save_top = tb.Button(
    header_actions,
    text="💾",
    width=3,
    bootstyle="primary",      # или "secondary" если нужно серым
    padding=BTN_PAD,
    command=lambda: save_data(True),
)
btn_save_top.grid(row=0, column=2, padx=(BTN_GAP, 0), sticky="nsew")

entry_panel = tb.Frame(entry_frame_container, style="Main.TFrame")
entry_panel.pack(fill=X)

input_row = tb.Frame(entry_panel, style="Main.TFrame")
input_row.pack(fill=X, pady=(0, 10))
input_frame = tb.Frame(input_row, style="Main.TFrame")
input_frame.pack(side=LEFT, fill=X, expand=True)
tb.Label(input_frame, text=TEXT_YEAR, font=FONT_PRIMARY, style="Main.TLabel").pack(
    side=LEFT, padx=5
)

year_var = tb.StringVar()
year_combo = tb.Combobox(
    input_frame, textvariable=year_var, font=FONT_PRIMARY, width=5, bootstyle="light"
)
year_combo["values"] = [f"{y:02}" for y in range(24, 100)]
try:
    year_combo.set(last_selected_year)
    year_var.set(last_selected_year)
except Exception:
    year_combo.current(1)
year_combo.pack(side=LEFT, padx=5)
year_combo.bind("<<ComboboxSelected>>", on_year_change)

tb.Label(input_frame, text=TEXT_NUMBER, font=FONT_PRIMARY, style="Main.TLabel").pack(
    side=LEFT, padx=5
)
entry_var = tb.StringVar()
num_box = _wrap_with_border(input_frame)
num_box.pack(side=LEFT, padx=5)

entry = tb.Entry(
    num_box, textvariable=entry_var, font=FONT_PRIMARY, width=10, bootstyle="light"
)
entry.pack(fill=X, padx=4, pady=3)
entry.bind("<Return>", lambda e: process_serial())

add_button = tb.Button(
    input_frame,
    text=TEXT_ADD_SERP,
    width=3,
    bootstyle="primary",
    padding=(5, 5),
    command=process_serial,
)
add_button.pack(side=LEFT, padx=5)

########

save_row = tb.Frame(entry_panel, style="Main.TFrame")
save_row.pack(fill=X, pady=(5, 10))

sort_mode_button = tb.Button(
    save_row,
    text="Сортировка",
    width=12,
    bootstyle="secondary",   # старт: режим заводы = серая
    command=cycle_sort_mode,
)
sort_mode_button.pack(side=LEFT, padx=5)

# 2) ОБНОВИТЬ (значок), просто перерисовать списки с учётом текущих правок
refresh_button = tb.Button(
    save_row,
    text="⟳",
    width=3,
    bootstyle="secondary",
    command=refresh_lists,
)
refresh_button.pack(side=LEFT, padx=5)
apply_vs6d_only_mode_ui()



# Подпись циклической кнопки привести к текущему режиму
update_sort_mode_button_visual()
# search_var нужен для search_product() (теперь ввод идёт из popup)
search_var = tb.StringVar()


main_area = tb.Frame(frame_main, style="Main.TFrame")
main_area.pack(fill=BOTH, expand=True, padx=10, pady=5)
main_area.columnconfigure(0, weight=3)  # ВС6Д шире
main_area.columnconfigure(1, weight=2)
main_area.rowconfigure(0, weight=1)

work_frame_container = tb.Frame(main_area, style="Main.TFrame")
work_frame_container.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=0)
work_frame_container.grid_rowconfigure(1, weight=1)

work_header = tb.Frame(work_frame_container, style="Header.TFrame")
work_header.pack(fill="x", pady=(0, PADY_HEADER))
work_header_label = tb.Label(work_header, text=TEXT_WORK_TITLE, style="Header1.TLabel")
work_header_label.pack(side="left", padx=10, pady=5)
update_vs13_button_state()
# рамка только вокруг содержимого
work_border = tb.LabelFrame(work_frame_container, padding=6, style="Main.TLabelframe")
work_border.pack(fill="both", expand=True, pady=(2, 0))

work_container = tb.Frame(work_border, style="WorkArea.TFrame")
work_container.pack(fill="both", expand=True)

# ===== ПРАВАЯ рабочая колонка: ВС7 =====
work7_frame_container = tb.Frame(main_area, style="Main.TFrame")
work7_frame_container.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=0)
work7_frame_container.grid_rowconfigure(1, weight=1)

work7_header = tb.Frame(work7_frame_container, style="Header.TFrame")
work7_header.pack(fill="x", pady=(0, PADY_HEADER))

work7_header_label = tb.Label(work7_header, text="ВС7 в работе", style="Header1.TLabel")
work7_header_label.pack(side="left", padx=10, pady=5)

# стрелка с количеством — всегда видна справа в шапке ВС7
toggle_storage_button = tb.Button(
    work7_header, text="◀ (0)", width=6, command=toggle_storage_visibility
)
toggle_storage_button.pack(side="right", padx=5, pady=5)

work7_border = tb.LabelFrame(work7_frame_container, padding=6, style="Main.TLabelframe")
work7_border.pack(fill="both", expand=True, pady=(2, 0))

work7_container = tb.Frame(work7_border, style="WorkArea.TFrame")
work7_container.pack(fill="both", expand=True)


work7_canvas = tb.Canvas(work7_container, background=BG_WORK_AREA)
work7_scrollbar = tb.Scrollbar(
    work7_container,
    orient=VERTICAL,
    command=work7_canvas.yview,
    bootstyle="dark-round",
    style="Wide.Vertical.TScrollbar",
)
work7_frame = tb.Frame(work7_canvas, style="WorkArea.TFrame")
work7_frame.bind(
    "<Configure>",
    lambda e: work7_canvas.configure(scrollregion=work7_canvas.bbox("all")),
)
work7_canvas.create_window((0, 0), window=work7_frame, anchor=NW)
work7_canvas.configure(yscrollcommand=work7_scrollbar.set, background=BG_WORK_AREA)
work7_scrollbar.pack(side=RIGHT, fill=Y)
work7_canvas.pack(side=LEFT, fill=BOTH, expand=True)
work7_canvas.config(height=300)

work_canvas = tb.Canvas(work_container, background=BG_WORK_AREA)
work_scrollbar = tb.Scrollbar(
    work_container,
    orient=VERTICAL,
    command=work_canvas.yview,
    bootstyle="dark-round",
    style="Wide.Vertical.TScrollbar",
)
work_frame = tb.Frame(work_canvas, style="WorkArea.TFrame")
work_frame.bind(
    "<Configure>", lambda e: work_canvas.configure(scrollregion=work_canvas.bbox("all"))
)
work_canvas.create_window((0, 0), window=work_frame, anchor=NW)
work_canvas.configure(yscrollcommand=work_scrollbar.set, background=BG_WORK_AREA)
work_scrollbar.pack(side=RIGHT, fill=Y)
work_canvas.pack(side=LEFT, fill=BOTH, expand=True)
work_canvas.config(height=300)
storage_frame_container = tb.Frame(main_area, style="Main.TFrame")
storage_frame_container.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=0)
storage_frame_container.grid_rowconfigure(1, weight=1)

# --- Шапка хранилища (пересобрано) ---
storage_header = tb.Frame(storage_frame_container, style="Header.TFrame")
storage_header.pack(fill="x", pady=(0, PADY_HEADER))

# 1) Слева: стрелка ЗАКРЫТЬ склад (перед заголовком)
storage_toggle_button = tb.Button(
    storage_header, text="▶", width=2, command=toggle_storage_visibility
)
storage_toggle_button.pack(side="left", padx=(10, 6), pady=5)

# 2) Заголовок — сразу за стрелкой
storage_header_label = tb.Label(
    storage_header, text=TEXT_STORAGE_TITLE, style="Header1.TLabel"
)
storage_header_label.pack(side="left", padx=(0, 10), pady=5)

# 3) Растяжка, чтобы увести кнопку вправо
tb.Frame(storage_header).pack(side="left", fill="x", expand=True)


storage_sort_button = tb.Button(
    storage_header,
    text="Сортировка",          # подпись обновим ниже через update_storage_sort_button_visual()
    width=12,
    bootstyle="secondary",      # по умолчанию "по датам" — синий
    command=lambda: cycle_storage_sort_mode(),
)
storage_sort_button.pack(side="right", padx=(0, 6), pady=5)

# 4) Справа: «Отправить все» (на месте старой правой стрелки)
send_all_storage_button = tb.Button(
    storage_header,
    text="Отправить все",
    bootstyle="secondary",
    command=send_all_to_storage,
)
send_all_storage_button.pack(side="right", padx=(0, 10), pady=5)


def _measure_btn(text, width=None, bootstyle="secondary"):
    tmp = tb.Button(storage_header, text=text, width=width, bootstyle=bootstyle)
    tmp.update_idletasks()
    w = tmp.winfo_reqwidth()
    tmp.destroy()
    return w

_right_offset = _measure_btn("Склад", width=8) + _measure_btn("✕", width=3) + 16  # +поля

# Сетку под шапку:

##

update_vs13_button_state()
storage_border = tb.LabelFrame(                         # ← рамка как у ВС6/ВС7
    storage_frame_container, padding=6, style="Main.TLabelframe"
)
storage_border.pack(fill="both", expand=True, pady=(2, 0))

storage_container = tb.Frame(storage_border, style="WorkArea.TFrame")
storage_container.pack(fill="both", expand=True)
storage_canvas = tb.Canvas(storage_container, background=BG_WORK_AREA)
storage_scrollbar = tb.Scrollbar(
    storage_container,
    orient=VERTICAL,
    command=storage_canvas.yview,
    bootstyle="dark-round",
    style="Wide.Vertical.TScrollbar",
)
storage_frame = tb.Frame(storage_canvas, style="WorkArea.TFrame")
storage_frame.bind(
    "<Configure>",
    lambda e: storage_canvas.configure(scrollregion=storage_canvas.bbox("all")),
)
storage_canvas.create_window((0, 0), window=storage_frame, anchor=NW)
storage_canvas.configure(yscrollcommand=storage_scrollbar.set, background=BG_WORK_AREA)
storage_scrollbar.pack(side=RIGHT, fill=Y)
storage_canvas.pack(side=LEFT, fill=BOTH, expand=True)
storage_canvas.config(height=300)
storage_frame_container.grid_remove()
update_vs13_button_state()
archive_frame = tb.Frame(frame_assembly, style="Main.TFrame")
archive_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
factory_buttons = {}
dpi_var = None
scaling_var = None


def apply_scaling():
    global scaling_factor
    new_scaling = scaling_var.get()
    scaling_factor = new_scaling
    save_data()
    messagebox.showinfo(
        "Масштабирование",
        "Масштабирование будет применено после перезапуска приложения.",
    )

def create_settings_tab():
    global scaling_var
    settings_container = tb.Frame(frame_settings, style="Settings.TFrame")
    settings_container.pack(fill=BOTH, expand=True, padx=20, pady=20)

    # (опциональный) пустой блок — можно оставить или удалить
    mode_frame = tb.Frame(settings_container, style="Settings.TFrame")
    mode_frame.pack(fill=X, pady=(0, 10))
    # === Отображение (крупный шрифт, фон = BG_MAIN) ===
    try:
        _s = tb.Style()
        # Лейблфрейм + его заголовок (лейбл)
        _s.configure("SettingsBig.TLabelframe",
                     font=("Arial", 12, "bold"),
                     background=BG_MAIN)
        _s.configure("SettingsBig.TLabelframe.Label",
                     font=("Arial", 12, "bold"),
                     background=BG_MAIN,
                     foreground=TEXT_MAIN)
        # Контейнеры и текстовые элементы
        _s.configure("SettingsBig.TFrame", background=BG_MAIN)
        _s.configure("SettingsBig.TLabel", font=("Arial", 11),
                     background=BG_MAIN, foreground=TEXT_MAIN)
        _s.configure("SettingsBig.Subheader.TLabel", font=("Arial", 11, "bold"),
                     background=BG_MAIN, foreground=TEXT_MAIN)
        # Чекбоксы
        _s.configure("SettingsBig.TCheckbutton", font=("Arial", 11),
                     background=BG_MAIN, foreground=TEXT_MAIN)
    except Exception:
        pass

    display_group = tb.LabelFrame(
        settings_container,
        text="Отображение",
        padding=18,
        style="SettingsBig.TLabelframe",
        bootstyle="primary",
    )
    display_group.pack(fill=X, pady=12)

    display_inner = tb.Frame(display_group, style="SettingsBig.TFrame")
    display_inner.pack(fill=X, expand=True)

    # Привязки переменных
    global display_full_serials_var, display_full_serials_storage_var, vs6d_only_mode_var, hide_added_date_var
    display_full_serials_var = tk.BooleanVar(value=bool(display_full_serials))
    display_full_serials_storage_var = tk.BooleanVar(value=bool(display_full_serials_storage))
    vs6d_only_mode_var = tk.BooleanVar(value=bool(vs6d_only_mode))
    hide_added_date_var = tk.BooleanVar(value=bool(hide_added_date))

    # Подзаголовок
    tb.Label(
        display_inner,
        text="Полный серийный номер:",
        style="SettingsBig.Subheader.TLabel",
    ).pack(anchor="w", padx=8, pady=(0, 6))

    # Галочки
    tb.Checkbutton(
        display_inner,
        text="Рабочая зона (ВС6Д / ВС7)",
        variable=display_full_serials_var,
        command=on_toggle_full_serials_work,
        bootstyle="secondary-round-toggle",
        style="SettingsBig.TCheckbutton",
        padding=6,
    ).pack(anchor="w", padx=16, pady=(0, 6))

    tb.Checkbutton(
        display_inner,
        text="Хранилище",
        variable=display_full_serials_storage_var,
        command=on_toggle_full_serials_storage,
        bootstyle="secondary-round-toggle",
        style="SettingsBig.TCheckbutton",
        padding=6,
    ).pack(anchor="w", padx=16, pady=(0, 10))

    tb.Checkbutton(
        display_inner,
        text="Скрыть дату добавления",
        variable=hide_added_date_var,
        command=on_toggle_hide_added_date,
        bootstyle="secondary-round-toggle",
        style="SettingsBig.TCheckbutton",
        padding=6,
    ).pack(anchor="w", padx=16, pady=(0, 10))

    # Разделитель
    tb.Separator(display_inner, bootstyle="secondary").pack(fill=X, padx=6, pady=8)

    # Режим ВС6Д
    tb.Checkbutton(
        display_inner,
        text="Режим VS6D",
        variable=vs6d_only_mode_var,
        command=on_toggle_vs6d_mode,
        bootstyle="secondary-round-toggle",
        style="SettingsBig.TCheckbutton",
        padding=6,
    ).pack(anchor="w", padx=8, pady=(0, 4))

    # ----- Масштабирование -----
    scaling_frame = tb.LabelFrame(
        settings_container,
        text="Масштабирование интерфейса",
        bootstyle="primary",
        padding=15,
        style="Settings.TLabelframe",
    )
    scaling_frame.pack(fill=X, pady=10)

    scale_value_label = tb.Label(
        scaling_frame,
        text=f"Текущее значение: {scaling_factor:.1f}",
        style="Settings.TLabel",
    )
    scale_value_label.pack(anchor=W, pady=5)

####


    tb.Label(
        scaling_frame,
        text="Коэффициент масштабирования (рекомендуется 1.5–2.5):",
        style="Settings.TLabel",
    ).pack(anchor=W, pady=5)

    scaling_var = tb.DoubleVar(value=scaling_factor)
    tb.Scale(
        scaling_frame,
        from_=1.0,
        to=4.0,
        orient=tk.HORIZONTAL,
        variable=scaling_var,
        bootstyle="primary",
        command=lambda v: scale_value_label.config(text=f"Текущее значение: {float(v):.1f}"),
    ).pack(fill=X, padx=5, pady=5)

    tb.Button(
        scaling_frame,
        text="Применить масштабирование",
        bootstyle="secondary",
        command=apply_scaling,
    ).pack(side=LEFT, anchor=W, padx=5, pady=5)



    # ----- Кнопка сохранения -----
    control_frame = tb.Frame(settings_container, style="Settings.TFrame")
    control_frame.pack(fill=X, pady=20)

    tb.Button(
        control_frame,
        text="Сохранить настройки",
        width=20,
        bootstyle="success",
        style="Bold.TButton",
        padding=5,
        command=save_settings,
    ).pack(side=RIGHT, padx=10, pady=10)
    update_date_label = tb.Label(
        settings_container,
        text= "Дата обновления: 23.12.2025",
        style="Settings.TLabel",
        font=("Arial", 10, "italic"),
    )
    update_date_label.pack(side=BOTTOM, anchor="e", padx=10, pady=5)
    tb.Label(
        settings_container,
        text="Версия программы: 13.1.7",
        style="Settings.TLabel",
        font=("Arial", 10, "italic"),
    ).pack(side=BOTTOM, anchor="e", padx=10, pady=5)



style.configure("Settings.TFrame", background=BG_SETTINGS)
style.configure(
    "Settings.TLabel", background=BG_SETTINGS, foreground=TEXT_MAIN, font=FONT_PRIMARY
)
style.configure(
    "SectionHeader.TLabel",
    font=("Arial", 12, "bold"),          # на 1 больше, чем сейчас у Header3 (11)
    background=BG_MAIN,
    foreground=TEXT_DATE         # тёмно-серый (#555555)
)
style.configure(
    "Settings.TRadiobutton",
    background=BG_SETTINGS,
    foreground=TEXT_MAIN,
    font=FONT_PRIMARY,
)
style.configure("Settings.TLabelframe", background=BG_SETTINGS, foreground=TEXT_MAIN)
style.configure(
    "Settings.TLabelframe.Label",
    background=BG_SETTINGS,
    foreground=TEXT_MAIN,
    font=FONT_BOLD,
)
notebook.bind("<<NotebookTabChanged>>", on_tab_change)
root.bind("<MouseWheel>", _on_mousewheel)


apply_vs6d_mode_visual()
root.bind("<Button-4>", _on_mousewheel)
root.bind("<Button-5>", _on_mousewheel)

create_settings_tab()
collapse_storage_on_start()  # сначала сворачиваем хранилище и ставим «◀ (N)»
update_product_list()  # теперь строим списки и пересчёт количества
update_assembly_archive()
update_button_colors()

# --- ЕДИНЫЙ обработчик пробела: открывать/закрывать хранилище, КРОМЕ VS6Д-only ---
def on_space_toggle(event=None):
    # пропускаем пробел, если печатают в поле ввода
    focus = root.focus_get()
    if focus and _is_text_input_widget(focus):
        return  # не перехватываем — пусть вставится пробел

    if vs6d_only_mode:
        # в режиме ВС6Д пробел блокируем (ничего не делаем)
        return "break"

    # обычный режим: переключаем видимость хранилища
    toggle_storage_visibility()
    return "break"



# Сначала снимаем любые старые бинды, затем вешаем один-единственный
try:
    root.unbind_all("<space>")
except Exception:
    pass



root.withdraw()
root.update_idletasks()
center_window(initial=True)
root.eval("tk::PlaceWindow . center")

root.deiconify()



def on_closing():
    save_all_comments()
    save_data()
    root.destroy()


root.protocol("WM_DELETE_WINDOW", on_closing)
last_window_state = root.state()
root.mainloop()
