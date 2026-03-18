import tkinter as tk
from tkinter import messagebox, Toplevel, filedialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from datetime import datetime
import threading
import time
import pyperclip
import json
import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import subprocess
import platform
import shutil
import keyboard
import re



FONT_SIZES = {"title": 28, "label": 16, "entry": 20, "button": 18, "serial": 20, "list_item": 20, "status": 18, "table": 14}
BUTTON_SIZES = {"width": 24, "padding": 12}
WINDOW_SIZE = "450x600"
SMALL_WINDOW_SIZE = "750x900"
ENTRY_WIDTH = 16
FACTORIES = {"1": "ВЕКТОР", "2": "ИНТЕГРАЛ", "3": "РЗП", "4": "КНИИТМУ", "5": "СВТ", "6": "СИГНАЛ"}
THRESHOLDS = {"1": 126, "2": 270, "3": 4, "4": 60, "5": 70, "6": 119}
PING_RESTORE_MS = 10_000
STATUS_MODE = "base"
PING_ATTEMPTS = 3
PING_GAP_SEC = 1.0

EXCEL_WIDTH_FACTOR = 1.6   # коэффициент «щедрости» ширины (было по сути 1.0)
EXCEL_WIDTH_CAP    = 120   # максимальная ширина колонки (в «символьных» единицах Excel)
EXCEL_MIN_WIDTH    = 12    # минимальная ширина «по умолчанию»
# Минимальные ширины для конкретных столбцов, чтобы наверняка всё влазило в одну строку
EXCEL_MIN_BY_HEADER = {
    "Примечания":      30,
    "Проверяющий":    18,
    "Зав. № (бирка, flash)":  18,
    "Завод":           16,
    "Версия прошивки ССФ1121":        15,
    "Цвет блока":      20,
    "MAC":             16,
    "Дата проверки":            20,
}

checked_blocks = []
global_blocks_cache = []
checked_serials_set = set()
full_counts = {}         # {user: int}
full_serial_owner = {}   # {serial: user}
current_serial = ""
copied_flag = False
current_year = datetime.now().strftime("%y")
current_user = ""
root_dir = Path("C:/ssb_data")
EXCEL_BOOK_PATH   = root_dir / "PNB_учёт.xlsx"
EXCEL_FILE = EXCEL_BOOK_PATH
EXCEL_PASSWORD    = "PNB"              # пароль защиты листов
EXCEL_FONT_NAME   = "Times New Roman"
EXCEL_FONT_SIZE   = 12
global_data_dir = root_dir / "global_data"
global_backup_dir = root_dir / "global_backup"
user_backup_dir = None
status_restore_after_id = None
global_history_file = global_data_dir / "global_history.json"
global_config_file = root_dir / "config_global.json"
full_count_file = root_dir / "full_count.json"
export_settings_file = root_dir / "export_settings.json"
up112_file = global_data_dir / "up_112.json"

user_data_file = None
global_data_loaded = False
last_checked_raw = None
space_pressed = False
ctrl_down = False
ctrl_combo = False
PING_BLINKING = False
PING_BLINK_AFTER_ID = None
highlight_tag_configured = False
window_size_large = True
ip_address = "192.168.10.25"
serp_list_file = root_dir / "serp_list_112.json"
last_copied_serial = None
counter_file = Path("C:/ssb_data/counter_name.json")
user_counters = {}
user_monthly_archives = {}
final_list_window = None

export_settings = {
    "firmware": "v. 1.6",
    "block_color": "серый",
    "notes": "проверка на новом стенде",
    "mac_prefix": "B8:17:22:ED:00",
}

try:
    import winsound
    SOUND_ENABLED = True
except ImportError:
    SOUND_ENABLED = False

def center_and_focus(win: tk.Toplevel | tb.Window, modal: bool = False):
    try:
        win.update_idletasks()
        w = win.winfo_width()
        h = win.winfo_height()
        if w <= 1 or h <= 1:
            geom = win.geometry().split("+")[0]
            try:
                w, h = map(int, geom.split("x"))
            except Exception:
                w, h = (600, 400)
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 3
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.lift()
        win.focus_force()
        if modal:
            try:
                win.transient(root)
            except Exception:
                pass
            win.grab_set()
    except Exception:
        pass

def play_sound(sound_type):
    if not SOUND_ENABLED: return
    try:
        if sound_type == "copy": winsound.Beep(1000, 100)
        elif sound_type == "add": winsound.Beep(800, 150)
        elif sound_type == "error": winsound.Beep(400, 300)
    except: pass

def only_digits_spaces(new_value: str):
    return all(ch.isdigit() or ch == " " for ch in new_value)

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'): return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def parse_input(raw: str):
    raw = raw.replace(" ", "")
    if len(raw) < 2 or len(raw) > 5 or not raw.isdigit(): return None
    factory_code = raw[-1]
    block_num = raw[:-1].zfill(4)
    return factory_code, block_num

def build_serial(factory_code: str, block4: str):
    return f"12{current_year}{factory_code}{block4}"

def generate_serial(raw_number: str):
    parsed = parse_input(raw_number)
    if not parsed: return None, None, None, None
    factory_code, block4 = parsed
    factory_name = FACTORIES.get(factory_code, "НЕИЗВЕСТНЫЙ ЗАВОД")
    serial = build_serial(factory_code, block4)
    return serial, factory_name, factory_code, block4

def check_block_threshold(factory_code: str, block_num: str):
    if factory_code not in THRESHOLDS: return False
    try:
        block_num_int = int(block_num or "0")
        return block_num_int <= THRESHOLDS[factory_code]
    except ValueError:
        return False

def get_global_block_info(serial: str):
    for block in global_blocks_cache:
        if block.get("full_serial") == serial: return block
    return None

def check_user_exists(): return bool(current_user)

def check_global_data_loaded():
    if not global_data_loaded:
        play_sound("error")
        messagebox.showerror("Ошибка", "Глобальные данные не загружены!")
        return False
    return True

def init_global_directories():
    global root_dir, global_data_dir, global_backup_dir, global_history_file, global_config_file
    root_dir.mkdir(exist_ok=True)
    global_data_dir.mkdir(exist_ok=True)
    global_backup_dir.mkdir(exist_ok=True)
    if not global_config_file.exists():
        with open(global_config_file, "w", encoding="utf-8") as f:
            json.dump({"users": []}, f, ensure_ascii=False, indent=4)
    if not global_history_file.exists():
        with open(global_history_file, "w", encoding="utf-8") as f:
            json.dump({"users": [], "blocks": []}, f, ensure_ascii=False, indent=4)
    if not up112_file.exists():
        with open(up112_file, "w", encoding="utf-8") as f:
            json.dump({"nums": []}, f, ensure_ascii=False, indent=2)
    load_export_settings()
    return True

def load_global_config():
    if global_config_file.exists():
        try:
            with open(global_config_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"users": []}
    return {"users": []}

def save_global_config(config):
    try:
        with open(global_config_file, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
        return True
    except Exception:
        return False

def load_export_settings():
    global export_settings
    try:
        if export_settings_file.exists():
            with open(export_settings_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    export_settings.update(data)
    except Exception:
        pass

def save_export_settings():
    try:
        with open(export_settings_file, "w", encoding="utf-8") as f:
            json.dump(export_settings, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

def create_user_backup():
    if not current_user or not user_data_file or not user_backup_dir: return
    user_backup_dir.mkdir(exist_ok=True)
    backup_file = user_backup_dir / "user_backup_latest.json"
    try:
        if user_data_file.exists(): shutil.copy2(user_data_file, backup_file)
    except Exception: pass

def create_global_backup():
    if not global_history_file or not global_backup_dir: return
    global_backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = global_backup_dir / f"global_backup_{timestamp}.json"
    try:
        if global_history_file.exists(): shutil.copy2(global_history_file, backup_file)
    except Exception: pass

def update_serp_list():
    try:
        serp_data = {"blocks": []}
        if global_history_file.exists():
            with open(global_history_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                serp_data["blocks"] = data.get("blocks", [])[:1000]
        with open(serp_list_file, "w", encoding="utf-8") as f:
            json.dump(serp_data, f, ensure_ascii=False, indent=2)
    except Exception: pass

def load_counters():
    global user_counters, user_monthly_archives
    try:
        if counter_file.exists():
            with open(counter_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                user_counters = data.get('current_counters', {})
                user_monthly_archives = data.get('archives', {})
        else:
            user_counters = {}
            user_monthly_archives = {}
    except Exception:
        user_counters = {}
        user_monthly_archives = {}

def save_counters():
    try:
        data = {'current_counters': user_counters, 'archives': user_monthly_archives}
        with open(counter_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception: pass

def archive_monthly_counters():
    global user_counters, user_monthly_archives
    current_month_year = datetime.now().strftime("%Y-%m")
    for user in list(user_counters.keys()):
        if user not in user_monthly_archives: user_monthly_archives[user] = {}
        for month in list(user_counters[user].keys()):
            if month != current_month_year:
                if month not in user_monthly_archives[user]: user_monthly_archives[user][month] = 0
                user_monthly_archives[user][month] += user_counters[user][month]
                del user_counters[user][month]
        if not user_counters[user]: del user_counters[user]
    save_counters()

def _ym_from_date_str(date_str: str) -> str | None:
    try:
        return datetime.strptime(date_str, "%d.%m.%Y").strftime("%Y-%m")
    except Exception:
        return None

def _adjust_user_counter(user: str, ym: str | None, delta: int):
    global user_counters, user_monthly_archives
    if not user or not ym:
        return
    changed = False
    if user in user_counters and ym in user_counters[user]:
        user_counters[user][ym] = max(0, user_counters[user][ym] + delta)
        if user_counters[user][ym] == 0:
            del user_counters[user][ym]
            if not user_counters[user]:
                del user_counters[user]
        changed = True
    elif user in user_monthly_archives and ym in user_monthly_archives[user]:
        user_monthly_archives[user][ym] = max(0, user_monthly_archives[user][ym] + delta)
        if user_monthly_archives[user][ym] == 0:
            del user_monthly_archives[user][ym]
            if not user_monthly_archives[user]:
                del user_monthly_archives[user]
        changed = True
    if changed:
        save_counters()
        update_status()

def update_user_counter():
    if current_user:
        current_month_year = datetime.now().strftime("%Y-%m")
        if current_user not in user_counters: user_counters[current_user] = {}
        if current_month_year not in user_counters[current_user]: user_counters[current_user][current_month_year] = 0
        user_counters[current_user][current_month_year] += 1
        save_counters()
        update_status()

def load_global_data():
    global global_blocks_cache, global_data_loaded
    if not global_history_file or not global_history_file.exists(): return {"users": [], "blocks": []}
    try:
        with open(global_history_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            global_blocks_cache = data.get("blocks", [])
            global_data_loaded = True
            return data
    except Exception:
        return {"users": [], "blocks": []}

def load_full_count():
        """Загрузить/инициализировать full_count.json"""
        global full_counts, full_serial_owner
        try:
            if full_count_file.exists():
                with open(full_count_file, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                    full_counts = dict(data.get("counts", {}))
                    full_serial_owner = dict(data.get("serial_owner", {}))
            else:
                full_counts = {}
                full_serial_owner = {}
                save_full_count()
        except Exception:
            full_counts = {}
            full_serial_owner = {}

def save_full_count():
        try:
            full_count_file.parent.mkdir(parents=True, exist_ok=True)
            with open(full_count_file, "w", encoding="utf-8") as f:
                json.dump({"counts": full_counts, "serial_owner": full_serial_owner}, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

def _fc_inc(user: str):
        if not user:
            return
        full_counts[user] = int(full_counts.get(user, 0) or 0) + 1

def _fc_dec(user: str):
        if not user:
            return
        cur = int(full_counts.get(user, 0) or 0)
        if cur <= 1:
            full_counts.pop(user, None)
        else:
            full_counts[user] = cur - 1

def fc_mark(user: str, serial: str):
        """
        Посчитать серийник за пользователя, если ещё не считали.
        Защита от повторов: один serial считается ровно один раз.
        """
        if not user or not serial:
            return
        if serial in full_serial_owner:
            return  # уже учтён кем-то
        full_serial_owner[serial] = user
        _fc_inc(user)
        save_full_count()

def fc_unmark(serial: str):
        """
        Убрать серийник из учёта (для удаления из списка или глобальной истории).
        """
        if not serial:
            return
        user = full_serial_owner.pop(serial, None)
        if user:
            _fc_dec(user)
            save_full_count()


def _dedup_blocks(blocks):
    seen = set()
    result = []
    for block in blocks:
        serial = block.get("full_serial")
        if serial and serial not in seen:
            seen.add(serial)
            result.append(block)
    return result

def load_up112_nums():
    try:
        if up112_file.exists():
            with open(up112_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return list(data.get("nums", []))
    except Exception:
        pass
    return []

def save_up112_nums(nums: list[str]):
    try:
        with open(up112_file, "w", encoding="utf-8") as f:
            json.dump({"nums": nums}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def serial_to_xxxxy(serial: str) -> str:
    if not serial or len(serial) < 9:
        return ""
    block4 = serial[5:9]
    factory = serial[4]
    return f"{block4}{factory}"

def copy_serial(event=None):
    global current_serial, copied_flag, last_copied_serial
    raw_number = entry.get().strip().replace(" ", "")
    if str(copy_btn['state']) == 'disabled' or not _can_copy_state(entry.get()):
        play_sound("error")
        return
    if len(raw_number) > 5:
        serial_label.config(text="НЕКОРРЕКТНЫЙ КОД")
        play_sound("error")
        messagebox.showerror("Ошибка", "Невозможно скопировать: код должен содержать не более 5 цифр.")
        return
    if not current_serial: return
    serial2, _, factory_code2, block4_ = generate_serial(entry.get().strip())
    block_num2 = raw_number[:-1]
    is_threshold = check_block_threshold(factory_code2, block_num2) if serial2 else False
    in_archive = bool(get_global_block_info(current_serial))
    in_prelist = (current_serial in checked_serials_set) and not in_archive and not is_threshold
    if in_archive or is_threshold or in_prelist:
        play_sound("error")
        if in_archive:
            messagebox.showwarning("Запрещено", " ССБ 112 с таким номером уже проверен и записан в архив. ")
        elif is_threshold:
            messagebox.showwarning("Запрещено",
                                   "ССБ 112 с таким номером уже проверен")
        else:
            messagebox.showwarning("Предварительный список",
                                   "Этот номер уже записан в текущем списке. Копирование отключено.")
        return
    pyperclip.copy(current_serial)
    copied_flag = True
    last_copied_serial = current_serial
    play_sound("copy")
    copy_btn.config(text="✅ СКОПИРОВАНО", bootstyle="success")

def add_block(event=None):
    global current_serial, copied_flag, last_checked_raw, last_copied_serial

    # NEW: «Записать» вообще не работает, когда «Копировать» недоступна
    if str(copy_btn['state']) == 'disabled' or not _can_copy_state(entry.get()):
        play_sound("error")
        messagebox.showwarning(
            "Недоступно",
            "Кнопка «Копировать» неактивна: ввод некорректен/дубликат/порог. Запись запрещена."
        )
        return

    if not check_user_exists():
        play_sound("error")
        messagebox.showerror("Ошибка", "Сначала введите имя пользователя!")
        user_entry.focus()
        return

    if not check_global_data_loaded():
        return

    raw_number = entry.get().strip()
    raw_number_no_spaces = raw_number.replace(" ", "")
    if len(raw_number_no_spaces) > 5:
        serial_label.config(text="НЕКОРРЕКТНЫЙ КОД")
        play_sound("error")
        messagebox.showerror("Ошибка", "Невозможно добавить блок: код должен содержать не более 5 цифр.")
        return

    if not current_serial:
        return

    if current_serial in checked_serials_set:
        play_sound("error")
        messagebox.showwarning("Предварительный список",
                               "Этот номер уже записан в текущем списке. Копирование отключено.")
        return

    serial, factory_name, factory_code, block4 = generate_serial(raw_number)
    if not serial:
        play_sound("error")
        messagebox.showerror("Ошибка ввода", "Введите минимум 2 цифры (пример: 1225 или 48 6).")
        return

    if factory_name == "НЕИЗВЕСТНЫЙ ЗАВОД":
        play_sound("error")
        messagebox.showerror("Ошибка", "Неизвестный код завода. Блок не может быть добавлен.")
        return

    global_block = get_global_block_info(serial)
    if global_block:
        play_sound("error")
        messagebox.showerror("Ошибка", f"Блок {serial} уже проверен пользователем {global_block.get('user', 'неизвестным')} {global_block.get('date', 'в неизвестную дату')}!")
        return

    block_num = raw_number_no_spaces[:-1]
    if check_block_threshold(factory_code, block_num):
        play_sound("error")
        messagebox.showwarning(
            "Предварительный список",
            "Вы не можете записать серийный уже проверенного блока ССБ 112"
        )
        return

    if serial in checked_serials_set:
        play_sound("error")
        messagebox.showwarning("Повтор", f"Блок {serial} уже проверен в текущей сессии!")
        return

    if not copied_flag:
        if not messagebox.askyesno("Подтверждение", f"Вы не скопировали {serial}. Точно добавить?"):
            return

    # --- успех: реально записываем ---
    block_data = {
        "date": datetime.now().strftime("%d.%m.%Y"),
        "raw_input": raw_number,
        "full_serial": serial,
        "factory_code": factory_code,
        "factory_name": factory_name,
        "year": current_year,
        "user": current_user
    }
    checked_blocks.insert(0, block_data)
    checked_serials_set.add(serial)
    fc_mark(current_user, serial)
    last_checked_raw = raw_number

    update_list()
    save_user_data()
    update_user_counter()
    play_sound("add")

    # NEW: «одноразовая вставка» — сбрасываем буфер на одиночный пробел
    try:
        pyperclip.copy(" ")
    except Exception:
        pass

    entry.delete(0, tk.END)
    entry.focus()
    current_serial = ""
    copied_flag = False
    last_copied_serial = None
    serial_label.config(text="[ СЕРИЙНЫЙ НОМЕР ]", bootstyle="default")
    try:
        copy_btn.config(text="📋 Копировать", bootstyle="info", state="disabled")
        add_btn.config(state="disabled")  # кнопка тоже в неактивное состояние до следующего валидного ввода
    except Exception:
        pass


def update_list():
    for widget in list_frame.winfo_children():
        widget.destroy()
    total = len(checked_blocks)
    for idx, block in enumerate(checked_blocks):
        order_num = total - idx
        item_frame = tb.Frame(list_frame, bootstyle="default")
        item_frame.pack(fill="x", pady=2, padx=5)
        order_label = tb.Label(
            item_frame,
            text=f"{order_num})",
            width=5,
            font=("Arial", FONT_SIZES["list_item"])
        )
        order_label.pack(side="left", padx=5)
        date_label = tb.Label(item_frame, text=block["date"], width=10, font=("Arial", FONT_SIZES["list_item"]))
        date_label.pack(side="left", padx=5)
        serial_text = block["full_serial"]
        serial_label2 = tb.Label(item_frame, text=serial_text, width=15, font=("Arial", FONT_SIZES["list_item"], "bold"))
        serial_label2.pack(side="left", padx=5)
        if idx == 0:
            serial_label2.config(foreground="green")
        factory_label = tb.Label(item_frame, text=block["factory_name"], width=15, font=("Arial", FONT_SIZES["list_item"]))
        factory_label.pack(side="left", padx=5, fill="x", expand=True)
        delete_btn = tb.Button(item_frame, text="❌", bootstyle="secondary", width=3, command=lambda b=block: delete_block(b))
        delete_btn.pack(side="right", padx=5)
    update_status()
    update_input_index_label()

def update_status():
    if STATUS_MODE == "overlay": return
    blocks_count = len(checked_blocks)
    current_month_year = datetime.now().strftime("%Y-%m")
    monthly_count = 0
    total_count = 0
    if current_user:
        if current_user in user_counters and current_month_year in user_counters[current_user]:
            monthly_count = user_counters[current_user][current_month_year]
        if current_user in user_monthly_archives:
            for month, count in user_monthly_archives[current_user].items(): total_count += count
        total_count += monthly_count
    status_text = f"   {current_user if current_user else 'не задан пользователь'} - Проверено за месяц: {monthly_count}"
    status_label_base.config(text=status_text)

def update_input_index_label():
    try:
        next_idx = len(checked_blocks) + 1
    except Exception:
        next_idx = 1
    order_live_label.config(text=f"{next_idx}:")

def delete_block(block):
    serial = block["full_serial"]
    if messagebox.askyesno("Подтверждение", f"Вы точно хотите удалить блок {serial} из списка?"):
        checked_blocks[:] = [b for b in checked_blocks if b["full_serial"] != serial]
        checked_serials_set.discard(serial)
        update_list()
        save_user_data()
        messagebox.showinfo("Удалено", f"Блок {serial} удален из списка.")
        fc_unmark(serial)
        ym = _ym_from_date_str(block.get("date", ""))
        _adjust_user_counter(block.get("user") or current_user, ym, -1)

def _can_copy_state(raw: str) -> bool:
    """
    Разрешить копирование только если:
    - введено 2–5 цифр (пробелы допускаются в raw),
    - корректно парсится,
    - завод известен,
    - номер не в пороге (threshold),
    - номера нет в текущем списке (prelist),
    - номера нет в архиве (global history).
    """
    raw_clean = (raw or "").replace(" ", "")
    if not (2 <= len(raw_clean) <= 5 and raw_clean.isdigit()):
        return False

    parsed = parse_input(raw)
    if not parsed:
        return False

    factory_code, block4 = parsed
    # Неизвестный завод — копировать нельзя
    if factory_code not in FACTORIES:
        return False

    serial = build_serial(factory_code, block4)

    # Уже в архиве?
    if get_global_block_info(serial):
        return False

    # Уже в текущем предварительном списке?
    if serial in checked_serials_set:
        return False

    # Порог (threshold)?
    block_num = raw_clean[:-1]  # всё кроме последней цифры-завода
    if check_block_threshold(factory_code, block_num):
        return False

    return True



def on_entry_change(*args):
    """
    Обновляет:
    - текст серийника (всегда одним стилем),
    - глобальный current_serial,
    - состояние/надпись кнопки «Копировать»,
    - состояние кнопки «Записать номер».
    """
    global current_serial, copied_flag, last_copied_serial

    raw = entry.get().strip()
    parsed = parse_input(raw)

    if not raw:
        current_serial = ""
        serial_label.config(text="[ СЕРИЙНЫЙ НОМЕР ]", bootstyle="default")
    elif parsed:
        factory_code, block4 = parsed
        factory_name = FACTORIES.get(factory_code, "НЕИЗВЕСТНЫЙ ЗАВОД")
        current_serial = build_serial(factory_code, block4)
        serial_label.config(
            text=f"12 {current_year} {factory_code} {block4} ({factory_name})",
            bootstyle="default"
        )
    else:
        current_serial = ""
        serial_label.config(text="НЕКОРРЕКТНЫЙ КОД", bootstyle="default")

    can_copy = _can_copy_state(raw)

    if can_copy:
        if copied_flag and last_copied_serial != current_serial:
            copied_flag = False
            copy_btn.config(text="📋 Копировать", bootstyle="info")
        copy_btn.config(state="normal")
        # NEW: разрешаем «Записать номер», когда можно копировать
        add_btn.config(state="normal")
    else:
        copied_flag = False
        copy_btn.config(state="disabled", text="📋 Копировать", bootstyle="info")
        # NEW: запрещаем «Записать номер», когда копирование недоступно
        add_btn.config(state="disabled")




def show_final_list():
    if check_copied_not_added(): return
    if not checked_blocks:
        messagebox.showinfo("Список пуст", "Вы еще не добавили ни одного блока.")
        return
    sorted_blocks = sorted(checked_blocks, key=lambda b: (b["factory_name"], b["full_serial"]))
    grouped = {}
    for b in sorted_blocks: grouped.setdefault(b["factory_name"], []).append(b)
    top = Toplevel(root)
    top.title("Блокнот")
    top.geometry("800x600")
    text = tk.Text(top, font=("Consolas", 16), wrap="none")
    text.pack(fill="both", expand=True)
    for factory, blocks in grouped.items():
        text.insert("end", f"{factory}:\n", "factory")
        for idx, b in enumerate(blocks, start=1): text.insert("end", f"   {idx}) {b['full_serial']}   ({b['date']})\n")
        text.insert("end", "\n")
    text.tag_configure("factory", font=("Arial", 18, "bold"))
    remember_btn = tb.Button(top, text="✅ Запомнил", bootstyle="success", command=lambda: confirm_remember(top), width=12, padding=8)
    remember_btn.place(relx=0.675, rely=0.95, anchor="sw")
    center_and_focus(top, modal=False)
    global final_list_window
    final_list_window = top

def confirm_remember(top_window):
    global global_blocks_cache, final_list_window
    if not checked_blocks:
        top_window.destroy(); final_list_window = None; return
    if not messagebox.askyesno("Подтверждение", "Вы точно запомнили номера проверенных блоков?\nСледующий список будет включать текущие блоки.\n\nОчистить список проверенных блоков?"):
        messagebox.showinfo("Информация", "Список оставлен без изменений.")
        top_window.destroy(); final_list_window = None; return
    global_data = load_global_data()
    blocks = global_data.get("blocks", [])
    existing = {b.get("full_serial") for b in blocks if b.get("full_serial")}
    added = 0
    new_up112_tokens = []
    for b in checked_blocks:
        s = b.get("full_serial")
        if s and s not in existing:
            blocks.insert(0, b); global_blocks_cache.insert(0, b)
            existing.add(s); added += 1
            token = serial_to_xxxxy(s)
            if token:
                new_up112_tokens.append(token)
    global_data["blocks"] = _dedup_blocks(blocks)
    try:
        with open(global_history_file, "w", encoding="utf-8") as f:
            json.dump(global_data, f, ensure_ascii=False, indent=4)
        create_global_backup(); update_serp_list()
        if new_up112_tokens:
            current_list = load_up112_nums()
            seen = set(current_list)
            for t in new_up112_tokens:
                if t not in seen:
                    current_list.append(t)
                    seen.add(t)
            save_up112_nums(current_list)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить глобальные данные: {e}")
        return
    archive_monthly_counters()
    checked_blocks.clear(); checked_serials_set.clear()
    update_list(); save_user_data()
    messagebox.showinfo("Успех", f"Добавлено {added} новых блоков.")
    top_window.destroy(); final_list_window = None

def save_user_data():
    if not current_user: return
    if user_data_file:
        user_data = {"user": current_user, "blocks": checked_blocks, "last_save": datetime.now().strftime("%d.%m.%Y %H:%M:%S")}
        try:
            with open(user_data_file, "w", encoding="utf-8") as f:
                json.dump(user_data, f, ensure_ascii=False, indent=4)
            create_user_backup()
        except Exception:
            pass

def save_all_data():
    if check_copied_not_added(): return
    save_user_data()
    messagebox.showinfo("Сохранено", "Данные успешно сохранены.")

def save_button_handler(event=None): save_all_data()

def load_user_data():
    global checked_blocks, checked_serials_set, current_user
    if not user_data_file: return
    try:
        if user_data_file.exists():
            with open(user_data_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                current_user = data.get("user", "")
                checked_blocks = data.get("blocks", [])
                checked_serials_set = {block["full_serial"] for block in checked_blocks}
                if current_user:
                    if user_entry and user_entry.winfo_exists(): user_entry.pack_forget()
                    if set_user_btn and set_user_btn.winfo_exists(): set_user_btn.pack_forget()
                    if user_label and user_label.winfo_exists():
                        user_label.config(text=f"{current_user}")
                        user_label.pack(side="left", padx=5)
                    if save_btn and save_btn.winfo_exists(): save_btn.pack(side="left", padx=5)
                    if settings_btn and settings_btn.winfo_exists(): settings_btn.pack(side="left", padx=5)
                    update_list()
    except Exception:
        pass

def update_year(event=None):
    global current_year
    current_year = year_var.get()
    on_entry_change()

def set_user(event=None):
    global current_user, user_data_file, user_backup_dir, save_btn
    new_user = user_entry.get().strip()
    if not new_user:
        play_sound("error")
        messagebox.showerror("Ошибка", "Введите имя пользователя!")
        return
    if " " in new_user:
        play_sound("error")
        messagebox.showerror("Ошибка", "Имя пользователя не должно содержать пробелов!")
        return
    global_config = load_global_config()
    if new_user in global_config.get("users", []):
        play_sound("error")
        messagebox.showerror("Ошибка", f"Пользователь '{new_user}' уже занят!")
        return
    if "users" not in global_config: global_config["users"] = []
    if new_user not in global_config["users"]:
        global_config["users"].append(new_user)
        if not save_global_config(global_config): return
    current_user = new_user
    user_data_file = Path("user_data.json")
    user_backup_dir = Path("backup")
    user_backup_dir.mkdir(exist_ok=True)
    user_entry.pack_forget()
    set_user_btn.pack_forget()
    user_label.config(text=f"{current_user}")
    user_label.pack(side="left", padx=5)
    save_btn.pack(side="left", padx=5)
    settings_btn.pack(side="left", padx=5)
    update_status()
    save_user_data()

def clear_list():
    if not checked_blocks:
        messagebox.showinfo("Список пуст", "Список и так пуст.")
        return
    result = messagebox.askyesno("Подтверждение", "Вы точно хотите очистить список проверенных блоков?")
    if result:
        archive_monthly_counters()
        checked_blocks.clear()
        checked_serials_set.clear()
        update_list()
        save_user_data()
        messagebox.showinfo("Успех", "Список очищен. Счетчики сохранены в архиве.")

def _valid_mac_prefix(prefix: str) -> bool:
    return re.fullmatch(r"(?i)([0-9a-f]{2}:){4}[0-9a-f]{2}", prefix.strip()) is not None

def _mac_from_serial(serial: str) -> str:
    prefix = export_settings.get("mac_prefix", "B8:17:22:ED:00")
    if not _valid_mac_prefix(prefix):
        prefix = "B8:17:22:ED:00"
    last4 = (serial[-4:] if serial else "0000").rjust(4, "0")
    b5 = last4[:2]
    b6 = last4[2:]
    return f"{prefix}:{b5}:{b6}".upper()

def open_export_settings():
    win = Toplevel(root)
    win.title("Настройки экспорта")
    win.geometry("820x550")
    pad = 10
    title = tb.Label(win, text="Параметры экспорта в Excel", font=("Arial", 18, "bold"))
    title.pack(pady=(14, 6))
    form = tb.Frame(win)
    form.pack(fill="both", expand=True, padx=pad, pady=pad)
    tb.Label(form, text="Версия прошивки:", font=("Arial", 14)).grid(row=0, column=0, sticky="w", padx=6, pady=8)
    fw_var = tk.StringVar(value=export_settings.get("firmware", "v. 1.6"))
    fw_entry = tb.Entry(form, textvariable=fw_var, width=28, font=("Arial", 14))
    fw_entry.grid(row=0, column=1, sticky="w", padx=6, pady=8)
    tb.Label(form, text="Цвет блока:", font=("Arial", 14)).grid(row=1, column=0, sticky="w", padx=6, pady=8)
    color_var = tk.StringVar(value=export_settings.get("block_color", "серый"))
    color_combo = tb.Combobox(form, textvariable=color_var, values=["серый", "чёрный", "синий", "зелёный", "красный"], width=26, font=("Arial", 14))
    color_combo.grid(row=1, column=1, sticky="w", padx=6, pady=8)
    tb.Label(form, text="MAC-префикс (5 байт):", font=("Arial", 14)).grid(row=2, column=0, sticky="w", padx=6, pady=8)
    mac_var = tk.StringVar(value=export_settings.get("mac_prefix", "B8:17:22:ED:00"))
    mac_entry = tb.Entry(form, textvariable=mac_var, width=28, font=("Arial", 14))
    mac_entry.grid(row=2, column=1, sticky="w", padx=6, pady=8)
    hint = tb.Label(form,
                    text="Формат: AA:BB:CC:DD:EE (последние 4 цифры серийника формируют XX:YY)",
                    bootstyle="secondary", font=("Arial", 11))
    hint.grid(row=3, column=1, sticky="w", padx=6, pady=(0,8))
    tb.Label(form, text="Примечания:", font=("Arial", 14)).grid(row=4, column=0, sticky="nw", padx=6, pady=8)
    notes_text = tk.Text(form, width=44, height=6, font=("Arial", 13))
    notes_text.grid(row=4, column=1, sticky="w", padx=6, pady=8)
    notes_text.insert("1.0", export_settings.get("notes", "проверка на новом стенде"))
    btns = tb.Frame(win); btns.pack(pady=(0, 12))
    def save_and_close():
        prefix = mac_var.get().strip()
        if not _valid_mac_prefix(prefix):
            messagebox.showerror("Ошибка", "MAC-префикс должен быть в формате AA:BB:CC:DD:EE")
            return
        export_settings["firmware"] = fw_var.get().strip() or "v. 1.6"
        export_settings["block_color"] = color_var.get().strip() or "серый"
        export_settings["notes"] = notes_text.get("1.0", "end").strip() or "проверка на новом стенде"
        export_settings["mac_prefix"] = prefix.upper()
        save_export_settings()
        messagebox.showinfo("Сохранено", "Настройки экспорта сохранены.")
        win.destroy()
    tb.Button(btns, text="💾 Сохранить", bootstyle="success", command=save_and_close, width=16, padding=8).pack(side="left", padx=6)
    tb.Button(btns, text="Закрыть", bootstyle="secondary", command=win.destroy, width=12, padding=8).pack(side="left", padx=6)
    center_and_focus(win, modal=True)

def _build_and_save_excel(blocks: list[dict]):
    """
    Стационарная книга C:/ssb_data/PNB_учёт.xlsx:
    - Лист 1: 'Основной учёт' (есть колонка 'ЗАВОД' сразу после 'Серийный номер')
    - Листы 2..7: по одному на каждый завод (без колонки 'ЗАВОД'), сортировка по 'Серийный номер'
    Оформление и защита применяются ко всем листам.
    """
    def row_from_block(b: dict, with_factory: bool) -> dict:
        serial = b.get("full_serial", "")
        row = {
            "Дата проверки": b.get("date", ""),
            "Зав. № (бирка, flash)": serial,
        }
        if with_factory:
            row["Завод"] = b.get("factory_name", "")
        row.update({
            "MAC": _mac_from_serial(serial),
            "Проверяющий": b.get("user", ""),
            "Версия прошивки ССФ1121": export_settings.get("firmware", "v. 1.6"),
            "Цвет блока": export_settings.get("block_color", "серый"),
            "Примечания": export_settings.get("notes", "проверка на новом стенде"),
        })
        return row

    # общий лист
    main_rows = [row_from_block(b, with_factory=True) for b in blocks]
    main_columns = ["Дата проверки", "Зав. № (бирка, flash)", "Завод", "MAC", "Проверяющий", "Версия прошивки ССФ1121", "Цвет блока", "Примечания"]

    # по заводам
    factory_names = list(FACTORIES.values())  # ["ВЕКТОР", "ИНТЕГРАЛ", ...]
    per_factory = {
        name: [row_from_block(b, with_factory=False) for b in blocks if b.get("factory_name") == name]
        for name in factory_names
    }
    for name in factory_names:
        per_factory[name].sort(key=lambda r: str(r.get("Зав. № (бирка, flash)", "")))  # по порядку

    # запись
    EXCEL_BOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(EXCEL_BOOK_PATH, engine="openpyxl") as writer:
        # 1) основной учёт
        pd.DataFrame(main_rows, columns=main_columns).to_excel(
            writer, index=False, sheet_name="Основной учёт"
        )

        # 2..7) по заводам (без 'ЗАВОД')
        factory_columns = ["Дата проверки", "Зав. № (бирка, flash)", "MAC", "Проверяющий", "Версия прошивки ССФ1121", "Цвет блока", "Примечания"]
        for name in factory_names:
            df_f = pd.DataFrame(per_factory[name], columns=factory_columns)
            if df_f.empty:
                df_f = pd.DataFrame(columns=factory_columns)
            df_f.to_excel(writer, index=False, sheet_name=name)

        # (иногда это помогает при старых версиях openpyxl)
        try:
            writer.book.save(EXCEL_BOOK_PATH)
        except Exception:
            pass

    # пост-оформление и защита
    wb = load_workbook(EXCEL_BOOK_PATH)
    for ws in wb.worksheets:
        _format_ws(ws)
        _protect_ws(ws)

    # (опционально) запретить изменение структуры книги
    try:
        wb.security.lockStructure = True
    except Exception:
        pass

    wb.save(EXCEL_BOOK_PATH)

def _format_ws(ws):
    """
    Оформление листа:
    - Times New Roman 14
    - Центрирование по горизонтали/вертикали
    - Без переноса строк (wrap_text=False) — всё в одну строку
    - Авто-ширина с повышенным коэффициентом + минимальные ширины для важных столбцов
    """
    header_font = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=True)
    cell_font   = Font(name=EXCEL_FONT_NAME, size=EXCEL_FONT_SIZE, bold=False)
    center_no_wrap = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # 1) шапка
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_no_wrap

    # 2) данные
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.alignment = center_no_wrap

    # 3) авто-ширина: посчитаем максимальную длину текста (включая заголовок),
    #    затем увеличим её коэффициентом, добавим запас, и применим
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # заголовок этой колонки
        header_val = ws.cell(row=1, column=col_idx).value
        header_txt = str(header_val) if header_val is not None else ""
        max_len = len(header_txt)

        # макс длина по всем строкам
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            s = str(val) if val is not None else ""
            if len(s) > max_len:
                max_len = len(s)

        # «щедрая» ширина: коэффициент + запас в 2 символа
        target_width = int((max_len + 2) * EXCEL_WIDTH_FACTOR)

        # учесть минимальные требования для конкретных столбцов
        min_for_header = EXCEL_MIN_BY_HEADER.get(header_txt, EXCEL_MIN_WIDTH)
        target_width = max(min_for_header, target_width)

        # и общий предел, чтобы не улетало в бесконечность
        target_width = min(target_width, EXCEL_WIDTH_CAP)

        ws.column_dimensions[col_letter].width = target_width

def _protect_ws(ws):
    """Защита от редактирования на уровне листа (пароль EXCEL_PASSWORD)."""
    ws.protection.sheet = True
    ws.protection.enable()
    try:
        ws.protection.set_password(EXCEL_PASSWORD)
    except Exception:
        pass

def open_global_history():
    try:
        if not global_history_file or not global_history_file.exists():
            messagebox.showinfo("Файл не найден", "Глобальный файл истории не существует.")
            return
        with open(global_history_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or "blocks" not in data:
            messagebox.showerror("Ошибка", "Неверный формат данных в файле.")
            return
        blocks = data["blocks"]
        unique_serials = {block.get("full_serial", "") for block in blocks}
        top = Toplevel(root)
        top.title("Глобальная история проверок")
        top.geometry("1100x700")
        archive_info = tb.Label(top, text=f"В архиве: {len(unique_serials)} проверенных блоков", font=("Arial", 18), bootstyle="dark")
        archive_info.pack(pady=5)
        style = tb.Style()
        style.configure("Treeview", font=("Arial", FONT_SIZES["table"]))
        style.configure("Treeview.Heading", font=("Arial", FONT_SIZES["table"], "bold"))
        columns = ("date", "serial", "user")
        tree = tb.Treeview(top, columns=columns, show="headings", height=25)
        tree.heading("date", text="Дата")
        tree.heading("serial", text="Серийный номер")
        tree.heading("user", text="Пользователь")
        tree.column("date", width=140, anchor="center")
        tree.column("serial", width=220, anchor="center")
        tree.column("user", width=180, anchor="center")
        style.configure("Treeview", rowheight=40)
        for block in blocks:
            tree.insert("", "end", values=(
                block.get("date", ""),
                block.get("full_serial", ""),
                block.get("user", "")
            ))
        scrollbar = tb.Scrollbar(top, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)
        button_frame = tb.Frame(top)
        button_frame.pack(fill="x", pady=10)
        top_button_frame = tb.Frame(button_frame)
        top_button_frame.pack(pady=5)
        def delete_selected():
            selected = tree.selection()
            if not selected: return
            item = selected[0]
            values = tree.item(item, "values")
            serial = values[1]
            if messagebox.askyesno("Подтверждение", f"Удалить блок {serial} из глобальной истории?"):
                global_data = load_global_data()
                global_data["blocks"] = [b for b in global_data["blocks"] if b["full_serial"] != serial]
                try:
                    with open(global_history_file, "w", encoding="utf-8") as f:
                        json.dump(global_data, f, ensure_ascii=False, indent=4)
                    tree.delete(item)
                    update_serp_list()
                    unique_serials.discard(serial)
                    archive_info.config(text=f"В архиве: {len(unique_serials)} проверенных блоков")
                    messagebox.showinfo("Успех", f"Блок {serial} удален из глобальной истории.")
                    load_counters()
                    fc_unmark(serial)
                    ym = _ym_from_date_str(values[0])
                    user_of_row = values[2]
                    _adjust_user_counter(user_of_row, ym, -1)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось удалить блок: {str(e)}")
        tb.Button(top_button_frame, text="❌ Удалить выбранное", bootstyle="danger", command=delete_selected, width=20, padding=10).pack(side="left", padx=10)
        bottom_button_frame = tb.Frame(button_frame)
        bottom_button_frame.pack(pady=5)
        def open_excel_book():
            try:
                # формируем книгу и сохраняем её в C:\ssb_data\Таблица проверок ССБ 112.xlsx
                _build_and_save_excel(blocks)
                messagebox.showinfo("Готово", f"Файл сохранён:\n{EXCEL_FILE}")
                # открываем в Excel / системной программе
                if platform.system() == "Windows":
                    os.startfile(EXCEL_FILE)  # type: ignore[attr-defined]
                elif platform.system() == "Darwin":
                    subprocess.run(["open", str(EXCEL_FILE)])
                else:
                    subprocess.run(["xdg-open", str(EXCEL_FILE)])
            except PermissionError:
                messagebox.showerror("Ошибка",
                                     "Не удалось сохранить файл. Возможно, он открыт в Excel. Закройте файл и попробуйте снова.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сформировать Excel: {e}")

        tb.Button(
            bottom_button_frame,
            text="📖 Открыть в Excel",
            bootstyle="success",
            command=lambda: build_excel_and_open(blocks),
            width=20,
            padding=10
        ).pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")

def build_excel_and_open(blocks: list[dict]):
    """Сформировать книгу и открыть её системной программой (обычно Excel)."""
    try:
        _build_and_save_excel(blocks)
        messagebox.showinfo("Готово", f"Файл сохранён:\n{EXCEL_BOOK_PATH}")
        if platform.system() == "Windows":
            os.startfile(EXCEL_BOOK_PATH)  # type: ignore[attr-defined]
        elif platform.system() == "Darwin":
            subprocess.run(["open", str(EXCEL_BOOK_PATH)])
        else:
            subprocess.run(["xdg-open", str(EXCEL_BOOK_PATH)])
    except PermissionError:
        messagebox.showerror(
            "Ошибка",
            "Не удалось сохранить файл. Возможно, он открыт в Excel. Закройте файл и попробуйте снова."
        )
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сформировать Excel: {e}")


def minimize_window(event=None): root.iconify()

def restore_window():
    root.deiconify()
    root.focus_force()
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))

def toggle_window_size(event=None):
    global window_size_large
    if window_size_large:
        root.geometry(SMALL_WINDOW_SIZE)
        window_size_large = False
    else:
        root.geometry(WINDOW_SIZE)
        window_size_large = True
    center_and_focus(root)

def close_final_list_window():
    global final_list_window
    if final_list_window and final_list_window.winfo_exists():
        final_list_window.destroy()
        final_list_window = None

def setup_global_hotkeys():
    keyboard.add_hotkey('alt', lambda: toggle_window())
    root.bind('<Escape>', toggle_window_size)
    root.bind('<KeyPress-Shift_L>', lambda e: ping_btn.invoke())
    root.bind('<Control-a>', lambda e: show_final_list())
    root.bind('<Control-z>', lambda e: close_final_list_window())
    root.bind('<Control-s>', lambda e: save_all_data())

def toggle_window():
    if root.state() == "iconic": restore_window()
    else: minimize_window(None)

def on_space_press(event):
    global last_checked_raw, space_pressed, highlight_tag_configured
    if not entry.get().strip() and last_checked_raw:
        entry.delete(0, tk.END)
        entry.insert(0, last_checked_raw)
        if not highlight_tag_configured:
            entry.tag_configure("highlight", background="#E6E6FA")
            entry.tag_configure("digit_highlight", background="#D8BFD8")
            highlight_tag_configured = True
        entry.tag_add("highlight", "1.0", "end")
        text = last_checked_raw.replace(" ", "")
        if len(text) >= 2:
            digit_count = 0
            pos = -1
            for i, char in enumerate(last_checked_raw):
                if char.isdigit():
                    digit_count += 1
                    if digit_count == len(text) - 1:
                        pos = i
                        break
            if pos != -1: entry.tag_add("digit_highlight", f"1.{pos}", f"1.{pos + 1}")
        space_pressed = True
        return "break"
    elif space_pressed:
        entry.delete(0, tk.END)
        space_pressed = False
        return "break"
    return None

def check_copied_not_added():
    global copied_flag, last_copied_serial
    if copied_flag and last_copied_serial:
        messagebox.showwarning("Предупреждение", f"Вы скопировали но не записали номер {last_copied_serial}. Проверьте записи!")
        return True
    return False

def _find_two_preceding_digit_indexes(s: str):
    digit_positions = [i for i, ch in enumerate(s) if ch.isdigit()]
    if len(digit_positions) < 3:
        return (None, None)
    tens_idx = digit_positions[-3]
    ones_idx = digit_positions[-2]
    return (tens_idx, ones_idx)

def _adjust_two_preceding_digits(delta: int) -> bool:
    s = entry.get()
    t_idx, o_idx = _find_two_preceding_digit_indexes(s)
    if t_idx is None or o_idx is None:
        return False
    tens = int(s[t_idx])
    ones = int(s[o_idx])
    val = tens * 10 + ones
    val = (val + delta) % 100
    new_tens = val // 10
    new_ones = val % 10
    s_list = list(s)
    s_list[t_idx] = str(new_tens)
    s_list[o_idx] = str(new_ones)
    entry.delete(0, tk.END)
    entry.insert(0, "".join(s_list))
    on_entry_change()
    return True

def on_arrow_up(event):
    _adjust_two_preceding_digits(+1)
    return "break"

def on_arrow_down(event):
    _adjust_two_preceding_digits(-1)
    return "break"

def on_mousewheel(event):
    delta = event.delta
    if delta == 0:
        return
    step = +1 if delta > 0 else -1
    _adjust_two_preceding_digits(step)
    return "break"

def on_button4(event):
    _adjust_two_preceding_digits(+1)
    return "break"

def on_button5(event):
    _adjust_two_preceding_digits(-1)
    return "break"

def on_digit_entry(event):
    global space_pressed
    if space_pressed and event.char.isdigit():
        current_text = entry.get().replace(" ", "")
        if len(current_text) >= 2:
            new_text = current_text[:-2] + event.char + current_text[-1:]
            entry.delete(0, tk.END)
            entry.insert(0, new_text)
            entry.tag_remove("highlight", "1.0", "end")
            entry.tag_remove("digit_highlight", "1.0", "end")
            space_pressed = False
            on_entry_change()
        return "break"
    return None

def on_closing():
    if checked_blocks:
        messagebox.showinfo("Предупреждение", "Пожалуйста, завершите проверку, чтобы не потерять номера проверенных блоков")
        show_final_list(); return
    if check_copied_not_added(): return
    if current_user: save_user_data()
    keyboard.unhook_all()
    root.destroy()

def open_settings(event=None):
    settings_win = Toplevel(root)
    settings_win.title("Настройки сети")
    settings_win.geometry("280x500")
    tb.Label(settings_win, text="IP-адрес ССБ 112:").pack(pady=10)
    ip_var = tk.StringVar(value=ip_address)
    ip_entry = tb.Entry(settings_win, textvariable=ip_var, width=20)
    ip_entry.pack(pady=7)
    def save_settings():
        global ip_address
        ip_address = ip_var.get().strip()
        settings_win.destroy()
    tb.Button(settings_win, text="Сохранить", command=save_settings).pack(pady=20)
    center_and_focus(settings_win, modal=True)

def ping_ip():
    ip = ip_address
    if not ip:
        show_overlay_status("⚠ Введите IP", WARNING, "warning", duration=2000)
        return
    status_label_base.config(text="")
    show_overlay_status("🔍 Поиск блока ССБ 112.", INFO, btn_style=None, duration=None)
    start_ping_blink()
    threading.Thread(target=_ping_worker, args=(ip,), daemon=True).start()

def _ping_once(ip):
    is_win = platform.system().lower() == "windows"

    # 1 пакет, IPv4, таймаут ~1с
    if is_win:
        cmd = ["ping", "-n", "1", "-w", "1000", "-4", ip]
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        creation = subprocess.CREATE_NO_WINDOW
    else:
        cmd = ["ping", "-c", "1", "-W", "1", ip]
        si = None
        creation = 0

    try:
        res = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=2,
            startupinfo=si,
            creationflags=creation
        )
        out = (res.stdout or "").lower()
        # «Старое» правило: успех, если ping вернул 0 ИЛИ в тексте есть ttl=
        return (res.returncode == 0) or ("ttl=" in out)
    except Exception:
        return False



def _ping_worker(ip):
    ok = False
    for _ in range(PING_ATTEMPTS):
        ok = _ping_once(ip)
        if ok:
            break
        time.sleep(PING_GAP_SEC)
    def finish():
        stop_ping_blink("success" if ok else "danger")
        if ok:
            show_overlay_status("✅ ССБ 112 найден", SUCCESS, "success", duration=PING_RESTORE_MS)
        else:
            show_overlay_status("❌ ССБ 112 не найден", DANGER, "danger", duration=PING_RESTORE_MS)
    root.after(0, finish)

def show_overlay_status(text, label_style=None, btn_style=None, duration=None):
    global STATUS_MODE, status_restore_after_id
    STATUS_MODE = "overlay"
    if status_restore_after_id:
        root.after_cancel(status_restore_after_id); status_restore_after_id = None
    status_label_base.pack_forget()
    if label_style is None: status_label_overlay.config(text=text, bootstyle="default")
    else: status_label_overlay.config(text=text, bootstyle=label_style)
    _ensure_overlay_size()
    status_overlay_box.pack(fill="x")
    if btn_style is not None: ping_btn.config(bootstyle=btn_style)
    if duration is not None: status_restore_after_id = root.after(duration, hide_overlay_status)

def _blink_ping_btn():
    global PING_BLINK_AFTER_ID, PING_BLINKING
    if not PING_BLINKING:
        return
    state = getattr(ping_btn, "_blink_state", 0)
    state ^= 1
    ping_btn._blink_state = state
    ping_btn.config(bootstyle=("info" if state else "secondary"))
    PING_BLINK_AFTER_ID = root.after(300, _blink_ping_btn)

def start_ping_blink():
    global PING_BLINKING, PING_BLINK_AFTER_ID
    if PING_BLINKING:
        return
    PING_BLINKING = True
    ping_btn._blink_state = 0
    _blink_ping_btn()

def stop_ping_blink(final_style=None):
    global PING_BLINKING, PING_BLINK_AFTER_ID
    PING_BLINKING = False
    if PING_BLINK_AFTER_ID:
        root.after_cancel(PING_BLINK_AFTER_ID)
        PING_BLINK_AFTER_ID = None
    if final_style:
        ping_btn.config(bootstyle=final_style)

def hide_overlay_status():
    global STATUS_MODE, status_restore_after_id
    STATUS_MODE = "base"; status_restore_after_id = None
    status_overlay_box.pack_forget()
    status_label_overlay.config(text="", bootstyle="")
    stop_ping_blink()
    ping_btn.config(bootstyle="secondary")
    status_label_base.pack(fill="x")
    update_status()

def _ensure_overlay_size():
    h = max(status_label_base.winfo_height(), status_label_base.winfo_reqheight())
    if h <= 1: h = int(FONT_SIZES["status"] * 1.6)
    status_overlay_box.configure(height=h)

scaling_factor = 1.2
root = tb.Window(themename="flatly")
root.tk.call('tk', 'scaling', scaling_factor)
root.withdraw()
def _center_on_screen(win, size_str):
    w, h = map(int, size_str.split("x"))
    win.update_idletasks()
    sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
    x = (sw - w) // 2
    y = (sh - h) // 3
    win.geometry(f"{w}x{h}+{x}+{y}")
_center_on_screen(root, WINDOW_SIZE)
root.title("112 PNB")
root.iconbitmap(resource_path("l2"
                              ".ico"))
root.protocol("WM_DELETE_WINDOW", on_closing)
root.deiconify()
root.lift()
root.focus_force()
if not init_global_directories():
    messagebox.showerror("Ошибка", "Не удалось инициализировать глобальные директории. Программа будет закрыта.")
    sys.exit(1)
load_counters()
load_global_data()
global_data_loaded = True
user_data_file = Path("user_data.json")
load_full_count()
user_backup_dir = Path("backup")
top_frame = tb.Frame(root)
top_frame.pack(fill="x", padx=20, pady=10)
button_frame = tb.Frame(top_frame)
button_frame.pack(side="right")
ping_btn = tb.Button(button_frame, text="🌐", bootstyle="secondary", width=3, padding=5, command=ping_ip)
ping_btn.pack(side="left", padx=5)
folder_btn = tb.Button(button_frame, text="📊", bootstyle="primary", width=3, padding=5, command=open_global_history)
folder_btn.pack(side="left", padx=5)
ping_btn.bind("<Button-3>", open_settings)
settings_frame = tb.Frame(top_frame)
settings_frame.pack(side="left")
year_frame = tb.Frame(settings_frame)
year_frame.pack(side="left", padx=10)
tb.Label(year_frame, text="Год:", font=("Arial", FONT_SIZES["label"])).pack(side="left")
year_var = tk.StringVar(value=current_year)
year_combo = tb.Combobox(year_frame, textvariable=year_var, values=[str(i).zfill(2) for i in range(20, 41)], width=5, font=("Arial", FONT_SIZES["label"]))
year_combo.pack(side="left", padx=5)
year_combo.bind("<<ComboboxSelected>>", update_year)
year_combo.bind("<FocusOut>", update_year)
user_frame = tb.Frame(settings_frame)
user_frame.pack(side="left", padx=10)
user_label = tb.Label(user_frame, text="Пользователь:", font=("Arial", FONT_SIZES["label"]))
user_label.pack(side="left", padx=5)
user_entry = tb.Entry(user_frame, font=("Arial", FONT_SIZES["entry"]), width=15)
user_entry.pack(side="left", padx=5)
user_entry.bind("<Return>", set_user)
set_user_btn = tb.Button(user_frame, text="✓", bootstyle="secondary", width=3, command=set_user)
set_user_btn.pack(side="left", padx=5)
save_btn = tb.Button(user_frame, text="💾", bootstyle="success", width=3, padding=5, command=save_button_handler)
save_btn.pack(side="left", padx=5); save_btn.pack_forget()
settings_btn = tb.Button(user_frame, text="⚙️", bootstyle="primary", width=3, padding=5, command=open_export_settings)
settings_btn.pack(side="left", padx=5); settings_btn.pack_forget()
style = tb.Style()
style.configure("TButton", font=("Arial", FONT_SIZES["button"]))
style.configure("TLabel", font=("Arial", FONT_SIZES["label"]))
style.configure("Big.TLabelframe.Label", font=("Arial", 40, "bold"))
entry_var = tk.StringVar()
entry_var.trace_add("write", on_entry_change)
vcmd = (root.register(only_digits_spaces), "%P")
tb.Label(root, text="Введите номер блока ССБ 112 : ").pack(pady=5)
input_row = tb.Frame(root)
input_row.pack(pady=6)
order_live_label = tb.Label(input_row, text="1:", font=("Arial", FONT_SIZES["label"]))
order_live_label.pack(side="left", padx=(0, 8))
entry = tb.Entry(input_row, textvariable=entry_var, font=("Consolas", FONT_SIZES["entry"]), justify="center", width=ENTRY_WIDTH, validate="key", validatecommand=vcmd)
entry.pack(side="left"); entry.focus()
update_input_index_label()
entry.bind('<space>', on_space_press)
entry.bind('<KeyPress>', on_digit_entry)
entry.bind('<Up>', on_arrow_up)
entry.bind('<Down>', on_arrow_down)
entry.bind("<MouseWheel>", on_mousewheel)
entry.bind("<Button-4>", on_button4)
entry.bind("<Button-5>", on_button5)
serial_label = tb.Label(root, text="[ СЕРИЙНЫЙ НОМЕР ]", font=("Arial", FONT_SIZES["serial"], "bold"))
serial_label.pack(pady=12)
copy_btn = tb.Button(
    root,
    text="📋 Копировать",
    bootstyle="info",
    command=copy_serial,
    state="disabled",              # ← добавь это
    width=BUTTON_SIZES["width"],
    padding=BUTTON_SIZES["padding"]
)
copy_btn.pack(pady=10)

btn_frame = tb.Frame(root)
btn_frame.pack(pady=15)
add_btn = tb.Button(
    btn_frame,
    text="🖊️ Записать номер",
    bootstyle="success",
    command=add_block,
    state="disabled",  # NEW: стартуем выключенной — включит on_entry_change()
    width=BUTTON_SIZES["width"],
    padding=BUTTON_SIZES["padding"]
)

add_btn.grid(row=0, column=0, padx=15, pady=10)
final_btn = tb.Button(btn_frame, text="📊 Завершить проверку", bootstyle="primary", command=show_final_list, width=BUTTON_SIZES["width"], padding=BUTTON_SIZES["padding"])
final_btn.grid(row=0, column=1, padx=15, pady=10)
lf_title = tb.Label(root, text="Проверенные блоки", font=("Arial", 16))
list_frame_container = tb.Labelframe(root,
                                     labelwidget=lf_title,
                                     padding=10,
                                     bootstyle="primary")
list_frame_container.pack(fill="both", expand=True, padx=20, pady=16)
canvas = tk.Canvas(list_frame_container)
scrollbar = tb.Scrollbar(list_frame_container, orient="vertical", command=canvas.yview)
scrollable_frame = tb.Frame(canvas)
scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")
list_frame = scrollable_frame
APP_BG = root.cget("bg")
status_area = tk.Frame(root, bg=APP_BG, bd=0, highlightthickness=0)
status_area.pack(fill="x", pady=5)
status_label_base = tb.Label(status_area, text="Пользователь не задан", font=("Arial", FONT_SIZES["status"]))
status_label_base.pack(fill="x")
status_overlay_box = tk.Frame(status_area, bg=APP_BG, bd=0, highlightthickness=0, relief="flat")
status_overlay_box.pack_forget()
status_overlay_box.pack_propagate(False)
status_label_overlay = tb.Label(status_overlay_box, text="", font=("Arial", FONT_SIZES["status"]))
status_label_overlay.pack()
def minimize_window(event=None): root.iconify()
def restore_window():
    root.deiconify()
    root.focus_force()
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))
center_and_focus(root, modal=False)
setup_global_hotkeys()
load_user_data()
root.bind("<Return>", copy_serial)
root.bind("<Control-Return>", add_block)
root.mainloop()